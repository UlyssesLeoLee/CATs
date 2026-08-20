# -*- coding: utf-8 -*-
"""CATs Rust 技术选型书 -> Excel 工作簿生成
工作表: 封面 / 选型总览 / 工具链 / 异步与并发 / Web / 数据库 / 消息 / 序列化 /
       可观测性 / 安全 / 时间 / 媒体 / 依赖版本锁定 / ADR 决策记录 / 风险
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Microsoft YaHei"
NAVY   = "1F3864"
HEADER = "2E5496"
SUB    = "8EAADB"
LIGHT  = "D9E1F2"
ZEBRA  = "F2F5FB"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True):   return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def topleft():         return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def title_block(ws, title, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.fill = fill(NAVY); c.font = f(13, True, WHITE); c.alignment = center(False)
    ws.row_dimensions[1].height = 28

def style_header(ws, row, ncols, fillc=HEADER):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(fillc); cell.font = f(10, True, WHITE)
        cell.alignment = center(); cell.border = BORDER

def grid(ws, r1, r2, c1, c2, zebra=True):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.vertical is None:
                cell.alignment = topleft()
            if zebra and (r - r1) % 2 == 1:
                if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = fill(ZEBRA)

def widths(ws, mp):
    for col, w in mp.items():
        ws.column_dimensions[col].width = w

# ============================================================
wb = Workbook()

# ---------------- 封面 ----------------
ws = wb.active
ws.title = "封面"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 26, "C": 56, "D": 3})
ws.merge_cells("B2:C2")
ws["B2"] = "Rust 技术选型书（ADR 深度版）"
ws["B2"].font = f(22, True, NAVY); ws["B2"].alignment = center(False)
ws.row_dimensions[2].height = 44
ws.merge_cells("B3:C3")
ws["B3"] = "CATs — 全媒体 AI 辅助翻译 SaaS 平台"
ws["B3"].font = f(12, True, "404040"); ws["B3"].alignment = center(False)
ws.row_dimensions[3].height = 24

meta = [
    ("文档编号", "CATs-TS-RUST-001"),
    ("文档名", "Rust 技术选型书（含 ADR 决策记录，Rust 生态深度）"),
    ("版本", "第 1.0 版（草稿）"),
    ("创建日", "2026-08-19"),
    ("作者", "架构师"),
    ("状态", "评审前草稿"),
    ("密级", "仅社内"),
    ("适用标准", "Rust API Guidelines / Rust Standard Library / RFC 标准化"),
    ("上游文档", "CATs 技术选型书 v2.0（横向补充：本书专门深化 Rust 生态）"),
]
r = 5
for k, v in meta:
    ws.cell(r, 2, k).fill = fill(SUB); ws.cell(r, 2).font = f(10, True, WHITE)
    ws.cell(r, 2).alignment = center()
    ws.cell(r, 3, v).alignment = left(); ws.cell(r, 3).font = f(10)
    ws.cell(r, 2).border = BORDER; ws.cell(r, 3).border = BORDER
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
ws.cell(r, 2, "v1.0 主要内容（10 大类）").font = f(10, True, NAVY)
r += 1
items = [
    "① 工具链（rustc/cargo/clippy/rustfmt/cargo-audit/d 等）",
    "② 异步运行时（Tokio）+ 并发原语（parking_lot / dashmap / arc-swap）",
    "③ Web 框架（Axum / Tauri）+ gRPC（tonic）+ HTTP 客户端（reqwest）",
    "④ 数据库（sqlx）+ 迁移（sqlx-migrate）+ 连接池（deadpool）",
    "⑤ 消息（rdkafka）+ 缓存（redis-rs / deadpool-redis）",
    "⑥ 序列化（serde / prost）+ 数据验证（validator）",
    "⑦ 可观测性（tracing / OpenTelemetry / metrics）",
    "⑧ 安全（jsonwebtoken / argon2 / ring / rustls / keyring）",
    "⑨ 测试（cargo-nextest / mockall / testcontainers-rs / proptest / criterion）",
    "⑩ 媒体处理（ffmpeg-next / image）+ 14 条 ADR 决策记录",
]
for c_text in items:
    ws.cell(r, 2, c_text).font = f(9); ws.cell(r, 2).alignment = topleft()
    ws.cell(r, 2).border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 18
    r += 1

# ---------------- 选型总览 ----------------
ws = wb.create_sheet("选型总览")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 16, "B": 22, "C": 14, "D": 30, "E": 30, "F": 14})
title_block(ws, "Rust 选型总览（必选）", 6)
headers = ["类别", "选型", "版本", "用途", "替代方案（不采用）", "优先级"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 6)
must = [
    ("异步运行时", "tokio", "1.40+", "异步运行时", "async-std / smol / monoio", "P0"),
    ("Web 服务", "axum", "0.7+", "HTTP 服务", "actix-web / warp / rocket", "P0"),
    ("Web 客户端", "tauri", "2.1+", "桌面客户端", "—", "P0"),
    ("HTTP 客户端", "reqwest", "0.12+", "HTTP 客户端", "ureq / isahc", "P0"),
    ("gRPC", "tonic", "0.12+", "gRPC 客户端/服务", "grpc-rs", "P0"),
    ("WebSocket", "tokio-tungstenite", "0.24+", "WS 客户端/服务", "async-tungstenite", "P0"),
    ("数据库", "sqlx", "0.8+", "PG 驱动", "diesel / sea-orm", "P0"),
    ("迁移", "sqlx-migrate", "0.8+", "Schema 迁移", "refinery", "P0"),
    ("Kafka", "rdkafka", "0.36+", "Kafka 客户端", "kafka-rust / rskafka", "P0"),
    ("缓存", "redis", "0.27+", "Redis/Valkey 客户端", "—", "P0"),
    ("缓存池", "deadpool-redis", "0.18+", "Redis 池", "bb8-redis", "P0"),
    ("S3", "aws-sdk-s3", "1.55+", "S3 客户端（MinIO 兼容）", "s3", "P0"),
    ("序列化", "serde", "1.0+", "序列化框架", "—", "P0"),
    ("JSON", "serde_json", "1.0+", "JSON", "simd-json", "P0"),
    ("Protobuf", "prost", "0.13+", "Protobuf 运行时", "—", "P0"),
    ("验证", "validator", "0.18+", "数据验证", "garde", "P0"),
    ("错误（库）", "thiserror", "1.0+", "结构化错误", "snafu", "P0"),
    ("错误（应用）", "anyhow", "1.0+", "简化错误传播", "eyre", "P0"),
    ("日志", "tracing", "0.1+", "结构化日志", "log", "P0"),
    ("追踪", "opentelemetry", "0.27+", "OTel API+SDK", "—", "P0"),
    ("指标", "metrics", "0.23+", "抽象指标", "prometheus", "P0"),
    ("JWT", "jsonwebtoken", "9.3+", "JWT 签发/校验", "—", "P0"),
    ("密码哈希", "argon2", "0.5+", "密码哈希", "bcrypt", "P0"),
    ("加密", "ring", "0.17+", "加密原语", "aws-lc-rs / openssl", "P0"),
    ("TLS", "rustls", "0.23+", "TLS 实现", "openssl", "P0"),
    ("密钥", "keyring", "3.6+", "OS 密钥库", "—", "P0"),
    ("敏感值", "secrecy", "0.8+", "敏感值包装", "—", "P0"),
    ("内存清零", "zeroize", "1.8+", "内存清零", "—", "P0"),
    ("时间", "time", "0.3+", "时间处理", "chrono", "P0"),
    ("Cron", "cron", "0.12+", "Cron 表达式", "—", "P0"),
    ("配置", "figment", "0.10+", "多源配置", "config / dotenv", "P0"),
    ("并行", "rayon", "1.10+", "CPU 并行", "—", "P1"),
    ("HashMap", "dashmap", "6.1+", "高并发 Map", "—", "P1"),
    ("锁", "parking_lot", "0.12+", "同步锁", "—", "P1"),
    ("K8s", "kube", "0.95+", "K8s 客户端", "—", "P1"),
    ("FFmpeg", "ffmpeg-next", "7.1+", "FFmpeg 绑定", "ffmpeg-sidecar", "P1"),
    ("图像", "image", "0.25+", "图像处理", "—", "P1"),
    ("测试", "cargo-nextest", "0.9+", "并行测试", "cargo test", "P0"),
    ("Mock", "mockall", "0.13+", "通用 mock", "—", "P0"),
    ("容器测试", "testcontainers", "0.20+", "真实容器", "—", "P0"),
    ("属性测试", "proptest", "1.5+", "属性测试", "—", "P1"),
    ("基准", "criterion", "0.5+", "性能基准", "—", "P1"),
    ("火焰图", "cargo-flamegraph", "0.6+", "火焰图", "—", "P1"),
    ("审计", "cargo-audit", "0.21+", "漏洞扫描", "—", "P0"),
    ("许可证", "cargo-deny", "1.16+", "许可证/重复", "—", "P0"),
]
r = 4
for row in must:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 6)
ws.freeze_panes = "A4"

# ---------------- ADR 决策记录 ----------------
ws = wb.create_sheet("ADR决策记录")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 10, "B": 30, "C": 20, "D": 50, "E": 20})
title_block(ws, "ADR 决策记录（Rust 选型）", 5)
headers = ["编号", "主题", "决策", "理由", "取舍"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
adrs = [
    ("ADR-R-01", "锁定 Rust 1.75+ MSRV", "1.75", "覆盖 2024 主流特性 + 1-2 年升级缓冲", "无法用 1.80+ 少数新特性（影响小）"),
    ("ADR-R-02", "默认禁止 unsafe", "强制 SAFETY 注释 + 双签", "Rust 内存安全是核心价值", "不可避免的 unsafe（Vec/FFI）需集中封装"),
    ("ADR-R-03", "依赖版本锁定策略", "Cargo.toml caret + 锁 Cargo.lock", "平衡灵活性与稳定性", "major 升级需 ADR"),
    ("ADR-R-04", "选 Tokio 异步运行时", "tokio", "生态最完整（axum/tonic/sqlx 全兼容）", "io_uring 利用较少（差距不大）"),
    ("ADR-R-05", "thiserror vs anyhow 分工", "库用 thiserror + 应用用 anyhow", "库需明确错误 + 应用只关心成败", "公开 API 需统一错误码映射"),
    ("ADR-R-06", "选 axum HTTP 框架", "axum", "与 Tokio 完美集成 + tower 中间件", "纯性能略低于 actix-web（< 10%）"),
    ("ADR-R-07", "选 tonic gRPC", "tonic", "纯 Rust + 共享 tower + HTTP/2 streaming", "少数高级特性需 workaround"),
    ("ADR-R-08", "Tauri 桌面客户端", "Tauri 2.x", "详见技术选型书 v2.0 ADR-14", "WebView2 需内嵌（局域网离线）"),
    ("ADR-R-09", "选 sqlx PostgreSQL 驱动", "sqlx", "async + 编译时 SQL 校验", "ORM 关联查询要手写 SQL"),
    ("ADR-R-10", "redis-rs + deadpool-redis 缓存", "redis-rs + deadpool-redis", "Rust 生态最广泛 + 稳定连接池", "Cluster/Sentinel 需额外配置"),
    ("ADR-R-11", "选 rdkafka Kafka 客户端", "rdkafka", "工业级 librdkafka 绑定 + 性能/可靠性", "需安装 librdkafka 系统库"),
    ("ADR-R-12", "选 tracing 日志框架", "tracing", "原生 span + 与 OTel 集成", "部分老库用 log（需 tracing-log 桥接）"),
    ("ADR-R-13", "选 rustls TLS", "rustls", "纯 Rust + 内存安全 + 性能", "不支持 TLS 1.0/1.1"),
    ("ADR-R-14", "选 time 时间库", "time", "编译期类型化 + 性能 + API 现代", "生态迁移成本（多数教程用 chrono）"),
]
r = 4
for row in adrs:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 5)

# ---------------- 依赖版本锁定 ----------------
ws = wb.create_sheet("依赖版本锁定")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 24, "B": 14, "C": 24, "D": 14, "E": 16})
title_block(ws, "关键依赖版本锁定（§17）", 5)
headers = ["Crate", "版本", "用途", "锁定期", "升级方式"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
deps = [
    ("tokio", "1.40+", "异步运行时", "6 月（minor）", "锁 minor，自动 patch"),
    ("axum", "0.7+", "Web 框架", "6 月", "锁 minor"),
    ("tonic", "0.12+", "gRPC", "6 月", "锁 minor"),
    ("sqlx", "0.8+", "数据库", "6 月", "锁 minor"),
    ("rdkafka", "0.36+", "Kafka", "6 月", "锁 minor"),
    ("serde", "1.0+", "序列化", "永久（1.x）", "永不 major 升级"),
    ("prost", "0.13+", "Protobuf", "12 月", "锁 minor"),
    ("tracing", "0.1+", "日志/追踪", "12 月", "锁 minor"),
    ("opentelemetry", "0.27+", "OTel", "6 月", "锁 minor"),
    ("reqwest", "0.12+", "HTTP 客户端", "6 月", "锁 minor"),
    ("redis", "0.27+", "缓存", "6 月", "锁 minor"),
    ("thiserror", "1.0+", "错误处理（库）", "永久（1.x）", "永不 major 升级"),
    ("anyhow", "1.0+", "错误处理（应用）", "永久（1.x）", "永不 major 升级"),
    ("validator", "0.18+", "数据验证", "12 月", "锁 minor"),
    ("jsonwebtoken", "9.3+", "JWT", "12 月", "锁 minor"),
    ("argon2", "0.5+", "密码哈希", "永久（0.x 稳定）", "锁 minor"),
    ("ring", "0.17+", "加密", "永久", "随系统"),
    ("rustls", "0.23+", "TLS", "6 月", "锁 minor"),
    ("tauri", "2.1+", "桌面框架", "6 月", "锁 minor"),
    ("criterion", "0.5+", "基准", "12 月", "锁 minor"),
    ("proptest", "1.5+", "属性测试", "12 月", "锁 minor"),
    ("mockall", "0.13+", "mock", "12 月", "锁 minor"),
    ("testcontainers", "0.20+", "容器测试", "6 月", "锁 minor"),
]
r = 4
for row in deps:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 5)
ws.freeze_panes = "A4"

# ---------------- 工具链 ----------------
ws = wb.create_sheet("工具链")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 14, "C": 30, "D": 14})
title_block(ws, "Rust 工具链（§3）", 4)
headers = ["工具", "版本", "用途", "强制级别"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
tools = [
    ("rustc / cargo / clippy / rustfmt", "1.83+", "编译器与官方工具", "强制"),
    ("MSRV", "1.75+", "最小支持版本", "强制"),
    ("Edition", "2021", "Rust Edition", "强制"),
    ("cargo-nextest", "0.9+", "并行测试", "强制"),
    ("cargo-audit", "0.21+", "漏洞扫描（RustSec）", "强制"),
    ("cargo-deny", "1.16+", "许可证/重复依赖", "强制"),
    ("cargo-machete", "0.7+", "未使用依赖检测", "采纳"),
    ("cargo-bloat", "0.11+", "二进制大小分析", "采纳"),
    ("cargo-geiger", "0.11+", "unsafe 使用统计", "采纳"),
    ("cargo-binstall", "1.10+", "快速安装预编译二进制", "采纳"),
    ("cargo-flamegraph", "0.6+", "火焰图", "强制（性能分析时）"),
    ("cargo-make", "0.37+", "构建任务编排", "可选"),
    ("just", "1.34+", "Makefile 替代", "备选"),
    ("perf (Linux)", "—", "性能采样", "强制（Linux）"),
    ("pprof (Rust)", "0.13+", "CPU profile", "采纳"),
    ("heaptrack", "1.5+", "堆内存分析", "采纳"),
    ("valgrind / callgrind", "3.22+", "内存错误分析", "采纳"),
    ("semgrep", "—", "SAST（跨语言）", "强制（PR 门禁）"),
    ("syft + grype", "—", "SBOM + 漏洞", "强制（PR 门禁）"),
]
r = 4
for row in tools:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 异步与并发 ----------------
ws = wb.create_sheet("异步与并发")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 24, "B": 14, "C": 30, "D": 14})
title_block(ws, "异步与并发原语（§4）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
async_ = [
    ("tokio", "1.40+", "异步运行时", "强制（唯一）"),
    ("async-std", "—", "备选", "不采用（生态弱）"),
    ("smol", "—", "嵌入式场景", "不采用"),
    ("monoio", "—", "io_uring 极致性能", "备选（如需）"),
    ("glommio", "—", "线程绑定 io_uring", "备选（如需）"),
    ("tokio::sync::*", "—", "跨 async 同步", "强制"),
    ("parking_lot", "0.12+", "同步锁（高性能）", "强制"),
    ("arc-swap", "1.7+", "无锁读多写少", "采纳"),
    ("crossbeam", "0.8+", "高性能 channel", "采纳"),
    ("dashmap", "6.1+", "高并发 HashMap", "采纳"),
    ("rayon", "1.10+", "CPU 并行（data parallel）", "采纳"),
    ("thiserror", "1.0+", "库层结构化错误", "强制（库）"),
    ("anyhow", "1.0+", "应用层简化错误", "强制（应用）"),
    ("eyre", "0.6+", "anyhow fork", "备选"),
]
r = 4
for row in async_:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- Web / 网络 ----------------
ws = wb.create_sheet("Web与网络")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 14, "C": 30, "D": 14})
title_block(ws, "Web 与网络（§5）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
web = [
    ("axum", "0.7+", "HTTP 服务", "强制"),
    ("actix-web", "4.9+", "性能领先 Web 框架", "备选（生态与 Tokio 不完全兼容）"),
    ("warp", "0.3+", "灵活 API", "不采用（维护放缓）"),
    ("rocket", "0.5+", "同步为主", "不采用"),
    ("reqwest", "0.12+", "HTTP 客户端", "强制"),
    ("ureq", "2.10+", "轻量同步 HTTP", "备选"),
    ("isahc", "1.7+", "curl 绑定", "不采用"),
    ("tonic", "0.12+", "gRPC 客户端/服务", "强制"),
    ("tonic-build", "0.12+", "gRPC 代码生成", "强制"),
    ("tonic-reflection", "0.12+", "gRPC reflection（调试）", "采纳"),
    ("tonic-health", "0.12+", "gRPC health check", "采纳"),
    ("tokio-tungstenite", "0.24+", "WebSocket", "强制"),
    ("axum::extract::ws", "—", "axum 内置 WS", "强制"),
    ("tower", "0.5+", "中间件抽象", "强制"),
    ("tower-http", "0.6+", "HTTP 中间件（trace/cors/...）", "强制"),
    ("tauri", "2.1+", "桌面客户端", "强制"),
    ("keyring", "3.6+", "OS 密钥库", "强制"),
]
r = 4
for row in web:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 数据库 ----------------
ws = wb.create_sheet("数据库")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 14, "C": 30, "D": 14})
title_block(ws, "数据持久化（§6）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
db = [
    ("sqlx", "0.8+", "PG 驱动（异步 + 编译时校验）", "强制"),
    ("tokio-postgres", "0.7+", "底层 PG 驱动", "采纳（sqlx 依赖）"),
    ("diesel", "2.2+", "同步 ORM", "不采用"),
    ("sea-orm", "1.1+", "async ORM", "不采用"),
    ("sqlx-migrate", "0.8+", "迁移子命令", "强制"),
    ("refinery", "0.8+", "独立迁移库", "备选"),
    ("deadpool-postgres", "0.14+", "PG 连接池", "采纳"),
    ("bb8-postgres", "0.8+", "备选池", "备选"),
    ("r2d2-postgres", "0.18+", "同步池", "不采用"),
    ("aws-sdk-s3", "1.55+", "S3 客户端（MinIO 兼容）", "强制"),
    ("s3", "0.13+", "轻量 S3 客户端", "备选"),
    ("redis", "0.27+", "Redis/Valkey 客户端", "强制"),
    ("deadpool-redis", "0.18+", "Redis 连接池", "强制"),
    ("fred", "9.x", "高性能集群客户端", "备选"),
    ("bb8-redis", "0.17+", "备选池", "备选"),
]
r = 4
for row in db:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 消息与序列化 ----------------
ws = wb.create_sheet("消息与序列化")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 14, "C": 30, "D": 14})
title_block(ws, "消息系统与序列化（§7-§8）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
msg = [
    ("rdkafka", "0.36+", "Kafka 客户端（C 绑定）", "强制（详见 ADR-R-11）"),
    ("kafka-rust", "0.10+", "纯 Rust Kafka", "备选（功能不全）"),
    ("rskafka", "0.5+", "纯 Rust 异步 Kafka", "备选"),
    ("serde", "1.0+", "序列化框架", "强制"),
    ("serde_json", "1.0+", "JSON", "强制"),
    ("serde_yaml", "0.9+", "YAML 配置", "采纳"),
    ("serde_with", "3.11+", "自定义序列化辅助", "强制"),
    ("simd-json", "0.13+", "高性能 JSON", "采纳"),
    ("prost", "0.13+", "Protobuf 运行时", "强制"),
    ("prost-build", "0.13+", "Protobuf 代码生成", "强制"),
    ("rmp-serde", "1.3+", "MessagePack", "备选"),
    ("bincode", "2.0+", "二进制序列化", "备选"),
    ("validator", "0.18+", "数据验证（derive）", "强制"),
    ("garde", "0.20+", "数据验证备选", "备选"),
    ("url", "2.5+", "URL 类型", "强制"),
    ("uuid", "1.10+", "UUID v4/v7", "强制"),
]
r = 4
for row in msg:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 可观测性 ----------------
ws = wb.create_sheet("可观测性")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 24, "B": 14, "C": 30, "D": 14})
title_block(ws, "可观测性（§9）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
obs = [
    ("tracing", "0.1+", "结构化日志 + span", "强制（详见 ADR-R-12）"),
    ("tracing-subscriber", "0.3+", "subscriber 实现", "强制"),
    ("tracing-bunyan-formatter", "0.3+", "Bunyan 风格 JSON", "采纳"),
    ("tracing-log", "0.2+", "log → tracing 桥接", "采纳"),
    ("log", "—", "旧式日志", "不推荐（用 tracing）"),
    ("opentelemetry", "0.27+", "OTel API", "强制"),
    ("opentelemetry_sdk", "0.27+", "OTel SDK", "强制"),
    ("opentelemetry-otlp", "0.27+", "OTLP 导出器", "强制"),
    ("tracing-opentelemetry", "0.28+", "tracing → OTel 桥接", "强制"),
    ("opentelemetry-semantic-conventions", "0.27+", "OTel 语义约定", "强制"),
    ("metrics", "0.23+", "抽象指标 API", "强制"),
    ("metrics-exporter-prometheus", "0.15+", "Prometheus 导出", "强制"),
    ("axum-prometheus", "0.7+", "axum HTTP 指标", "采纳"),
    ("autometrics", "1.0+", "函数级指标", "采纳"),
]
r = 4
for row in obs:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 安全 ----------------
ws = wb.create_sheet("安全")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 14, "C": 30, "D": 14})
title_block(ws, "安全与加密（§12）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
sec = [
    ("jsonwebtoken", "9.3+", "JWT 签发/校验", "强制"),
    ("oauth2", "5.0+", "OAuth 2.0 客户端", "采纳"),
    ("openidconnect", "3.5+", "OpenID Connect", "采纳"),
    ("argon2", "0.5+", "密码哈希（OWASP 推荐）", "强制"),
    ("ring", "0.17+", "通用加密原语", "强制（详见 ADR-R-13 关联）"),
    ("aws-lc-rs", "1.10+", "ring 替代（AWS 维护）", "备选"),
    ("rustls", "0.23+", "TLS 实现", "强制（详见 ADR-R-13）"),
    ("openssl", "—", "OpenSSL", "不推荐（仅 C 库透传时）"),
    ("keyring", "3.6+", "OS 密钥库", "强制"),
    ("secrecy", "0.8+", "敏感值包装", "强制"),
    ("zeroize", "1.8+", "内存清零", "强制"),
]
r = 4
for row in sec:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 时间 ----------------
ws = wb.create_sheet("时间与配置")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 14, "C": 30, "D": 14})
title_block(ws, "时间 / 配置 / Cron（§11、§13）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
tm = [
    ("time", "0.3+", "时间处理（编译期类型化）", "强制（详见 ADR-R-14）"),
    ("chrono", "0.4+", "成熟时间库", "备选"),
    ("timeago", "0.4+", "相对时间显示", "采纳"),
    ("cron", "0.12+", "Cron 表达式解析", "强制（worker-service 调度）"),
    ("figment", "0.10+", "多源配置（env + file + 默认）", "强制"),
    ("config", "0.14+", "老牌配置库", "备选"),
    ("dotenvy", "0.15+", ".env 文件加载", "采纳"),
    ("envy", "0.4+", "环境变量 → 结构体", "采纳"),
]
r = 4
for row in tm:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 媒体处理 ----------------
ws = wb.create_sheet("媒体处理")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 14, "C": 30, "D": 14})
title_block(ws, "媒体处理 / 容器化 / K8s（§14-§16）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
mm = [
    ("ffmpeg-next", "7.1+", "FFmpeg Rust 绑定", "强制（render-writer）"),
    ("ffmpeg-sidecar", "—", "sidecar 方式调用 FFmpeg", "备选（更安全）"),
    ("image", "0.25+", "图像处理（PIL 替代）", "强制"),
    ("gif", "0.13+", "GIF 解码", "采纳"),
    ("webp", "0.3+", "WebP 解码", "采纳"),
    ("lopdf", "0.34+", "PDF 解析与生成", "采纳（轻量）"),
    ("pdf", "0.14+", "高层 PDF API", "备选"),
    ("kube", "0.95+", "K8s 客户端（Operator）", "强制（如需）"),
    ("k8s-openapi", "0.22+", "K8s OpenAPI 类型", "强制（如需）"),
    ("bollard", "0.15+", "Docker API 客户端", "采纳"),
]
r = 4
for row in mm:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 测试 ----------------
ws = wb.create_sheet("测试")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 14, "C": 30, "D": 14})
title_block(ws, "测试工具（§3.4）", 4)
headers = ["Crate", "版本", "用途", "决策"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
test = [
    ("cargo test", "—", "Rust 内置测试", "强制"),
    ("cargo-nextest", "0.9+", "并行测试运行", "强制"),
    ("mockall", "0.13+", "通用 mock（自动 mock trait）", "强制"),
    ("wiremock", "0.6+", "HTTP/gRPC mock server", "采纳"),
    ("mockito", "1.5+", "HTTP mock", "备选"),
    ("testcontainers", "0.20+", "真实容器测试", "强制"),
    ("proptest", "1.5+", "属性测试（property-based）", "采纳"),
    ("rstest", "0.21+", "参数化测试夹具", "采纳"),
    ("criterion", "0.5+", "性能基准", "采纳"),
    ("tokio-test", "0.4+", "异步测试辅助", "采纳"),
    ("pretty_assertions", "1.4+", "友好断言输出", "采纳"),
]
r = 4
for row in test:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 风险 ----------------
ws = wb.create_sheet("风险")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 10, "B": 36, "C": 8, "D": 8, "E": 50})
title_block(ws, "Rust 选型风险登记册（§18）", 5)
headers = ["编号", "风险", "概率", "影响", "缓解措施"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
risks = [
    ("RUST-01", "Tokio 升级引入行为变更", "中", "高", "锁 MSRV + 6 月升级窗口 + 充分测试"),
    ("RUST-02", "unsafe 误用导致内存安全问题", "低", "极高", "cargo geiger 检测 + 集中封装 + 100% 测试覆盖"),
    ("RUST-03", "C 依赖（librdkafka）系统兼容问题", "中", "中", "Docker 镜像固定系统库 + CI 多 OS 测试"),
    ("RUST-04", "生态快速迭代导致依赖频繁 breaking", "中", "中", "SemVer caret + 6 月 review + ADR for major 升级"),
    ("RUST-05", "编译时间长影响 CI 效率", "高", "中", "cargo-nextest + sccache + Docker layer 缓存"),
    ("RUST-06", "团队 Rust 熟练度不足", "中", "中", "内部培训 + code review 双签 + 资深人员 review"),
    ("RUST-07", "异步测试不稳定（flaky）", "中", "中", "tokio-test + 超时控制 + CI 重试 + flaky 标记"),
    ("RUST-08", "Tauri 跨平台兼容性（尤其 Linux）", "中", "中", "3 OS 平台持续集成 + 兼容性测试矩阵"),
    ("RUST-09", "sqlx 编译时 SQL 校验在 CI 与本地 DATABASE_URL 不匹配时失败", "中", "低", ".env 模板 + CI 注入 + 文档"),
    ("RUST-10", "性能退化难以发现", "中", "中", "criterion 基准 + cargo flamegraph + 性能基线对比"),
]
r = 4
for row in risks:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    for col in (3, 4):
        v = ws.cell(r, col).value
        if v in ("高", "极高"):
            ws.cell(r, col).fill = fill("F8CBAD")
        elif v == "中":
            ws.cell(r, col).fill = fill("FFE699")
        elif v == "低":
            ws.cell(r, col).fill = fill("C6E0B4")
    r += 1
grid(ws, 4, r - 1, 1, 5)
ws.freeze_panes = "A4"

# ---------------- Save ----------------
_here = os.path.dirname(os.path.abspath(__file__))
out_dir = os.path.dirname(_here)  # 上级 = 02-基础设计/技术选型/ 目录
out = os.path.join(out_dir, "CATs_Rust技术选型书_v1.0.xlsx")
wb.save(out)
print(f"Saved {out} with sheets: {wb.sheetnames}")
