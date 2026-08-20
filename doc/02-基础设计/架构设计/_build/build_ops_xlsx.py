# -*- coding: utf-8 -*-
"""CATs 可热插拔部署与运维设计书 -> Excel 工作簿生成
工作表: 封面 / 部署单元 / 镜像规范 / 事件Topic / DLQ管理 / 节点池 / HPA / 插件接口 /
       能力宣告 / 特性开关 / Feature Bundle / 管理员模块 / RBAC / 审批 / 风险 / 路线图
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Microsoft YaHei"
NAVY   = "1F3864"
HEADER = "2E5496"
SUB    = "8EAADB"
LIGHT  = "D9E1F2"
ZEBRA  = "F2F5FB"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True):   return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def topleft():         return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def title_block(ws, title, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.fill = fill(NAVY); c.font = f(13, True, WHITE); c.alignment = center(False)
    ws.row_dimensions[1].height = 28

def style_header(ws, row, ncols, fillc=HEADER):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(fillc); cell.font = f(10, True, WHITE)
        cell.alignment = center(); cell.border = BORDER

def grid(ws, r1, r2, c1, c2, zebra=True):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.vertical is None:
                cell.alignment = topleft()
            if zebra and (r - r1) % 2 == 1:
                if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = fill(ZEBRA)

def widths(ws, mp):
    for col, w in mp.items():
        ws.column_dimensions[col].width = w

# ============================================================
wb = Workbook()

# ---------------- 封面 ----------------
ws = wb.active
ws.title = "封面"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 26, "C": 56, "D": 3})
ws.merge_cells("B2:C2")
ws["B2"] = "可热插拔部署与运维设计"
ws["B2"].font = f(22, True, NAVY); ws["B2"].alignment = center(False)
ws.row_dimensions[2].height = 44
ws.merge_cells("B3:C3")
ws["B3"] = "CATs — 全媒体 AI 辅助翻译 SaaS 平台"
ws["B3"].font = f(12, True, "404040"); ws["B3"].alignment = center(False)
ws.row_dimensions[3].height = 24

meta = [
    ("文档编号", "CATs-ARCH-OPS-001"),
    ("文档名", "可热插拔部署与运维设计书"),
    ("版本", "第 1.0 版（草稿）"),
    ("创建日", "2026-08-19"),
    ("作者", "架构师"),
    ("状态", "评审前草稿"),
    ("密级", "仅社内"),
    ("上游文档", "CATs 微服务架构设计书 v1.0 / CATs 技术选型书 v2.0"),
    ("下游文档", "CATs 接口设计书 v2.0 / CATs 数据库设计书 v2.0 / CATs 模块设计书 v2.0 / CATs 测试设计书 v1.0"),
]
r = 5
for k, v in meta:
    ws.cell(r, 2, k).fill = fill(SUB); ws.cell(r, 2).font = f(10, True, WHITE)
    ws.cell(r, 2).alignment = center()
    ws.cell(r, 3, v).alignment = left(); ws.cell(r, 3).font = f(10)
    ws.cell(r, 2).border = BORDER; ws.cell(r, 3).border = BORDER
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
ws.cell(r, 2, "v1.0 主要补齐内容").font = f(10, True, NAVY)
r += 1
items = [
    "① 原子化部署的契约层（镜像规范、GitOps 流程、回滚契约）",
    "② 中心事件管理（DLQ、回放、追踪、保留策略）",
    "③ App 集群管理（节点池/HPA/PDB）",
    "④ 可热插拔架构（服务接口版本化、能力宣告、配置热更新、特性开关、插件接口、可选服务降级）",
    "⑤ 管理员运维界面（8 大模块 + RBAC + 审批工作流）",
    "⑥ 单功能独立升级（Feature Bundle 概念 + Canary + Expand-Contract 迁移）",
    "⑦ API 设计规范（针对热插拔的版本化、兼容性、错误码、幂等性、健康检查、追踪）",
    "⑧ 存储过程设计规范（数据库隔离、Outbox、兼容性迁移、跨服务数据访问、归档/备份）",
    "⑨ 风险登记册（15 条运维风险 + 缓解措施）",
    "⑩ 实施路线图（与项目 M1-M5 同步）",
]
for c_text in items:
    ws.cell(r, 2, c_text).font = f(9); ws.cell(r, 2).alignment = topleft()
    ws.cell(r, 2).border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 18
    r += 1

# ---------------- 部署单元 ----------------
ws = wb.create_sheet("部署单元")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 10, "B": 22, "C": 30, "D": 18, "E": 40})
title_block(ws, "三级部署单元定义", 5)
headers = ["级别", "单元", "原子性边界", "适用场景", "可回滚时长"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
units = [
    ("L1", "镜像（OCI）", "单个 OCI 镜像的构建/推送/拉取", "持续集成产物", "N/A（不可变）"),
    ("L2", "服务（K8s Deployment 资源集）", "单个服务的一次部署/回滚", "通用服务升级", "< 5min"),
    ("L3", "功能（Feature Bundle 跨服务组合）", "一个业务功能的一次发布/回滚", "跨服务的功能灰度", "< 5min（无 schema）/ < 30min（含 schema）"),
]
r = 4
for row in units:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 5)

# ---------------- 镜像规范 ----------------
ws = wb.create_sheet("镜像规范")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 16, "B": 40, "C": 36, "D": 18})
title_block(ws, "OCI 镜像命名规范与发布策略", 4)
headers = ["字段", "规则", "示例", "强制级别"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
img = [
    ("{service}", "小写连字符，与 K8s 资源名一致", "office-converter", "强制"),
    ("{version}", "SemVer 2.0（major.minor.patch），major 不轻易变更", "1.3.2", "强制"),
    ("{feature_set}", "stable / canary / beta / 特定 feature 名", "stable", "强制"),
    ("{git_sha}", "7 字符短哈希", "a1b2c3d", "强制"),
    ("{build_num}", "CI 构建序号（单调递增）", "42", "强制"),
    ("完整 tag", "{service}-{version}-{feature_set}-{git_sha}-{build_num}", "office-converter-1.3.2-stable-a1b2c3d-42", "—"),
    ("latest 策略", "仅供开发使用，**禁止生产引用**", "—", "强制"),
    ("不可变标签", "Harbor 强制开启；CI 重试不重 tag 而是新 build_num", "—", "强制"),
    ("回滚标签", "稳定版本可回滚 → 引用 {version} 标签", "office-converter:1.3.2", "强制"),
]
r = 4
for row in img:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 事件Topic ----------------
ws = wb.create_sheet("事件Topic")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 10, "B": 36, "C": 12, "D": 16, "E": 30, "F": 14})
title_block(ws, "Kafka 主题命名规范（{domain}.{entity}.{event_type}.v{n}）", 6)
headers = ["域", "主题", "分区数", "副本数", "保留期", "压缩"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 6)
topics = [
    ("task", "task.task.created.v1", 6, 3, "7d", "log"),
    ("task", "task.task.updated.v1", 6, 3, "7d", "log"),
    ("task", "task.task.canceled.v1", 6, 3, "7d", "log"),
    ("task", "task.task.completed.v1", 6, 3, "7d", "log"),
    ("task", "task.task.failed.v1", 6, 3, "7d", "log"),
    ("task", "task.media.asr.requested.v1", 12, 3, "7d", "log"),
    ("task", "task.media.asr.completed.v1", 12, 3, "7d", "log"),
    ("task", "task.media.ocr.requested.v1", 6, 3, "7d", "log"),
    ("task", "task.media.ocr.completed.v1", 6, 3, "7d", "log"),
    ("task", "task.media.subtitle.completed.v1", 6, 3, "7d", "log"),
    ("task", "task.media.office.completed.v1", 6, 3, "7d", "log"),
    ("task", "task.media.render.completed.v1", 6, 3, "7d", "log"),
    ("file", "file.file.uploaded.v1", 3, 3, "7d", "log"),
    ("file", "file.file.version_added.v1", 3, 3, "7d", "log"),
    ("user", "user.user.created.v1", 3, 3, "7d", "log"),
    ("user", "user.org.member_added.v1", 3, 3, "7d", "log"),
    ("user", "user.org.subscription_changed.v1", 3, 3, "7d", "log"),
    ("project", "project.project.glossary_updated.v1", 3, 3, "7d", "log"),
    ("project", "project.project.tm_updated.v1", 3, 3, "7d", "log"),
    ("project", "project.project.compliance_policy_changed.v1", 3, 3, "7d", "log"),
    ("audit", "audit.*.*.v1", 6, 3, "365d", "none"),
    ("notification", "notification.*.*.v1", 3, 3, "1d", "none"),
    ("task.dlq", "task.dlq.v1", 6, 3, "30d", "none"),
    ("file.dlq", "file.dlq.v1", 3, 3, "30d", "none"),
    ("user.dlq", "user.dlq.v1", 3, 3, "30d", "none"),
    ("project.dlq", "project.dlq.v1", 3, 3, "30d", "none"),
]
r = 4
for row in topics:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 6)
ws.freeze_panes = "A4"

# ---------------- DLQ管理 ----------------
ws = wb.create_sheet("DLQ管理")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 24, "B": 50, "D": 12})
title_block(ws, "DLQ 管理机制（中心化）", 3)
headers = ["机制", "说明", "SLA"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 3)
dlq = [
    ("DLQ Topic 命名", "每域独立 DLQ：{domain}.dlq.v1", "—"),
    ("DLQ 消息结构", "原消息 + 错误码/错误信息/重试次数/原始 offset/trace_id", "—"),
    ("DLQ 告警", "DLQ 消息数 > 阈值 → Slack/钉钉告警", "5min 内送达"),
    ("DLQ 重放", "单条/批量重放到原 topic", "需 platform_admin"),
    ("DLQ 永久删除", "归档后永久删除", "需 ≥ 2 platform_admin 审批"),
    ("DLQ 保留", "30 天", "自动过期"),
    ("自动修复策略", "可配置重试上限（默认 3 次）", "—"),
    ("消费者故障转移", "消费者组 rebalance 时未确认消息回到原 topic", "—"),
]
r = 4
for row in dlq:
    for i, v in enumerate(row[:3], 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 3)

# ---------------- 节点池 ----------------
ws = wb.create_sheet("节点池与HPA")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 12, "C": 30, "D": 14, "E": 14, "F": 16})
title_block(ws, "K8s 节点池、HPA 与 PDB 配置", 6)
headers = ["服务/对象", "节点池", "用途", "min副本", "max副本", "触发指标"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 6)
hpa = [
    ("auth-service",       "general",   "登录/认证", 2, 10, "CPU > 70%"),
    ("user-service",       "general",   "用户/组织管理", 2, 6, "CPU > 70%"),
    ("project-service",    "general",   "项目/术语库/TM", 2, 8, "CPU > 70%"),
    ("task-service",       "general",   "任务编排核心", 2, 10, "CPU > 70%"),
    ("file-service",       "general",   "文件/对象存储", 2, 6, "CPU > 70%"),
    ("notification-service","general",  "通知/WS", 2, 6, "CPU > 70%"),
    ("report-service",     "general",   "用量/计费", 1, 4, "CPU > 70%"),
    ("audit-service",      "general",   "审计", 1, 4, "CPU > 70%"),
    ("translation-core",   "general",   "翻译核心", 3, 30, "Kafka consumer lag"),
    ("ingestion-service",  "media-cpu", "媒体探测", 1, 5, "Kafka consumer lag"),
    ("asr-service",        "media-gpu", "语音转写", 2, 10, "Kafka consumer lag"),
    ("ocr-service",        "media-cpu", "图片/PDF OCR", 2, 10, "Kafka consumer lag"),
    ("subtitle-service",   "media-cpu", "字幕生成", 1, 5, "Kafka consumer lag"),
    ("office-converter-service", "media-cpu", "Office 转换", 2, 8, "Kafka consumer lag"),
    ("render-writer-service", "media-cpu", "渲染写回", 1, 5, "Kafka consumer lag"),
    ("worker-service",     "general",   "Cron/对账/批量导入", 1, 3, "Cron schedule"),
]
r = 4
for row in hpa:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 6)
ws.freeze_panes = "A4"

# ---------------- 插件接口 ----------------
ws = wb.create_sheet("插件接口")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 36, "C": 28, "D": 30})
title_block(ws, "插件接口设计（trait/Protocol）", 4)
headers = ["插件", "接口契约", "实现", "注入方式"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
plug = [
    ("Renderer（渲染器）", "trait Renderer { fn render(&self, job: RenderJob) -> Result<RenderOutput> }", "subtitle_burn_in / pdf_relayout / office_passthrough / gif_reencode / webp_reencode", "main.rs 按 render_kind 注入"),
    ("OcrEngine（OCR 引擎）", "trait OcrEngine { fn recognize(&self, image: &Image) -> Result<TextRegions> }", "PaddleOCR / Tesseract", "DI 容器"),
    ("AsrModel（ASR 模型）", "trait AsrModel { fn transcribe(&self, audio: &Audio) -> Result<Transcript> }", "faster-whisper-medium / faster-whisper-large", "DI 容器"),
    ("ModelProvider（翻译模型提供者）", "trait ModelProvider { fn translate(&self, segment: &Segment) -> Result<Translation> }", "OpenAI/Claude/Gemini/DeepSeek/本地 Qwen", "LiteLLM 路由"),
    ("Storage（存储后端）", "trait Storage { fn put(&self, key: &str, data: Bytes) -> Result<()> }", "MinIO / NFS", "DI 容器"),
    ("NotificationChannel（通知渠道）", "trait NotificationChannel { fn send(&self, msg: &Message) -> Result<()> }", "WS / 邮件 / 桌面 / 钉钉", "DI 容器"),
]
r = 4
for row in plug:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 能力宣告 ----------------
ws = wb.create_sheet("能力宣告")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 16, "B": 14, "C": 30, "D": 30, "E": 20})
title_block(ws, "服务能力宣告（capabilities 端点摘要）", 5)
headers = ["服务", "能力", "详细", "依赖", "降级行为"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
cap = [
    ("translation-core", "translate", "supported_languages / max_segment_length / streaming / modes", "project-service, ai-gateway", "返回错误"),
    ("translation-core", "tm", "exact_match / fuzzy_match / semantic_match / min_score", "project-service", "降级到无 TM 模式"),
    ("translation-core", "glossary", "enforcement / fallback", "project-service", "降级到无术语模式"),
    ("translation-core", "tag_protection", "placeholders / html_tags", "—", "—"),
    ("translation-core", "compliance", "local_model_required / fallback_on_local_unavailable", "ai-gateway", "fail-closed"),
    ("asr-service", "asr", "models / languages / min/max audio duration", "faster-whisper runtime", "服务不可用时返回 503"),
    ("ocr-service", "ocr", "engines / languages / tables / layout", "PaddleOCR/Tesseract", "服务不可用时返回 503"),
    ("subtitle-service", "subtitle", "formats / time_alignment / languages", "translation-core", "仅返回 ASR 文本"),
    ("office-converter-service", "office", "formats / version support / conversion", "LibreOffice/python-docx/openpyxl", "返回 503"),
    ("render-writer-service", "render", "render_kinds / output_formats / dub (reserved)", "ffmpeg", "仅返回翻译文本"),
    ("file-service", "storage", "backends / presigned_url / versions", "MinIO", "降级到直连文件系统"),
    ("notification-service", "notify", "channels / templates / preferences", "WebSocket", "降级到 DB 站内信"),
]
r = 4
for row in cap:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 5)

# ---------------- 特性开关 ----------------
ws = wb.create_sheet("特性开关")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 30, "B": 14, "C": 14, "D": 14, "E": 30})
title_block(ws, "特性开关（按 OpenFeature 分类）", 5)
headers = ["Flag Key", "类别", "默认值", "评估时机", "影响范围"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
ff = [
    ("enable_pdf_relayout_v2", "Release", "false", "用户请求时", "PDF 翻译功能用户"),
    ("pdf_relayout_v2_rollout_pct", "Rollout", "10", "用户请求时（按 %）", "灰度期间"),
    ("tm_match_algorithm", "Experiment", "v1", "启动时", "全局"),
    ("disable_asr_for_maintenance", "Ops", "false", "立即生效", "ASR 服务全局"),
    ("force_local_model_for_org_X", "Compliance", "false", "用户请求时", "特定 org"),
    ("enable_jira_mode", "Release", "false", "用户请求时", "Jira 模式用户"),
    ("enable_game_localization", "Release", "false", "用户请求时", "游戏本地化功能"),
    ("enable_multi_model_voting", "Release", "false", "用户请求时", "L3 模式用户"),
    ("enable_redis_cache", "Ops", "true", "启动时", "全局"),
    ("enable_strict_compliance_audit", "Compliance", "true", "立即生效", "审计服务"),
    ("ocr_engine_primary", "Ops", "paddleocr", "启动时", "OCR 服务"),
    ("subtitle_format_default", "Experiment", "srt", "用户请求时", "视频翻译用户"),
    ("enable_dlq_auto_replay", "Ops", "false", "立即生效", "DLQ 消息"),
    ("enable_dark_launch_compliance", "Experiment", "true", "用户请求时", "敏感项目用户"),
    ("render_quality_priority", "Experiment", "balanced", "用户请求时", "渲染服务"),
]
r = 4
for row in ff:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 5)

# ---------------- Feature Bundle ----------------
ws = wb.create_sheet("FeatureBundle")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 40, "C": 28, "D": 30})
title_block(ws, "Feature Bundle 样例（业务功能跨服务组合）", 4)
headers = ["Bundle 名", "包含服务（特定版本）", "业务能力", "Canary 策略"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
fb = [
    ("pdf-translation-v2", "ingestion v1.5.0 + ocr v2.0.0 + translation-core v2.1.0 + render-writer v1.3.0", "PDF 文档翻译 V2（增强版面保留）", "5%-25%-50%-100% × 30min"),
    ("video-subtitle-burn-in", "ingestion v1.5.0 + asr v1.2.0 + subtitle v1.1.0 + render-writer v1.3.0", "视频字幕烧录", "5%-25%-50%-100% × 30min"),
    ("compliance-local-model", "translation-core v2.1.0 + ai-gateway v1.5.0", "合规本地模型路由", "5%-50%-100% × 1h（敏感路径严控）"),
    ("tm-matching-v3", "translation-core v3.0.0 + project-service v1.6.0", "TM 匹配算法 V3（语义召回增强）", "1%-10%-50%-100% × 2h"),
    ("multi-model-voting", "translation-core v3.1.0 + ai-gateway v1.6.0", "L3 多模型投票", "1%-10%-50%-100% × 2h"),
    ("game-localization", "ingestion v1.6.0 + ocr v2.1.0 + translation-core v2.2.0 + render-writer v1.4.0", "游戏本地化全流程", "5%-25%-50%-100% × 30min"),
    ("ocr-table-recognition", "ocr v2.2.0 + ingestion v1.6.0", "OCR 表格识别增强", "10%-50%-100% × 1h"),
    ("jira-mode", "ingestion v1.7.0 + ocr v2.3.0 + translation-core v2.3.0", "Jira 工单翻译/总结/标签生成", "5%-25%-50%-100% × 1h"),
]
r = 4
for row in fb:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 管理员模块 ----------------
ws = wb.create_sheet("管理员模块")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 30, "C": 30, "D": 16})
title_block(ws, "管理员运维界面模块清单", 4)
headers = ["模块", "主要功能", "关键指标/视图", "优先级"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
admin = [
    ("服务总览",  "16 个服务的健康/版本/流量/延迟/错误率", "健康卡片 / 流量趋势", "M1"),
    ("部署管理",  "升级/回滚/暂停/继续/Canary 推进/审批", "版本列表 / 部署进度", "M1"),
    ("审计日志",  "所有运维操作记录（不可篡改）", "操作列表 + 过滤", "M1"),
    ("事件总线监视", "Topic 详情 / Consumer lag / DLQ 列表 / 事件回放", "Topic 列表 / 消费趋势", "M2"),
    ("特性开关管理", "Flag 启/停 / 灰度切换 / A/B 实验", "Flag 列表 + 评估规则", "M2"),
    ("数据库迁移", "Forward/Rollback 执行 / Schema 对比 / 备份恢复", "迁移历史 / 待执行", "M2"),
    ("日志/指标", "Loki 日志 / Prometheus 指标 / Tempo Trace 嵌入", "聚合查询面板", "M2"),
    ("DR 演练", "触发演练 / 回滚演练 / 验证清单", "演练历史 / 恢复时长", "M3"),
    ("插件管理", "查看已注册插件 / 启/停 / 升级", "插件列表", "M3"),
    ("审批工作流", "查看待审批 / 通过/拒绝", "审批队列", "M1"),
]
r = 4
for row in admin:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- RBAC ----------------
ws = wb.create_sheet("RBAC")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 60, "C": 18})
title_block(ws, "RBAC 角色与权限矩阵", 3)
headers = ["角色", "权限", "典型使用者"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 3)
rbac = [
    ("viewer",        "只读：查看服务状态、Metrics、日志、Trace", "开发/产品"),
    ("operator",      "viewer + 重启/扩缩容/暂停同步/查看 DLQ", "SRE"),
    ("deployer",      "operator + 升级/回滚/Canary（仅 STG/PRE）", "测试/SRE"),
    ("org_admin",     "viewer + 切换 org 维度的 Feature Flag + 查看 org 用量", "客户管理员"),
    ("platform_admin","全部权限（含生产部署审批/DLQ 删除/Schema 变更/DR 演练）", "平台管理员"),
    ("auditor",       "只读：所有操作 + 审计日志（不可被其他角色改动）", "合规/审计"),
]
r = 4
for row in rbac:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 3)

# ---------------- 审批 ----------------
ws = wb.create_sheet("审批")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 30, "B": 12, "C": 24, "D": 30})
title_block(ws, "运维操作审批工作流", 4)
headers = ["操作", "是否需审批", "审批人要求", "备注"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
appr = [
    ("STG/PRE 升级", "否", "—", "自动同步"),
    ("PROD 升级", "是", "≥ 1 platform_admin", "—"),
    ("PROD Canary 推进到 50%/100%", "是", "≥ 2 platform_admin", "严控"),
    ("任何回滚", "是", "≥ 1 platform_admin", "—"),
    ("DLQ 永久删除", "是", "≥ 2 platform_admin", "—"),
    ("Schema 变更", "是", "≥ 2 platform_admin（含 DBA）", "—"),
    ("DR 演练", "是", "≥ 1 platform_admin + 1 SRE", "—"),
    ("特性开关影响范围 > 10% 用户", "是", "1 platform_admin", "—"),
    ("紧急暂停（Suspend）", "否（事后审计）", "—", "事后补审批"),
    ("重启单 Pod（< 5 副本服务）", "否", "—", "自动放行"),
    ("重启单 Pod（≥ 5 副本服务）", "否", "—", "PDB 已保护"),
    ("删除生产数据", "是", "≥ 2 platform_admin + DBA", "强制双人"),
]
r = 4
for row in appr:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- 风险 ----------------
ws = wb.create_sheet("风险")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 10, "B": 30, "C": 8, "D": 8, "E": 50, "F": 30})
title_block(ws, "运维风险登记册", 6)
headers = ["编号", "风险", "概率", "影响", "缓解措施", "备注"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 6)
risks = [
    ("OPR-01", "镜像不可变未严格执行，CI 重试覆盖了 tag", "中", "高", "Harbor 强制不可变标签；CI 重试不重 tag；生产禁止 latest 标签", "—"),
    ("OPR-02", "灰度回滚时长 > 5min", "低", "高", "预先用 k6 压测回滚流程；Argo CD 自动化；演练", "—"),
    ("OPR-03", "Kafka 中心化后故障影响所有跨服务协同", "中", "高", "Kafka 3 broker HA + RF=3；监控；DR 演练", "架构 R-08"),
    ("OPR-04", "DLQ 累积未及时处理", "中", "中", "DLQ 告警阈值 + 运维 UI DLQ 管理；SLA：24h 内处理", "—"),
    ("OPR-05", "事件 schema 演进破坏消费者", "中", "高", "Schema Registry + CI 兼容性校验；破坏性变更强制升 v；共期 ≥ 2 发布周期", "—"),
    ("OPR-06", "能力宣告与实际不一致", "中", "中", "capabilities 端点 CI 验证（与代码断言一致）；金丝雀用例覆盖", "—"),
    ("OPR-07", "特性开关误操作影响生产", "中", "高", "影响范围 > 10% 用户需审批；全审计；可一键回滚", "—"),
    ("OPR-08", "插件注入失败", "中", "中", "启动时强校验必需插件；DI 失败时启动失败；fallback 机制", "—"),
    ("OPR-09", "可选服务缺失时调用方未降级", "中", "中", "启动时依赖检查 + 降级路径测试；运维 UI 标记状态", "—"),
    ("OPR-10", "运维 UI 越权", "中", "高", "RBAC 强制 + 中间件 + BFF 双重；越权拒绝；全审计", "—"),
    ("OPR-11", "Feature Bundle 依赖冲突", "中", "中", "manifest 声明依赖 + CI 校验；版本约束", "—"),
    ("OPR-12", "数据库迁移回滚失败", "低", "高", "Expand-Contract 强制；每步前先备份；reverse 脚本验证", "—"),
    ("OPR-13", "Canary 自动 abort 误判", "中", "中", "多指标综合判定；保守阈值；可手动覆盖", "—"),
    ("OPR-14", "审计日志被篡改", "低", "高", "审计事件 append-only（PostgreSQL trigger 禁止 UPDATE/DELETE）；定时异地归档", "—"),
    ("OPR-15", "运维 UI 单点故障", "中", "中", "3 副本 + 跨 AZ；只读操作可通过 CLI 替代；紧急 kubectl", "—"),
]
r = 4
for row in risks:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    for col in (3, 4):
        v = ws.cell(r, col).value
        if v in ("高", "极高"):
            ws.cell(r, col).fill = fill("F8CBAD")
        elif v == "中":
            ws.cell(r, col).fill = fill("FFE699")
        elif v == "低":
            ws.cell(r, col).fill = fill("C6E0B4")
    r += 1
grid(ws, 4, r - 1, 1, 6)
ws.freeze_panes = "A4"

# ---------------- 路线图 ----------------
ws = wb.create_sheet("路线图")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 16, "C": 50, "D": 18})
title_block(ws, "实施路线图（与项目 M1-M5 同步）", 4)
headers = ["阶段", "时间窗", "主要工作", "里程碑"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
phases = [
    ("阶段一（M1）",  "2026-09 ~ 2026-10",
     "① 镜像命名规范落地；② GitOps 流程（Argo CD）；③ STG 自动同步；"
     "④ 单一 Kafka 集群 + DLQ 基础；⑤ 事件 schema v1 注册；⑥ 节点池标签；⑦ HPA（基于 CPU）；"
     "⑧ 服务能力宣告端点（最小集）；⑨ 媒体处理插件 trait 实现；⑩ 运维 UI（服务总览 + 部署管理 + 审计）；"
     "⑪ 单服务级别升级；⑫ API 规范全部落地；⑬ Outbox + Expand-Contract 落地",
     "M1 MVP 闭环"),
    ("阶段二（M2）", "2026-11 ~ 2027-01",
     "① Feature Bundle manifest 格式；② PROD Canary 自动化；"
     "③ 事件回放 API；④ DLQ 运维 UI 完善；⑤ Schema Registry 落地；"
     "⑥ HPA 基于 Kafka consumer lag（KEDA）；⑦ PDB 全部服务；"
     "⑧ OpenFeature 落地；⑨ 运行时配置热更新（ConfigMap + Reloader）；"
     "⑩ 运维 UI（事件总线 + 特性开关 + 数据库迁移模块）；"
     "⑪ 跨服务 Feature Bundle 试点（PDF 翻译 V2）；"
     "⑫ 历史数据归档 job；⑬ 备份恢复流程",
     "M2 数据底座"),
    ("阶段三（M3-M5）", "2027-02 ~ 2027-Q4",
     "① 跨集群多环境；② Federation 能力评估；"
     "③ 事件版本化 2 套并存；④ 事件流分析工具；"
     "⑤ Cluster Autoscaler 弹性节点；⑥ 跨节点池调度优化；"
     "⑦ 第三方插件注册机制；⑧ 插件市场（内部）；"
     "⑨ 完整运维 UI 全部模块上线；⑩ DR 演练模块；"
     "⑪ 完整 Feature Bundle 灰度 + 自动 abort + 一键回滚；"
     "⑫ gRPC streaming 增强；⑬ multipart upload；"
     "⑭ PITR 落地；⑮ DR 演练每季度",
     "M3-M5 文档/工作流/增强能力"),
]
r = 4
for row in phases:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)
ws.freeze_panes = "A4"

# ---------------- Save ----------------
_here = os.path.dirname(os.path.abspath(__file__))
out_dir = os.path.dirname(_here)  # 上级 = 02-基础设计/架构设计/ 目录
out = os.path.join(out_dir, "CATs_可热插拔部署与运维设计_v1.0.xlsx")
wb.save(out)
print(f"Saved {out} with sheets: {wb.sheetnames}")
