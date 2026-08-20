# -*- coding: utf-8 -*-
"""CATs 测试设计书 v2.0 -> Excel 工作簿生成
按 JIS X 0129 (ISO/IEC/IEEE 29119) + IPA/SEC 实践重写：
- 保留 v1.0 的 11 个数据型工作表
- 新增 V 字模型矩阵、覆盖率与缺陷注入率、RACI 矩阵、文档模板、术语集、引用标准
- 适配 v2.0 .md 的章节结构
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Microsoft YaHei"
NAVY   = "1F3864"
HEADER = "2E5496"
SUB    = "8EAADB"
LIGHT  = "D9E1F2"
ZEBRA  = "F2F5FB"
WHITE  = "FFFFFF"
GREEN  = "548235"
ORANGE = "BF8F00"
RED    = "C00000"
GREY   = "A6A6A6"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True):   return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def topleft():         return Alignment(horizontal="left",   vertical="top",    wrap_text=True)
def topright():        return Alignment(horizontal="right",  vertical="top",    wrap_text=True)

def title_block(ws, title, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.fill = fill(NAVY); c.font = f(13, True, WHITE); c.alignment = center(False)
    ws.row_dimensions[1].height = 28

def style_header(ws, row, ncols, fillc=HEADER):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(fillc); cell.font = f(10, True, WHITE)
        cell.alignment = center(); cell.border = BORDER

def grid(ws, r1, r2, c1, c2, zebra=True):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.vertical is None:
                cell.alignment = topleft()
            if zebra and (r - r1) % 2 == 1:
                if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = fill(ZEBRA)

def widths(ws, mp):
    for col, w in mp.items():
        ws.column_dimensions[col].width = w

# ============================================================
wb = Workbook()

# ---------------- Sheet 1: 封面 ----------------
ws = wb.active
ws.title = "封面"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 26, "C": 56, "D": 3})
ws.merge_cells("B2:C2")
ws["B2"] = "测试设计书 v2.0（IPA 标准化版）"
ws["B2"].font = f(22, True, NAVY); ws["B2"].alignment = center(False)
ws.row_dimensions[2].height = 44
ws.merge_cells("B3:C3")
ws["B3"] = "CATs — 全媒体 AI 辅助翻译 SaaS 平台"
ws["B3"].font = f(12, True, "404040"); ws["B3"].alignment = center(False)
ws.row_dimensions[3].height = 24

meta = [
    ("文档编号", "CATs-TST-001"),
    ("文档名", "测试设计书（Test Design Specification，IPA 标准化版）"),
    ("版本", "第 2.0 版（草稿）"),
    ("创建日", "2026-08-19"),
    ("作者", "测试负责人"),
    ("状态", "评审前草稿"),
    ("密级", "仅社内（Confidential / Internal Only）"),
    ("适用标准", "JIS X 0129-1/2/3/4:2013-2016（ISO/IEC/IEEE 29119 シリーズ）+ IPA/SEC"),
    ("上游文档", "OFCAT 需求定义书 v1.1 / CATs 微服务架构设计书 v1.0 / CATs 技术选型书 v2.0 / CATs 接口设计书 v2.0 / CATs 数据库设计书 v2.0 / CATs 模块设计书 v2.0"),
]
r = 5
for k, v in meta:
    ws.cell(r, 2, k).fill = fill(SUB); ws.cell(r, 2).font = f(10, True, WHITE)
    ws.cell(r, 2).alignment = center()
    ws.cell(r, 3, v).alignment = left(); ws.cell(r, 3).font = f(10)
    ws.cell(r, 2).border = BORDER; ws.cell(r, 3).border = BORDER
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
ws.cell(r, 2, "v1.0 → v2.0 主要变更").font = f(10, True, NAVY)
r += 1
changes = [
    "① 按 JIS X 0129（ISO/IEC/IEEE 29119）系列 + IPA/SEC 实践重写章节结构",
    "② 追加 V 字模型映射与各测试级别详细技法（同値分割/境界値/状態遷移/ユースケース/ペアワイズ等）",
    "③ 追加结构化测试覆盖率（ステートメント/ブランチ/MC/DC）与缺陷注入率（ミューテーション）基线",
    "④ 追加风险驱动测试（リスクベースドテスト）与项目级风险登记册",
    "⑤ 追加测试体制·RACI 矩阵·教育训练",
    "⑥ 追加缺陷去除率阶段目标（4 阶段 ≥ 99%）",
    "⑦ 追加引用标准清单（JIS/ISO/IPA/OWASP/CWE）+ JIS X 0129 术语集（中日英对照）",
    "⑧ 追加 JIS X 0129-3 测试文档模板（测试用例规格书/结果报告/测试日志）",
]
for c_text in changes:
    ws.cell(r, 2, c_text).font = f(9); ws.cell(r, 2).alignment = topleft()
    ws.cell(r, 2).border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 18
    r += 1

r += 1
ws.cell(r, 2, "统计摘要").font = f(10, True, NAVY)
r += 1
stats = [
    ("覆盖服务数（含客户端/平台）", '=COUNTA(测试项总览!A4:A200)'),
    ("测试用例总条数（按各服务汇总）", '=COUNTA(测试项总览!A4:A2000)'),
    ("E2E 核心场景数", '=COUNTA(媒体类型E2E!A4:A100)'),
    ("MVP 功能追溯条数（F1-F11）", '=COUNTA(需求追溯矩阵!A4:A100)'),
    ("风险登记条数", '=COUNTA(风险与缓解!A4:A100)'),
    ("性能 SLO 条数", '=COUNTA(性能SLO!A4:A100)'),
    ("里程碑数", '=COUNTA(里程碑!A4:A100)'),
    ("覆盖率基线服务数", '=COUNTA(覆盖率与缺陷注入率!A4:A100)'),
    ("RACI 活动数", '=COUNTA(RACI矩阵!A4:A100)'),
    ("引用标准数", '=COUNTA(引用标准!A4:A100)'),
]
for k, fml in stats:
    ws.cell(r, 2, k).fill = fill(LIGHT); ws.cell(r, 2).font = f(10, True)
    ws.cell(r, 2).alignment = center()
    ws.cell(r, 3, fml).font = f(10, True, NAVY); ws.cell(r, 3).alignment = left(False)
    ws.cell(r, 2).border = BORDER; ws.cell(r, 3).border = BORDER
    ws.row_dimensions[r].height = 20
    r += 1

# ---------------- Sheet 2: 测试项总览 ----------------
ws = wb.create_sheet("测试项总览")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 12, "B": 22, "C": 14, "D": 18, "E": 40, "F": 10, "G": 10, "H": 10, "I": 12})
title_block(ws, "测试项总览（按服务 × 测试类型 × V 字模型级别）", 9)
headers = ["编号", "服务/对象", "测试类型", "子模块/重点", "关键测试项描述", "数量级", "优先级", "关联需求/原则", "来源文档"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 9)

items = [
    ("S-01", "auth-service", "单元", "Token/密码/RBAC", "Token 签发/校验、Argon2 密码哈希、RBAC 角色判定、Session 状态机", "30+", "P1", "C2 鉴权", "接口设计书 §3.1"),
    ("S-01", "auth-service", "集成", "数据表/迁移", "users_credential/sessions/roles 表 CRUD、迁移脚本前后、跨表外键", "15+", "P1", "§5.3 数据一致性", "数据库设计书 §4.1"),
    ("S-01", "auth-service", "契约", "gRPC/REST", "gRPC AuthInternal.AuthCheck、REST OpenAPI、auth.events schema", "10+", "P1", "§5.3 接口契约", "接口设计书 §3.1"),
    ("S-01", "auth-service", "接口", "6 个 endpoint", "login/oidc-callback/refresh/logout/roles/role-bindings、错误码 11 种全覆盖", "30+", "P1", "§5.1 接口", "接口设计书 §1.4/§3.1"),
    ("S-01", "auth-service", "安全", "鉴权攻击面", "密码爆破/Token 重放/JWT 算法替换/refresh_token 滥用/CSRF（web 登录）", "15+", "P0", "C2 安全", "架构设计书 §1.2 原则 1-5"),
    ("S-02", "user-service", "单元", "套餐/邀请", "套餐限额判定、邀请 token 生成、组织/用户关系遍历", "20+", "P2", "—", "接口设计书 §3.2"),
    ("S-02", "user-service", "集成", "数据表/跨库一致", "orgs/users_profile/org_members/subscriptions CRUD、跨库 user_id 引用一致性", "15+", "P2", "§5.3 数据一致性", "数据库设计书 §4.2"),
    ("S-02", "user-service", "契约", "OpenAPI/事件", "OpenAPI、user.events schema（含 org.subscription_changed）", "8+", "P2", "§5.3 接口契约", "接口设计书 §3.2"),
    ("S-02", "user-service", "接口", "7 个 endpoint", "/v1/users 系列、跨 org 越权防护", "20+", "P1", "§5.1 接口", "接口设计书 §3.2"),
    ("S-02", "user-service", "安全", "越权/邀请滥用", "跨 org 越权、邀请链接 token 一次性、邮箱枚举", "10+", "P0", "C2 安全", "架构设计书 §1.2 原则 1"),
    ("S-03", "project-service", "单元", "TM/术语/向量核心算法", "术语强制校验、TM 匹配 L0/L1 判定、pgvector 相似度、多语种维度约束", "40+", "P0", "F2/F3/F4 确定性优先", "需求定义书 §5.2/接口设计书 §3.3"),
    ("S-03", "project-service", "集成", "数据表/分区/索引", "projects/terms/glossary_versions/translation_memory/tm_vectors CRUD、HASH 分区、HNSW 索引", "25+", "P0", "§5.3 数据一致性", "数据库设计书 §4.3"),
    ("S-03", "project-service", "契约", "gRPC/OpenAPI/事件", "gRPC ProjectInternal（GetGlossary/SearchTM）、OpenAPI、project.events schema", "10+", "P1", "§5.3 接口契约", "接口设计书 §3.3"),
    ("S-03", "project-service", "接口", "REST 系列/TM 搜索", "/v1/projects 系列、TM 搜索（精确/语义）", "20+", "P1", "F2/F3", "接口设计书 §3.3"),
    ("S-03", "project-service", "性能", "TM 匹配 SLO", "L0 < 100ms p95、L1 < 300ms p95、pgvector 百万级 HNSW 召回延迟", "10+", "P1", "性能 SLO L0/L1", "需求定义书 §6.2"),
    ("S-03", "project-service", "安全", "越权读取/修改", "跨 org 访问、术语库越权读取、敏感策略越权修改", "8+", "P0", "C2/C3 合规", "需求定义书 §6.5"),
    ("S-04", "task-service", "单元", "状态机/编排", "任务状态机迁移校验、媒体子任务编排规则、SLA 超时判定、Stage 进度聚合", "50+", "P0", "F6/F8 状态机", "接口设计书 §3.4/§6"),
    ("S-04", "task-service", "集成", "任务表/Outbox", "tasks/task_media_items/media_assets/asr_transcripts/subtitle_segments/task_events_outbox CRUD、双写一致性", "30+", "P0", "§5.3 数据一致性", "数据库设计书 §4.4"),
    ("S-04", "task-service", "契约", "OpenAPI/事件/内部 API", "OpenAPI、task.events schema、task.media.*.requested/completed、stage-progress 内部 API", "20+", "P0", "§5.3 接口契约", "接口设计书 §3.4/§6"),
    ("S-04", "task-service", "接口", "5 对外+SSE+内部", "/v1/tasks 系列(POST/GET/cancel/retry)+ SSE 事件流 + Stage 进度上报", "30+", "P0", "§5.1 接口", "接口设计书 §3.4"),
    ("S-04", "task-service", "数据一致性", "双写+对账", "业务表 + Outbox 同事务、崩溃后事件补齐、对账 job", "10+", "P0", "单一权威原则", "架构设计书 §7"),
    ("S-04", "task-service", "性能", "高并发/长任务", "500 并发创建任务、8h 视频状态查询与 SSE 推送", "10+", "P1", "性能 SLO", "需求定义书 §6.2"),
    ("S-05", "file-service", "单元", "预签名/版本", "预签名 URL 生成/校验、文件元数据校验、版本号派生", "15+", "P1", "—", "接口设计书 §3.5"),
    ("S-05", "file-service", "集成", "文件表/对象存储", "files 表 CRUD、MinIO 桶策略、软删除保留", "10+", "P1", "—", "数据库设计书 §4.5"),
    ("S-05", "file-service", "契约", "OpenAPI/事件", "OpenAPI、file.events schema", "6+", "P1", "§5.3 接口契约", "接口设计书 §3.5"),
    ("S-05", "file-service", "接口", "6 endpoint/直传", "6 个 endpoint + 预签名 PUT 直传", "20+", "P1", "§5.1 接口", "接口设计书 §3.5"),
    ("S-05", "file-service", "安全", "预签名/越权", "预签名 URL 过期/盗链/越权、跨 org、软删除前访问阻断", "10+", "P0", "C2 安全", "架构设计书 §1.2 原则 1"),
    ("S-05", "file-service", "性能", "大文件/高并发", "2GB 分片上传、预签名 URL 1000 QPS 生成", "8+", "P2", "—", "需求定义书 §6.2"),
    ("S-06", "notification-service", "单元", "模板/偏好/通道", "通知模板渲染、用户偏好判定、多通道分发", "15+", "P2", "—", "接口设计书 §3.6"),
    ("S-06", "notification-service", "集成", "表/WS 连接", "notification_db CRUD、WebSocket 连接状态", "10+", "P2", "—", "数据库设计书 §4.7"),
    ("S-06", "notification-service", "契约", "WS 消息/事件", "WebSocket 消息 schema、notification.events 消费 schema", "8+", "P2", "§5.3 接口契约", "接口设计书 §3.6"),
    ("S-06", "notification-service", "接口", "WS/5 endpoint", "WebSocket 长连接、5 个 REST endpoint", "15+", "P1", "§5.1 接口", "接口设计书 §3.6"),
    ("S-06", "notification-service", "性能", "高并发 WS", "10000 并发 WS 稳定推送、推送延迟 p95 < 1s", "5+", "P1", "性能 SLO", "—"),
    ("S-07", "report-service", "单元", "聚合/计费/QA", "用量聚合、计费规则、QA 命中率统计", "15+", "P1", "—", "接口设计书 §3.7"),
    ("S-07", "report-service", "集成", "表/跨库只读", "report_db CRUD、svc_report_ro 权限验证", "8+", "P1", "单一权威/只读", "数据库设计书 §2"),
    ("S-07", "report-service", "契约", "OpenAPI/事件", "OpenAPI、task.events 消费 schema", "5+", "P2", "§5.3 接口契约", "接口设计书 §3.7"),
    ("S-07", "report-service", "接口", "3 只读 endpoint", "3 个只读 endpoint、查询权限校验", "10+", "P1", "§5.1 接口", "接口设计书 §3.7"),
    ("S-07", "report-service", "数据准确性", "跨 org 隔离/核算", "批用人工核算数据校对、跨 org 数据隔离", "8+", "P1", "—", "—"),
    ("S-08", "audit-service", "单元", "规范化/过滤", "审计事件规范化、查询过滤器解析", "10+", "P1", "C3 合规", "接口设计书 §3.8"),
    ("S-08", "audit-service", "集成", "表/跨库只读", "audit_db CRUD、跨库只读性能", "5+", "P1", "—", "数据库设计书 §4.8"),
    ("S-08", "audit-service", "契约", "事件/OpenAPI", "audit.events 消费 schema、OpenAPI", "5+", "P1", "§5.3 接口契约", "接口设计书 §3.8"),
    ("S-08", "audit-service", "接口", "2 只读 endpoint", "2 个只读 endpoint、platform_admin 角色校验", "8+", "P1", "§5.1 接口", "接口设计书 §3.8"),
    ("S-08", "audit-service", "合规", "关键操作落库", "术语变更/策略变更/敏感项目访问/登录失败/权限变更 100% 落库", "15+", "P0", "C3 合规 fail-closed", "架构设计书 §1.2 原则 3"),
    ("S-09", "translation-core", "单元", "LangGraph 节点/确定性", "TM 匹配/术语注入/标签保护/模型翻译/术语校验/QA 节点确定性", "60+", "P0", "F2-F7 确定性优先", "需求定义书 §5.2/接口设计书 §3.10"),
    ("S-09", "translation-core", "集成", "读/缓存/gRPC 网关", "project_db 只读、Valkey 缓存、gRPC Client、LiteLLM 网关", "30+", "P0", "—", "模块设计书 §1.3"),
    ("S-09", "translation-core", "契约", "gRPC/事件", "gRPC TranslationCore（Translate 流式/TranslateBatch/TMMatch/QACheck）、project.events 消费", "15+", "P0", "§5.3 接口契约", "接口设计书 §3.10"),
    ("S-09", "translation-core", "接口", "4 gRPC/2 REST", "4 个 gRPC method + 2 个内部 REST、错误码 COMPLIANCE_BLOCKED/QA_BLOCKED/UPSTREAM_*", "30+", "P0", "§5.1 接口", "接口设计书 §3.10/§1.4"),
    ("S-09", "translation-core", "性能", "延迟分层 SLO", "L0/L1/L2/L3 全部达标、流式首字延迟、TranslateBatch 批量效率", "15+", "P0", "性能 SLO L0-L3", "需求定义书 §6.2"),
    ("S-09", "translation-core", "合规", "fail-closed/校验", "敏感项目 fail-closed、本地模型不可达行为、术语强制 100%", "15+", "P0", "F10/C3 合规", "需求定义书 §5.2 F10"),
    ("S-10", "ingestion-service", "单元", "媒体探测", "ffprobe/PyMuPDF/LibreOffice 探测规则", "10+", "P1", "—", "接口设计书 §4.1"),
    ("S-10", "ingestion-service", "集成", "子任务规划", "task_media_items 规划写入", "10+", "P1", "—", "数据库设计书 §4.4"),
    ("S-10", "ingestion-service", "异步事件", "事件消费/产出", "file.events+task.events 消费、task.media.*.requested 产出", "10+", "P1", "§5.3 接口契约", "接口设计书 §4.1"),
    ("S-11", "asr-service", "单元", "faster-whisper", "模型选择/分段策略/CUDA_OOM 降级", "15+", "P0", "F6 L2", "模块设计书 §1.4/§4.2"),
    ("S-11", "asr-service", "集成", "转写写入/去重", "asr_transcripts 写入、Valkey 幂等去重", "10+", "P0", "—", "数据库设计书 §4.4"),
    ("S-11", "asr-service", "异步事件", "事件消费/产出", "task.media.asr.requested 消费、.completed 产出", "8+", "P0", "§5.3 接口契约", "接口设计书 §4.2"),
    ("S-11", "asr-service", "性能", "CPU/GPU 推理", "CPU/GPU 推理时长、降级到 CPU 行为", "5+", "P1", "R-10 资源争抢", "技术选型书 R-10"),
    ("S-11", "asr-service", "故障注入", "OOM/降级", "CUDA_OOM 触发降级到 CPU、不无限重试", "5+", "P0", "R-10", "模块设计书 §4.2"),
    ("S-12", "ocr-service", "单元", "PaddleOCR/抽帧", "PaddleOCR/Tesseract 切换、抽帧密度", "10+", "P1", "—", "接口设计书 §4.3"),
    ("S-12", "ocr-service", "集成", "结构化落盘", "结构化 JSON 落 file-service", "8+", "P1", "—", "接口设计书 §4.3"),
    ("S-12", "ocr-service", "异步事件", "事件消费/产出", "task.media.ocr.requested 消费、.completed 产出", "8+", "P1", "§5.3 接口契约", "接口设计书 §4.3"),
    ("S-12", "ocr-service", "性能", "OCR 推理", "PaddleOCR 中日英推理时长、抽帧吞吐", "4+", "P2", "—", "—"),
    ("S-13", "subtitle-service", "单元", "时间轴/格式", "srt/vtt/ass 格式生成、时间轴切分规则", "15+", "P0", "—", "接口设计书 §4.4"),
    ("S-13", "subtitle-service", "集成", "段落写入/gRPC", "subtitle_segments 写入、gRPC 调 translation-core", "10+", "P0", "—", "数据库设计书 §4.4"),
    ("S-13", "subtitle-service", "异步事件", "事件消费/产出", "task.media.asr.completed 消费、.subtitle.completed 产出", "10+", "P0", "§5.3 接口契约", "接口设计书 §4.4"),
    ("S-14", "office-converter-service", "单元", "结构化库抽取/回填", "python-docx/openpyxl/python-pptx/odfpy 抽取与回填", "20+", "P0", "—", "接口设计书 §4.5"),
    ("S-14", "office-converter-service", "集成", "LibreOffice 进程池", "LibreOffice Headless 进程池、超时熔断", "10+", "P0", "R-09", "技术选型书 R-09"),
    ("S-14", "office-converter-service", "异步事件", "事件消费/产出", "task.media.office.requested 消费、.completed 产出", "10+", "P0", "§5.3 接口契约", "接口设计书 §4.5"),
    ("S-14", "office-converter-service", "格式覆盖", "6+ 格式", "docx/xlsx/pptx/odt/ods/odp/老 doc/xls/ppt", "10+", "P1", "—", "需求定义书 §3.5"),
    ("S-14", "office-converter-service", "故障注入", "进程僵死", "LibreOffice 进程僵死后被强制 kill + 任务重试", "5+", "P1", "R-09", "技术选型书 R-09"),
    ("S-15", "render-writer-service", "单元", "5 种 Renderer", "subtitle_burn_in/pdf_relayout/office_passthrough/gif_reencode/webp_reencode trait 实现", "20+", "P0", "—", "模块设计书 §1.2/§4.5"),
    ("S-15", "render-writer-service", "集成", "落盘/上报", "file-service 落盘、task-service 上报", "10+", "P0", "—", "接口设计书 §4.6"),
    ("S-15", "render-writer-service", "异步事件", "事件消费/产出", "task.media.render.requested 消费、.completed 产出", "10+", "P0", "§5.3 接口契约", "接口设计书 §4.6"),
    ("S-16", "worker-service", "单元", "Cron/对账/清洗", "Cron 调度、对账超时判定、批量导入清洗规则", "10+", "P2", "—", "接口设计书 §3.9"),
    ("S-16", "worker-service", "集成", "API 间接读写", "经 task-service API 间接读写、对账结果", "8+", "P2", "—", "接口设计书 §3.9"),
    ("C-01", "Tauri 客户端", "单元", "Rust 核心/Svelte", "Rust commands、Svelte 组件", "40+", "P0", "F1/F7/F8", "模块设计书 §2"),
    ("C-01", "Tauri 客户端", "E2E", "12 核心场景", "12 个 E2E 核心场景覆盖（见媒体类型E2E 表）", "55+", "P0", "F1-F10", "需求定义书 §5.2"),
    ("C-01", "Tauri 客户端", "离线队列", "断网/恢复", "断网→编辑→恢复→自动重放、0 丢失 0 重复、冲突提示", "10+", "P0", "F9/§6.5 安全", "模块设计书 §2.3"),
    ("C-01", "Tauri 客户端", "安全", "Keyring/更新签名", "OS 密钥库集成、自动更新签名校验", "5+", "P0", "C2 安全/R-06", "模块设计书 §2.2/技术选型 R-06"),
    ("C-01", "Tauri 客户端", "兼容性", "3 OS", "Windows 10/11、macOS 12+ (M1+Intel)、Ubuntu 22.04/Fedora", "15+", "P1", "R-06 WebView2", "§4.8 兼容性"),
    ("C-02", "Next.js 控制台", "单元", "中间件/Route", "middleware、API Route Handler", "20+", "P1", "—", "模块设计书 §3.2"),
    ("C-02", "Next.js 控制台", "E2E", "Playwright", "主要页面 + BFF 聚合", "15+", "P1", "—", "模块设计书 §3"),
    ("C-02", "Next.js 控制台", "鉴权", "中间件/二次校验", "中间件穿透、/admin/* 仅 platform_admin", "10+", "P0", "C2 安全", "模块设计书 §3.2"),
    ("C-02", "Next.js 控制台", "兼容性", "3 浏览器", "Chrome/Edge/Firefox 最新+前一LTS", "10+", "P2", "—", "§4.8 兼容性"),
    ("P-01", "Envoy Gateway", "接口/限流/熔断", "7 大能力", "HTTPRoute/GRPCRoute 分流、JWT 校验、限流、熔断、灰度权重、TLS、内部 CA 证书轮换", "30+", "P0", "—", "架构设计书 §1.1/§14"),
    ("P-02", "PostgreSQL 集群", "部署/HA/迁移", "5 大能力", "CloudNativePG 部署、PITR、主从延迟、逻辑复制槽、HA failover、Expand-Contract 迁移", "25+", "P0", "单一权威", "架构设计书 §5/§15.3"),
    ("P-03", "Valkey", "部署/切换/限流", "4 维度", "部署、持久化、故障切换、限流计数器", "10+", "P1", "禁做主存储", "架构设计书 §11"),
    ("P-04", "Kafka 集群", "KRaft/分区/ACL", "5 维度", "KRaft 部署、分区副本、consumer lag 告警、Topic 自动创建、ACL 权限", "20+", "P0", "—", "架构设计书 §6"),
    ("P-05", "Debezium CDC", "LSN/演进/恢复", "4 维度", "复制槽 LSN 推进、Schema 演进、故障恢复不丢事件", "15+", "P0", "Outbox 原子性", "架构设计书 §7"),
    ("P-06", "Harbor", "部署/扫描/签名", "4 维度", "部署、Trivy 扫描、复制策略、签名校验", "10+", "P1", "R-11 单点", "架构设计书 §1.1/技术选型 R-11"),
    ("P-07", "Prometheus+Grafana", "指标/告警/仪表板", "3 维度", "指标抓取、告警触发、仪表板可读性", "15+", "P1", "—", "架构设计书 §13"),
    ("P-08", "OTel + Tempo", "Trace 传播/查询", "3 维度", "Trace 上下文贯穿同步+异步、采样率、查询性能", "10+", "P0", "§4.1 全链路 Trace", "架构设计书 §13.2"),
    ("P-09", "Loki + Promtail", "日志/关联", "3 维度", "日志采集完整性、Trace ID 关联、保留策略", "8+", "P1", "—", "架构设计书 §13"),
    ("P-10", "Argo CD", "同步/审批/回滚", "3 维度", "GitOps 同步、审批闸门、回滚（git revert + 自动同步）", "10+", "P1", "回滚 < 5min", "架构设计书 §15/§1.1"),
    ("P-11", "E2E 业务流", "端到端", "12 核心场景", "E2E-①~⑫ 全场景跨服务全链路验证", "60+", "P0", "F1-F10/M3-M5", "需求定义书 §5/§3.5"),
]

r = 4
for it in items:
    for i, v in enumerate(it, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    pri = it[6]
    pc = ws.cell(r, 7)
    if pri == "P0": pc.fill = fill("F8CBAD")
    elif pri == "P1": pc.fill = fill("FFE699")
    elif pri == "P2": pc.fill = fill("C6E0B4")
    r += 1
grid(ws, 4, r - 1, 1, 9)
ws.freeze_panes = "A4"

# ---------------- Sheet 3: V字模型矩阵 ----------------
ws = wb.create_sheet("V字模型矩阵")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})
title_block(ws, "V 字模型 × 测试类型 覆盖矩阵（按 JIS X 0129）", 7)
headers = ["服务/对象", "单元测试\n（コンポーネント）", "集成测试\n（結合）", "契约测试", "系统测试\n（システム）", "验收测试\n（受入）", "数量级"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 7)

v_matrix = [
    ("S-01 auth-service",       "✓", "✓", "✓", "✓", "",  "100+"),
    ("S-02 user-service",       "✓", "✓", "✓", "✓", "",  "73+"),
    ("S-03 project-service",    "✓", "✓", "✓", "✓", "",  "113+"),
    ("S-04 task-service",       "✓", "✓", "✓", "✓", "",  "150+"),
    ("S-05 file-service",       "✓", "✓", "✓", "✓", "",  "69+"),
    ("S-06 notification-service","✓", "✓", "✓", "✓", "",  "53+"),
    ("S-07 report-service",     "✓", "✓", "✓", "✓", "",  "46+"),
    ("S-08 audit-service",      "✓", "✓", "✓", "✓", "",  "43+"),
    ("S-09 translation-core",   "✓", "✓", "✓", "✓", "",  "165+"),
    ("S-10 ingestion-service",  "✓", "✓", "",  "✓", "",  "30+"),
    ("S-11 asr-service",        "✓", "✓", "",  "✓", "",  "43+"),
    ("S-12 ocr-service",        "✓", "✓", "",  "✓", "",  "30+"),
    ("S-13 subtitle-service",   "✓", "✓", "",  "✓", "",  "35+"),
    ("S-14 office-converter",   "✓", "✓", "",  "✓", "",  "55+"),
    ("S-15 render-writer",      "✓", "✓", "",  "✓", "",  "40+"),
    ("S-16 worker-service",     "✓", "✓", "",  "✓", "",  "18+"),
    ("C-01 Tauri 客户端",        "✓", "✓", "",  "✓", "✓", "125+"),
    ("C-02 Next.js 控制台",      "✓", "✓", "",  "✓", "✓", "55+"),
    ("P-01 Envoy Gateway",      "",  "✓", "",  "✓", "",  "30+"),
    ("P-02 PostgreSQL",         "",  "✓", "",  "✓", "",  "25+"),
    ("P-03 Valkey",             "",  "✓", "",  "✓", "",  "10+"),
    ("P-04 Kafka",              "",  "✓", "",  "✓", "",  "20+"),
    ("P-05 Debezium CDC",       "",  "✓", "",  "✓", "",  "15+"),
    ("P-06 Harbor",             "",  "✓", "",  "✓", "",  "10+"),
    ("P-07 Prometheus+Grafana", "",  "✓", "",  "✓", "",  "15+"),
    ("P-08 OTel + Tempo",       "",  "✓", "",  "✓", "",  "10+"),
    ("P-09 Loki + Promtail",    "",  "✓", "",  "✓", "",  "8+"),
    ("P-10 Argo CD",            "",  "✓", "",  "✓", "",  "10+"),
    ("P-11 端到端 E2E",         "",  "",  "",  "✓", "✓", "60+"),
]
r = 4
for row in v_matrix:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = center() if i > 1 else left()
    r += 1
grid(ws, 4, r - 1, 1, 7)
ws.freeze_panes = "B4"

# ---------------- Sheet 4: 需求追溯矩阵 ----------------
ws = wb.create_sheet("需求追溯矩阵")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 8, "B": 16, "C": 40, "D": 14, "E": 24, "F": 14, "G": 24})
title_block(ws, "需求追溯矩阵（OFCAT 需求 F1-F11 + 架构原则 → 测试项）", 7)
headers = ["需求 ID", "需求名", "需求描述（简）", "涉及服务", "对应测试类型", "需求章节", "测试项位置"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 7)

matrix = [
    ("F1", "选区捕获", "获取用户在网页上选中的文本作为翻译对象", "Tauri 客户端、translation-core", "E2E + 单元 + 兼容性", "§5.2 F1", "测试项总览 C-01 + S-09"),
    ("F2", "TM 匹配", "精确匹配/模糊匹配", "project-service、translation-core", "单元 + 集成 + 性能", "§5.2 F2", "测试项总览 S-03/S-09"),
    ("F3", "术语匹配与注入", "术语抽取 + 提示词注入", "project-service、translation-core", "单元 + 集成 + 契约", "§5.2 F3", "测试项总览 S-03/S-09"),
    ("F4", "术语强制校验", "译后扫描 + 违规替换", "translation-core", "单元 + 合规", "§5.2 F4", "测试项总览 S-09 + 性能SLO"),
    ("F5", "标签/占位符保护", "哨兵替换 + 数量校验", "translation-core", "单元 + 集成", "§5.2 F5", "测试项总览 S-09"),
    ("F6", "单模型流式翻译", "默认 L2 路径", "translation-core、task-service", "单元 + 接口 + 性能", "§5.2 F6", "测试项总览 S-09/S-04 + 性能SLO L2"),
    ("F7", "行内 overlay 编辑", "可编辑/可确认", "Tauri 客户端、translation-core", "E2E + 单元", "§5.2 F7", "测试项总览 C-01 + S-09"),
    ("F8", "写回页面", "富文本安全写回", "Tauri 客户端", "E2E + 兼容性 + 安全", "§5.2 F8", "测试项总览 C-01"),
    ("F9", "保存进 TM", "回存本地 SQLite → PostgreSQL", "translation-core、project-service、Tauri", "E2E + 集成 + 数据一致性", "§5.2 F9", "测试项总览 S-09/S-03 + C-01 离线队列"),
    ("F10", "合规路由", "敏感强制本地", "translation-core、ai-gateway、project-service", "合规 + 安全 + 故障注入", "§5.2 F10", "测试项总览 S-09 + 风险与缓解 C-3"),
    ("F11", "存量数据导入", "清洗 + 映射 + 去重", "worker-service、project-service", "E2E + 集成（M2 阶段）", "§5.2 F11", "测试项总览 S-16/S-03"),
]

r = 4
for row in matrix:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 7)
ws.freeze_panes = "A4"

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c = ws.cell(r, 1, "架构原则 → 测试活动追溯（V 字模型右臂覆盖验证）")
c.fill = fill(NAVY); c.font = f(11, True, WHITE); c.alignment = center()
ws.row_dimensions[r].height = 22
r += 1
ph_headers = ["原则", "原则名", "原则说明", "测试活动", "测试类型/级别", "架构章节", "测试项位置"]
for i, h in enumerate(ph_headers, 1):
    ws.cell(r, i, h)
style_header(ws, r, 7)
r += 1
ph_rows = [
    ("P-1", "确定性优先", "TM/术语/标签/时间轴对齐等确定性逻辑用算法保证正确性", "F2/F4/F5 算法准确性、QA 一致性", "单元 + 集成 + 性能", "架构设计书 §1.2 原则 1", "测试项总览 S-09/S-03/S-13"),
    ("P-2", "数据存储单一权威", "PostgreSQL 唯一权威；Valkey 仅缓存；Kafka 仅事件", "双写一致性、跨服务最终一致、pgvector 与主表一致", "集成 + 数据一致性 + 故障注入", "架构设计书 §1.2 原则 2", "测试项总览 S-04/S-03 + 风险与缓解"),
    ("P-3", "合规 fail-closed", "敏感内容失败时中止而非降级泄露", "敏感项目全链路本地、合规测试、网抓验证零上云", "合规 + 安全 + 故障注入", "架构设计书 §1.2 原则 3", "测试项总览 S-09 + 风险与缓解 C-3"),
    ("P-4", "服务边界清晰", "按数据/媒体类型划分，不按团队组织架构", "媒体处理服务无独立库、跨服务数据访问禁止", "集成 + 安全", "架构设计书 §1.2 原则 4", "测试项总览 S-10~S-15"),
    ("P-5", "异步优先不滥用", "只在必要时引入异步", "同步/异步链路选择合理性、Trace ID 贯通", "接口 + 可观测性 + 性能", "架构设计书 §1.2 原则 5", "测试项总览 P-04/P-05/P-08"),
]
for row in ph_rows:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, r - len(ph_rows), r - 1, 1, 7)

# ---------------- Sheet 5: 服务×测试类型矩阵 ----------------
ws = wb.create_sheet("服务测试类型矩阵")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 8, "C": 8, "D": 8, "E": 8, "F": 8, "G": 8, "H": 8, "I": 8, "J": 8, "K": 8})
title_block(ws, "服务 × 测试类型 覆盖矩阵（含 JIS X 0129-4 测试技法）", 11)
headers = ["服务/对象", "单元", "集成", "契约", "接口", "E2E", "性能", "安全", "合规", "故障注入", "数量级"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 11)

matrix_rows = [
    ("S-01 auth-service",       "✓", "✓", "✓", "✓", "",  "",  "✓", "",  "",  "100+"),
    ("S-02 user-service",       "✓", "✓", "✓", "✓", "",  "",  "✓", "",  "",  "73+"),
    ("S-03 project-service",    "✓", "✓", "✓", "✓", "",  "✓", "✓", "",  "",  "113+"),
    ("S-04 task-service",       "✓", "✓", "✓", "✓", "",  "✓", "",  "",  "",  "150+"),
    ("S-05 file-service",       "✓", "✓", "✓", "✓", "",  "✓", "✓", "",  "",  "69+"),
    ("S-06 notification-service","✓", "✓", "✓", "✓", "",  "✓", "",  "",  "",  "53+"),
    ("S-07 report-service",     "✓", "✓", "✓", "✓", "",  "",  "",  "",  "",  "46+"),
    ("S-08 audit-service",      "✓", "✓", "✓", "✓", "",  "",  "",  "✓", "",  "43+"),
    ("S-09 translation-core",   "✓", "✓", "✓", "✓", "",  "✓", "",  "✓", "",  "165+"),
    ("S-10 ingestion-service",  "✓", "✓", "",  "✓", "",  "",  "",  "",  "",  "30+"),
    ("S-11 asr-service",        "✓", "✓", "",  "✓", "",  "✓", "",  "",  "✓", "43+"),
    ("S-12 ocr-service",        "✓", "✓", "",  "✓", "",  "✓", "",  "",  "",  "30+"),
    ("S-13 subtitle-service",   "✓", "✓", "",  "✓", "",  "",  "",  "",  "",  "35+"),
    ("S-14 office-converter",   "✓", "✓", "",  "✓", "",  "",  "",  "",  "✓", "55+"),
    ("S-15 render-writer",      "✓", "✓", "",  "✓", "",  "",  "",  "",  "✓", "40+"),
    ("S-16 worker-service",     "✓", "✓", "",  "",  "",  "",  "",  "",  "",  "18+"),
    ("C-01 Tauri 客户端",        "✓", "✓", "",  "",  "✓", "✓", "✓", "",  "✓", "125+"),
    ("C-02 Next.js 控制台",      "✓", "✓", "",  "",  "✓", "",  "✓", "",  "",  "55+"),
    ("P-01 Envoy Gateway",      "",  "✓", "",  "✓", "",  "✓", "✓", "",  "✓", "30+"),
    ("P-02 PostgreSQL",         "",  "✓", "",  "",  "",  "✓", "✓", "",  "✓", "25+"),
    ("P-03 Valkey",             "",  "✓", "",  "",  "",  "✓", "",  "",  "✓", "10+"),
    ("P-04 Kafka",              "",  "✓", "",  "",  "",  "✓", "✓", "",  "✓", "20+"),
    ("P-05 Debezium CDC",       "",  "✓", "",  "",  "",  "",  "",  "",  "✓", "15+"),
    ("P-06 Harbor",             "",  "✓", "",  "",  "",  "",  "✓", "",  "✓", "10+"),
    ("P-07 Prometheus+Grafana", "",  "✓", "",  "",  "",  "",  "",  "",  "",  "15+"),
    ("P-08 OTel + Tempo",       "",  "✓", "",  "",  "",  "",  "",  "",  "",  "10+"),
    ("P-09 Loki + Promtail",    "",  "✓", "",  "",  "",  "",  "",  "",  "",  "8+"),
    ("P-10 Argo CD",            "",  "✓", "",  "",  "",  "",  "",  "",  "✓", "10+"),
    ("P-11 端到端 E2E",         "",  "",  "",  "",  "✓", "✓", "",  "✓", "✓", "60+"),
]
r = 4
for row in matrix_rows:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = center() if i > 1 else left()
    r += 1
grid(ws, 4, r - 1, 1, 11)
ws.freeze_panes = "B4"

# ---------------- Sheet 6: 媒体类型E2E ----------------
ws = wb.create_sheet("媒体类型E2E")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 12, "B": 24, "C": 50, "D": 40, "E": 20})
title_block(ws, "媒体类型 → E2E 场景 → 涉及服务（V 字模型系统测试层）", 5)
headers = ["场景 ID", "媒体类型", "E2E 场景描述", "涉及服务链路", "对应里程碑"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)

e2e = [
    ("E2E-①", "text 网页", "用户在 Tauri 客户端打开网页 → 选中文本 → 翻译 → 确认 → 写回", "Tauri → Envoy → auth/project/task → translation-core (TM 命中优先) → 写回 (全部 16 个 + Tauri)", "M1"),
    ("E2E-②", "video", "上传视频 → ffprobe 探测 → ASR + 字幕轨 OCR → 字幕分段 → 翻译 → 字幕烧录 → 下载 mp4", "file → ingestion → asr(+ocr) → subtitle → translation → render-writer", "M5"),
    ("E2E-③", "audio", "上传音频 → ffprobe 探测 → ASR → 段落翻译 → 写回 MP3 元数据/SRT", "file → ingestion → asr → translation → render-writer", "M5"),
    ("E2E-④", "pdf 文字层", "上传 PDF → PyMuPDF 文字层抽取 → 段落翻译 → PyMuPDF+ReportLab 版面重排 → 下载", "file → ingestion → translation → render-writer (pdf_relayout)", "M3"),
    ("E2E-⑤", "pdf 扫描件", "上传 PDF → PaddleOCR 识别 → 段落翻译 → 版面重排 → 下载", "file → ingestion → ocr → translation → render-writer (pdf_relayout)", "M3"),
    ("E2E-⑥", "docx", "上传 docx → python-docx 段落抽取 → 翻译 → 段落回填 → 下载", "file → ingestion → office-converter → translation → office-converter (回填)", "M3"),
    ("E2E-⑦", "xlsx", "上传 xlsx → openpyxl 单元格抽取 → 翻译 → 单元格回填 → 下载", "file → ingestion → office-converter → translation → office-converter (回填)", "M3"),
    ("E2E-⑧", "pptx", "上传 pptx → python-pptx 文本框抽取 → 翻译 → 文本框回填 → 下载", "file → ingestion → office-converter → translation → office-converter (回填)", "M3"),
    ("E2E-⑨", "gif", "上传 GIF → ffmpeg 抽帧 → PaddleOCR 逐帧 → 翻译 → Pillow 重编码 → 下载", "file → ingestion → ocr (抽帧) → translation → render-writer (gif_reencode)", "M5"),
    ("E2E-⑩", "webp", "上传 WebP → ffmpeg 抽帧 → PaddleOCR 逐帧 → 翻译 → WebP 重编码 → 下载", "file → ingestion → ocr (抽帧) → translation → render-writer (webp_reencode)", "M5"),
    ("E2E-⑪", "offline-recovery", "Tauri 客户端断网 → 编辑 → 恢复联网 → 离线队列自动重放", "Tauri 离线队列 → 后端（自动重放）", "M2"),
    ("E2E-⑫", "compliance-sensitive", "敏感项目请求 → 全程本地 LLM → 本地不可达时 fail-closed → 网抓验证零上云", "Tauri → auth/project (policy) → translation-core (合规) → ai-gateway (本地)", "M1/M5"),
]
r = 4
for row in e2e:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 5)
ws.freeze_panes = "A4"

# ---------------- Sheet 7: 测试环境与工具 ----------------
ws = wb.create_sheet("测试环境与工具")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 24, "C": 36, "D": 24})
title_block(ws, "测试环境分层 + 工具一览（含 JIS X 0129-2 选型原则）", 4)
headers = ["类别", "名称", "用途", "备注"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
env_rows = [
    ("环境分层", "DEV 开发者本机", "单元测试、本地集成测试", "mock + 测试 fixtures，单机"),
    ("环境分层", "CI Runner", "PR 门禁：单测 + 契约 + 轻量集成", "容器化 GitOps 流水线"),
    ("环境分层", "STG テスト環境", "完整集成 + 接口 + E2E + 性能基线", "1 控制面 + 2 工作节点；与生产 1:1 缩放"),
    ("环境分层", "PRE 预生产", "验收 + 性能压测 + 故障注入 + DR 演练", "3 控制面 + N 工作节点（与生产同规模）"),
    ("环境分层", "PROD 生产", "灰度发布验证 + 监控告警验证", "3 控制面 + N 工作节点；仅 release tag 触发"),
    ("单测/集成", "cargo test + cargo-tarpaulin", "Rust 单测 + 覆盖率", "覆盖率 ≥ 80%"),
    ("单测/集成", "pytest + pytest-cov", "Python 单测 + 覆盖率", "覆盖率 ≥ 80%"),
    ("单测/集成", "vitest + @vitest/coverage-v8", "TS/Svelte 单测 + 覆盖率", "覆盖率 ≥ 70%"),
    ("单测/集成", "testcontainers-python/rust", "起 PG/Redis/Kafka 临时实例", "集成测试"),
    ("变异测试", "mutmut（Python）", "变异测试验证测试有效性", "可选；M3 后启用"),
    ("变异测试", "cargo-mutants（Rust）", "变异测试验证测试有效性", "可选；M3 后启用"),
    ("契约", "pact (Python/Rust/JS) + pact-provider-verifier", "OpenAPI 双向契约", "Consumer + Provider"),
    ("契约", "buf lint + buf breaking", "Protobuf 兼容性门禁", "CI 阻断"),
    ("契约", "JSON Schema 校验（Kafka Event）", "Event schema 演进", "Topic < 10 用 JSON Schema 文件"),
    ("接口/E2E", "tauri-driver", "Tauri WebDriver 协议", "客户端 E2E"),
    ("接口/E2E", "playwright", "Next.js 浏览器自动化", "控制台 E2E"),
    ("接口/E2E", "pytest 集成测试套件", "全链路后端 E2E", "部署到 K3s cats-test Namespace"),
    ("性能", "k6", "HTTP/WS 压测", "p95/p99 延迟统计"),
    ("性能", "ghz", "gRPC 基准", "translation-core 流式"),
    ("性能", "wrk2", "高精度延迟", "—"),
    ("性能", "kafka-producer-perf-test.sh", "Kafka 吞吐基线", "—"),
    ("安全/合规", "semgrep", "SAST（Python/Rust/TS）", "PR 门禁"),
    ("安全/合规", "bandit / cargo-audit / pip-audit / npm audit", "依赖与代码安全", "—"),
    ("安全/合规", "syft + grype", "SBOM 生成 + 漏洞扫描", "CI 阻断已知严重"),
    ("安全/合规", "gitleaks", "密钥扫描", "历史+当前"),
    ("安全/合规", "OWASP ZAP", "DAST baseline", "Next.js 公开面"),
    ("安全/合规", "mitmproxy", "合规抓包验证", "敏感内容零上云"),
    ("安全/合规", "jwt_tool", "JWT 签名/算法替换攻击验证", "手工"),
    ("故障注入", "chaos-mesh", "K3s 内 Pod/网络/IO 故障", "推荐 K8s 原生"),
    ("故障注入", "tc netem", "网络延迟/丢包/分区", "手动"),
    ("故障注入", "kubectl / crictl", "Pod/容器故障", "—"),
    ("可观测性", "Prometheus + Grafana + Alertmanager", "指标 + 仪表板 + 告警", "—"),
    ("可观测性", "OpenTelemetry Collector + Tempo", "链路追踪", "贯穿同步+异步"),
    ("可观测性", "Loki + Promtail + Grafana Explore", "日志检索 + Trace ID 关联", "—"),
    ("可观测性", "k9s / stern / kubectl logs", "K3s 实时运维", "—"),
]
r = 4
for row in env_rows:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)
ws.freeze_panes = "A4"

# ---------------- Sheet 8: 缺陷等级与SLA ----------------
ws = wb.create_sheet("缺陷等级与SLA")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 8, "B": 14, "C": 50, "D": 30, "E": 12})
title_block(ws, "缺陷等级定义与 SLA（按 JIS X 0129-1 缺陷管理实践）", 5)
headers = ["等级", "名称", "定义", "示例", "修复 SLA"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
sev = [
    ("P0", "阻断级（致命）", "核心功能不可用 / 数据丢失 / 安全漏洞", "TM 匹配崩溃、合规内容上云、PG 主从不一致、登录全挂", "24h"),
    ("P1", "严重级", "主要功能不可用但有 workaround", "某媒体类型翻译链路全断、性能 SLO 偏差 > 50%", "3d"),
    ("P2", "一般级", "次要功能异常", "单个 endpoint 报错、UI 某浏览器某页错位、告警文案不准确", "2w"),
    ("P3", "轻微级", "体验/文案问题", "提示文案不通顺、图标错误、文档错误", "下迭代"),
    ("P4", "建议级", "优化建议", "性能可优化点、可读性", "评估"),
]
r = 4
for row in sev:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    pri = ws.cell(r, 1)
    if pri.value == "P0": pri.fill = fill("F8CBAD")
    elif pri.value == "P1": pri.fill = fill("FFD966")
    elif pri.value == "P2": pri.fill = fill("C6E0B4")
    elif pri.value == "P3": pri.fill = fill("BDD7EE")
    elif pri.value == "P4": pri.fill = fill("D9D9D9")
    r += 1
grid(ws, 4, r - 1, 1, 5)

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, "缺陷度量指标（含缺陷注入率与去除率）")
c.fill = fill(NAVY); c.font = f(11, True, WHITE); c.alignment = center()
ws.row_dimensions[r].height = 22
r += 1
m_headers = ["指标", "计算方式", "目标", "数据来源", "统计周期"]
for i, h in enumerate(m_headers, 1):
    ws.cell(r, i, h)
style_header(ws, r, 5)
r += 1
metrics = [
    ("缺陷密度（欠陥密度）", "缺陷数 / 千行代码", "< 1.0（按服务）", "SonarQube/Cobertura + 缺陷库", "每迭代"),
    ("缺陷去除率（欠陥除去率）", "各阶段发现/总注入", "≥ 99%（4 阶段累计，见 §10.6）", "缺陷库 + 注入记录", "每里程碑"),
    ("缺陷注入率（欠陥挿入率）", "故意注入/总代码量", "核心 5‰/一般 3‰/客户端 2‰", "变异测试工具", "每里程碑"),
    ("变异得分（ミューテーションスコア）", "被检出变异体/总变异体", "≥ 缺陷发现率目标（90/80/70%）", "mutmut/cargo-mutants", "每里程碑"),
    ("P0 缺陷遗留", "当前未关闭的 P0 数", "= 0", "缺陷库", "实时"),
    ("平均修复时长", "新建→关闭", "P0 < 1d / P1 < 3d / P2 < 14d", "缺陷库", "每周"),
    ("缺陷逃逸率", "线上发现 / 总缺陷", "< 5%", "缺陷库", "每迭代"),
    ("重打开率", "重打开数 / 总关闭数", "< 10%", "缺陷库", "每迭代"),
    ("测试用例执行率", "已执行 / 总用例", "100%", "测试管理工具", "每里程碑"),
    ("测试通过率", "通过 / 已执行", "> 95%", "测试管理工具", "每里程碑"),
    ("缺陷平均定位时间", "发现到定位根因", "< 30 min", "缺陷库", "每迭代"),
]
for row in metrics:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, r - len(metrics), r - 1, 1, 5)

# ---------------- Sheet 9: 性能SLO ----------------
ws = wb.create_sheet("性能SLO")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 14, "B": 26, "C": 20, "D": 18, "E": 30})
title_block(ws, "性能 SLO（初版，待 PRE 校准；按 IPA 非機能要求グレード）", 5)
headers = ["类别", "场景/层级", "指标", "目标", "测试方法"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
slo = [
    ("翻訳遅延", "L0 TM 100% 匹配", "端到端 p95", "< 100ms", "k6 + ghz，10k 样本"),
    ("翻訳遅延", "L1 TM 模糊匹配", "端到端 p95", "< 300ms", "k6 + ghz"),
    ("翻訳遅延", "L2 单模型流式", "首字 p95", "< 1s", "k6 流式 + SSE 截首字"),
    ("翻訳遅延", "L2 单模型流式", "完成 p95", "< 3s", "k6 流式"),
    ("翻訳遅延", "L3 多模型投票", "完成 p95", "< 10s", "k6 多模型并行"),
    ("翻訳遅延", "L0 精确匹配不调云端", "网络抓包", "0 次云端 API 调用", "mitmproxy 验证"),
    ("翻訳遅延", "L1 模糊匹配不调云端", "网络抓包", "0 次云端 API 调用（语义向量查询在本地）", "mitmproxy 验证"),
    ("負荷テスト", "500 并发用户", "核心服务错误率", "< 0.1%", "k6 混合负载 30min"),
    ("負荷テスト", "1000 并发用户", "核心服务错误率", "< 1%", "k6 混合负载 30min"),
    ("ストレステスト", "超出 1500 并发", "降级而非崩溃", "请求拒绝而非 5xx", "k6 阶梯压测"),
    ("スパイクテスト", "瞬时并发翻倍", "无雪崩", "RPS 不归零", "k6 spike profile"),
    ("ソークテスト", "24h 持续", "无内存泄漏", "RSS 稳定 ±10%", "k6 sustained + pprof"),
    ("容量テスト", "视频任务（30min 1080p）", "端到端 p95", "< 5min", "E2E-② 全链路"),
    ("容量テスト", "PDF 任务（200 页）", "端到端 p95", "< 2min", "E2E-④ 全链路"),
    ("容量テスト", "音频任务（30min）", "端到端 p95", "< 2min", "E2E-③ 全链路"),
    ("容量テスト", "GIF（50 帧）", "端到端 p95", "< 30s", "E2E-⑨ 全链路"),
    ("可用性", "全平台 SLO 月度", "可用性", "≥ 99.5%", "Prometheus + 错误率统计"),
    ("数据库", "PostgreSQL 主从延迟", "replica lag p95", "< 1s", "pg_stat_replication"),
    ("数据库", "PostgreSQL 主节点 failover", "自动恢复时长", "< 30s", "手动 pg_ctl stop 演练"),
    ("消息队列", "Kafka consumer lag", "告警阈值", "< 10000 条", "Prometheus + Alertmanager"),
    ("消息队列", "Debezium 复制槽推进", "LSN 推进延迟", "< 5s", "pg_replication_slots"),
    ("可观测性", "Trace 同步+异步贯通率", "贯通率", "100% 关键 E2E 场景", "Tempo 抽样 100 个 E2E"),
    ("客户端", "Tauri 首屏", "WebView 可交互 p95", "< 2s", "tauri-driver"),
    ("客户端", "Tauri 离线队列重放", "100 条重放 p95", "< 30s", "tauri-driver + 网络模拟"),
    ("通知", "WebSocket 推送延迟", "端到端 p95", "< 1s", "playwright + 10000 并发"),
]
r = 4
for row in slo:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 5)
ws.freeze_panes = "A4"

# ---------------- Sheet 10: 覆盖率与缺陷注入率 ----------------
ws = wb.create_sheet("覆盖率与缺陷注入率")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 24, "B": 14, "C": 14, "D": 14, "E": 14, "F": 16, "G": 18})
title_block(ws, "覆盖率基线 + 缺陷注入率（按 JIS X 0129-2 测试充分性）", 7)
headers = ["服务/对象", "语句覆盖\n（Stmts）", "分支覆盖\n（Branch）", "MC/DC\n（关键判定）", "缺陷注入率", "变异得分目标", "备注"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 7)
cov = [
    ("S-01 auth-service",       "≥ 85%", "≥ 75%", "≥ 80%", "5‰", "≥ 90%", "安全敏感，基线最高"),
    ("S-02 user-service",       "≥ 80%", "≥ 70%", "≥ 70%", "3‰", "≥ 80%", "业务规则较多"),
    ("S-03 project-service",    "≥ 85%", "≥ 75%", "≥ 80%", "5‰", "≥ 90%", "TM/术语算法核心"),
    ("S-04 task-service",       "≥ 85%", "≥ 75%", "≥ 80%", "5‰", "≥ 90%", "状态机核心"),
    ("S-05 file-service",       "≥ 80%", "≥ 70%", "≥ 70%", "3‰", "≥ 80%", "—"),
    ("S-06 notification-service","≥ 75%", "≥ 65%", "≥ 70%", "3‰", "≥ 80%", "—"),
    ("S-07 report-service",     "≥ 80%", "≥ 70%", "≥ 70%", "3‰", "≥ 80%", "统计准确性"),
    ("S-08 audit-service",      "≥ 85%", "≥ 75%", "≥ 80%", "5‰", "≥ 90%", "合规追溯"),
    ("S-09 translation-core",   "≥ 90%", "≥ 80%", "≥ 85%", "5‰", "≥ 90%", "算法核心，基线最高"),
    ("S-10 ingestion-service",  "≥ 80%", "≥ 70%", "≥ 70%", "3‰", "≥ 80%", "—"),
    ("S-11 asr-service",        "≥ 80%", "≥ 70%", "≥ 70%", "3‰", "≥ 80%", "—"),
    ("S-12 ocr-service",        "≥ 80%", "≥ 70%", "≥ 70%", "3‰", "≥ 80%", "—"),
    ("S-13 subtitle-service",   "≥ 85%", "≥ 75%", "≥ 80%", "3‰", "≥ 80%", "时间轴正确性"),
    ("S-14 office-converter",   "≥ 80%", "≥ 70%", "≥ 70%", "3‰", "≥ 80%", "—"),
    ("S-15 render-writer",      "≥ 85%", "≥ 75%", "≥ 80%", "3‰", "≥ 80%", "渲染正确性"),
    ("S-16 worker-service",     "≥ 80%", "≥ 70%", "≥ 70%", "3‰", "≥ 80%", "—"),
    ("C-01 Tauri（Rust 核心）",   "≥ 75%", "≥ 65%", "—",    "2‰", "≥ 70%", "—"),
    ("C-01 Tauri（Svelte 前端）", "≥ 65%", "≥ 55%", "—",    "2‰", "≥ 70%", "—"),
    ("C-02 Next.js 控制台",      "≥ 70%", "≥ 60%", "—",    "2‰", "≥ 70%", "—"),
]
r = 4
for row in cov:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 7)
ws.freeze_panes = "A4"

# 缺陷去除率表
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c = ws.cell(r, 1, "缺陷去除率（欠陥除去率）阶段目标")
c.fill = fill(NAVY); c.font = f(11, True, WHITE); c.alignment = center()
ws.row_dimensions[r].height = 22
r += 1
d_headers = ["阶段", "目标", "累计", "测试活动", "测量方式", "—", "—"]
for i, h in enumerate(d_headers, 1):
    ws.cell(r, i, h)
style_header(ws, r, 7)
r += 1
dre = [
    ("单元测试",   "≥ 60%", "60%",   "函数/方法/类、状态机、纯算法", "覆盖率 + 变异得分", "—", "—"),
    ("集成测试",   "≥ 20%", "80%",   "服务间契约、数据流、状态机", "契约测试 + 数据一致性测试", "—", "—"),
    ("系统测试",   "≥ 15%", "95%",   "E2E 业务流、性能、可靠性、安全", "E2E + 性能 + 故障注入", "—", "—"),
    ("验收测试",   "≥ 4%",  "99%",   "业务场景、效果指标", "UAT + KPI 验证", "—", "—"),
    ("交付后",     "< 1%",  "—",     "生产监控", "线上缺陷跟踪", "—", "—"),
]
for row in dre:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, r - len(dre), r - 1, 1, 7)

# ---------------- Sheet 11: 风险与缓解 ----------------
ws = wb.create_sheet("风险与缓解")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 10, "B": 30, "C": 8, "D": 8, "E": 50, "F": 30})
title_block(ws, "测试风险登记册 + 缓解措施（按 JIS X 0129-2 リスクベースドテスト）", 6)
headers = ["编号", "风险", "概率", "影响", "缓解措施", "关联架构风险/章节"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 6)
risks = [
    ("R-06", "Tauri WebView2 依赖（Windows 离线首次运行）", "中", "高", "打包内嵌 WebView2 Evergreen Bootstrapper 离线安装包，STG 兼容性测试覆盖 Windows 多版本", "架构设计书 R-06"),
    ("R-07", "K3s 控制面节点故障域集中", "高", "高", "3 控制面节点跨物理机/机架部署，PD B + 反亲和性；故障注入演练 1/3 节点宕机", "架构设计书 R-07"),
    ("R-08", "Kafka 复制槽（Debezium）积压拖垮 PostgreSQL WAL", "高", "高", "监控复制槽 lag 告警；性能测试模拟 Debezium 暂停；HA：WAL 保留上限 + 紧急手动清理", "架构设计书 R-08"),
    ("R-09", "LibreOffice Headless 进程僵死/内存泄漏", "中", "中", "office-converter 进程池超时熔断 + 进程僵死 kill 演练；定期 worker 进程回收", "技术选型书 R-09"),
    ("R-10", "faster-whisper GPU 资源争抢", "中", "中", "K3s 节点打 GPU 污点+资源配额；ASR 高并发 + CUDA_OOM 降级到 CPU 演练", "技术选型书 R-10"),
    ("R-11", "Harbor 单点故障阻断全部镜像拉取", "高", "中", "Harbor HA 部署（多副本+外置对象存储）+ 节点本地镜像缓存；Harbor 不可用演练", "架构设计书 R-11 / 技术选型 R-11"),
    ("R-12", "微服务数量膨胀导致运维复杂度失控", "中", "中", "新服务接入对整体链路 Trace 完整性、性能、监控覆盖度回归测试", "架构设计书 R-12"),
    ("T-01", "K3s 测试集群搭建复杂、CI 全链路 E2E 慢", "高", "高", "用 k3d（K3s in Docker）起本地集群；E2E 拆'快速套件'（5 场景/15min）+'完整套件'（11 场景/60min）", "—"),
    ("T-02", "测试数据准备成本高（媒体样本大）", "中", "中", "媒体样本分级：核心样本（小，1MB 内，常驻） + 全量样本（大，按需下载）", "—"),
    ("T-03", "Kafka/PostgreSQL 容器化测试不稳定（flaky）", "中", "中", "用 cloudnative-pg 与 kraft-mode kafka 官方 Docker 镜像；CI 固定重试 1 次 + flaky 失败需人工排查", "—"),
    ("T-04", "性能测试结果难复现", "中", "中", "固定硬件（CI Runner 标定规格）；多次跑取 p50；环境基线监控", "—"),
    ("T-05", "故障注入风险（误伤 STG/PRE）", "中", "高", "故障注入**仅在 PRE 环境**；PRE 与 STG 物理隔离；故障注入工具使用 RBAC 限制", "—"),
    ("T-06", "合规测试的本地 LLM 资源占用", "中", "中", "STG 单独部署小型本地 LLM（量化版 Qwen 7B），与压测环境错峰跑", "R-10"),
    ("T-07", "跨团队服务的契约测试协调", "中", "中", "CI 强制门禁 + 契约变更需双方开发同 PR review + 接口变更公告频道", "—"),
    ("T-08", "自动化测试自身维护成本", "中", "中", "每个 sprint 拨 5% 工时做'测试卫生'（重跑失败标记 flaky、清理废弃用例、更新测试数据）", "—"),
    ("T-09", "Tauri 客户端 E2E 工具不成熟（tauri-driver 较新）", "高", "中", "优先用 WebDriver 协议覆盖核心路径；UI 细节测试用 Svelte 组件测试替代", "R-06"),
    ("C-1", "P0 缺陷带病合并", "中", "高", "P0 阻断 CI 合并门禁；缺陷库与 PR 状态联动", "—"),
    ("C-2", "故障注入演练误伤生产", "低", "高", "故障注入仅 PRE；K3s RBAC 限制 chaos-mesh 仅 namespace 内；演练前通知", "—"),
    ("C-3", "合规 fail-closed 失效（敏感内容上云）", "低", "极高", "P0 阻断级合规测试（翻译+网络抓包双重验证）；月度合规专项演练；变异测试对 fail-closed 路径", "架构设计书 §1.2 原则 3"),
]
r = 4
for row in risks:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    for col in (3, 4):
        v = ws.cell(r, col).value
        if v == "高" or v == "极高":
            ws.cell(r, col).fill = fill("F8CBAD")
        elif v == "中":
            ws.cell(r, col).fill = fill("FFE699")
        elif v == "低":
            ws.cell(r, col).fill = fill("C6E0B4")
    r += 1
grid(ws, 4, r - 1, 1, 6)
ws.freeze_panes = "A4"

# ---------------- Sheet 12: RACI矩阵 ----------------
ws = wb.create_sheet("RACI矩阵")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 28, "B": 14, "C": 14, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12})
title_block(ws, "测试活动 RACI 矩阵（R=执行 A=问责 C=咨询 I=知情）", 8)
headers = ["活动", "测试负责人", "测试工程师", "开发", "架构", "SRE", "产品", "合规"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 8)
raci = [
    ("测试设计书编写",        "R/A", "C",  "C", "C", "C", "I", "I"),
    ("测试用例编写",          "A",   "R",  "C", "I", "I", "C", "I"),
    ("单元测试实施",          "C",   "I",  "R/A","I", "I", "I", "I"),
    ("集成测试实施",          "A",   "R",  "C", "I", "I", "I", "I"),
    ("系统测试实施",          "A",   "R",  "C", "C", "C", "I", "I"),
    ("验收测试实施",          "C",   "C",  "I", "I", "I", "R/A","C"),
    ("合规测试",              "A",   "R",  "C", "I", "I", "C", "R/A"),
    ("缺陷修复",              "I",   "C",  "R/A","C", "I", "I", "I"),
    ("缺陷验证",              "A",   "R",  "C", "I", "I", "I", "I"),
    ("发布准出签字",          "R/A", "C",  "C", "C", "C", "A", "C"),
    ("故障注入演练",          "A",   "R",  "C", "C", "R", "I", "I"),
    ("DR 演练",               "C",   "C",  "I", "I", "R/A","I", "I"),
    ("CI 流水线维护",         "A",   "R",  "C", "C", "C", "I", "I"),
    ("测试数据准备",          "A",   "R",  "C", "I", "C", "I", "C"),
    ("覆盖率与变异报告",      "A",   "R",  "C", "C", "I", "I", "I"),
    ("测试结果报告（TM）",    "R/A", "R",  "C", "C", "C", "C", "C"),
    ("UAT 效果指标验证",      "C",   "C",  "I", "I", "I", "R/A","C"),
]
r = 4
for row in raci:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = center() if i > 1 else left()
        # 染色
        if "R/A" in v:
            cell.fill = fill("F8CBAD")  # 强
        elif v == "A":
            cell.fill = fill("FFD966")  # 中
        elif v == "R":
            cell.fill = fill("C6E0B4")  # 弱
    r += 1
grid(ws, 4, r - 1, 1, 8)
ws.freeze_panes = "B4"

# ---------------- Sheet 13: 里程碑 ----------------
ws = wb.create_sheet("里程碑")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 12, "B": 24, "C": 24, "D": 50, "E": 40})
title_block(ws, "测试里程碑 × 项目里程碑", 5)
headers = ["项目里程碑", "测试里程碑", "主要测试活动", "出口准则", "关联需求/架构"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
ms = [
    ("M1 — MVP 闭环", "TM-1 MVP 验证", "E2E-① 文本选区翻译 + F1-F10 全部功能/接口/性能/合规测试", "§11.2 M1 出口准则全达成", "需求定义书 §10 M1 / 架构设计书 §18"),
    ("M2 — 数据底座", "TM-2 数据底座验证", "F11 存量导入测试 + TM 百万级性能基线 + 离线/同步测试 + S-16 worker-service 全部", "§11.2 M2 出口准则全达成 + E2E-⑪ 离线恢复通过", "需求定义书 §10 M2"),
    ("M3 — 文档场景", "TM-3 文档场景验证", "E2E-④⑤⑥⑦⑧ PDF/Office 全场景 + 跨语言字幕", "§11.2 M3 出口准则全达成 + S-14 office-converter 格式覆盖完成", "需求定义书 §10 M3"),
    ("M4 — 工作流场景", "TM-4 工作流验证", "Jira 模式（范围外延）/ 文件夹自动导入", "§11.2 M4 出口准则全达成", "需求定义书 §10 M4"),
    ("M5 — 增强能力", "TM-5 增强能力验证", "E2E-②③⑨⑩ 视频/音频/动图、OCR、多模型投票、本地 LLM、文档级实体记忆", "§11.2 M5 出口准则全达成 + L3 性能达标", "需求定义书 §10 M5"),
    ("持续", "TM-Q 每季度", "DR 演练 + 故障注入全套 + 性能复测 + 安全复测 + 缺陷注入率验证", "全部 P2 阻断项通过", "—"),
    ("持续", "TM-M 每月", "合规专项演练（含网抓验证）", "敏感内容上云测试 100% 阻断", "架构设计书 §1.2 原则 3"),
]
r = 4
for row in ms:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 5)
ws.freeze_panes = "A4"

# ---------------- Sheet 14: 交付物清单 ----------------
ws = wb.create_sheet("交付物清单")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 8, "B": 28, "C": 40, "D": 18})
title_block(ws, "测试交付物清单（含 JIS X 0129-3 测试文档）", 4)
headers = ["编号", "交付物", "路径/工具", "时机"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
dels = [
    ("D-01", "测试设计书（Markdown）", "doc/04-测试/测试设计书/CATs_测试设计书_v1.0.md", "评审前 3 天"),
    ("D-02", "测试设计书（Excel）", "doc/04-测试/测试设计书/CATs_测试设计书_v1.0.xlsx", "评审前 3 天"),
    ("D-03", "构建脚本", "doc/04-测试/测试设计书/_build/build_xlsx.py", "与 D-02 同步"),
    ("D-04", "测试用例集（テストケース仕様書）", "doc/04-测试/测试用例/CATs_测试用例_<服务/模块>_v1.0.{md,xlsx}", "各 TM 里程碑前 2 周开始编写，TM 前 1 周冻结"),
    ("D-05", "测试结果报告（テスト結果報告）", "doc/04-测试/测试报告/CATs_测试报告_<TM>_v1.0.md", "每个 TM 结束"),
    ("D-06", "测试日志（テストログ）", "doc/04-测试/测试日志/CATs_测试日志_<日期>.log", "每次测试执行"),
    ("D-07", "自动化测试代码", "各服务仓库 tests/ 目录", "持续维护"),
    ("D-08", "CI 流水线配置", "仓库 .github/workflows/ 或 Gitea Actions / Jenkinsfile", "M1 前就绪"),
    ("D-09", "测试环境部署脚本", "Helm/Argo CD Application 配置（GitOps）", "STG/PRE 环境就绪"),
    ("D-10", "测试数据集", "内部测试素材库（Git LFS 或对象存储）", "STG 就绪前"),
    ("D-11", "缺陷跟踪看板", "与项目管理工具对接", "持续维护"),
    ("D-12", "覆盖率报告", "CI 仪表板（Cobertura/Grafana）", "持续"),
    ("D-13", "DR 演练报告", "doc/04-测试/测试报告/DR_演练_<日期>.md", "每季度"),
    ("D-14", "变异测试/缺陷注入率验证报告", "doc/04-测试/测试报告/变异测试_<服务>_<日期>.md", "每里程碑"),
]
r = 4
for row in dels:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ---------------- Sheet 15: 文档模板 ----------------
ws = wb.create_sheet("文档模板")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 20, "B": 80})
title_block(ws, "JIS X 0129-3 测试文档模板（快速参考）", 2)
headers = ["文档类型", "模板内容"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 2)
tpls = [
    ("テストケース仕様書\n（Test Case Specification）",
     "# 测试用例规格书 - <服务/模块名>\n\n"
     "## 标识\n"
     "- 编号 / 关联需求 (F1-F11) / 关联测试项 / 作者/日期 / 状态\n\n"
     "## 用例规格\n"
     "- 测试项标识 / 输入说明（含等价类/边界值）/ 输出说明\n"
     "- 环境需求 / 依赖（前置用例/数据）\n"
     "- 步骤 / 预期结果\n\n"
     "## 追踪\n"
     "- 关联缺陷 / 关联风险"),
    ("テスト結果報告\n（Test Result Report）",
     "# 测试结果报告 - <TM 里程碑>\n\n"
     "## 摘要\n"
     "- 测试周期 / 范围 / 通过-失败-阻塞数\n"
     "- 缺陷总数 / 覆盖率达成 / 缺陷注入率达成 / 出口准则达成度\n\n"
     "## 详细结果\n"
     "- 各服务测试结果 / 关键缺陷 / SLO 达成 / 故障注入演练结果\n\n"
     "## 风险与建议\n"
     "- 遗留风险 / 改进建议\n\n"
     "## 追溯\n"
     "- 关联需求 / 关联缺陷 / 关联风险"),
    ("テストログ\n（Test Log）",
     "时间戳 | 测试 ID | 步骤 | 输入 | 预期 | 实际 | 状态 | 缺陷 ID | Trace ID\n\n"
     "（CSV/Markdown 表格形式，CI 自动化生成）"),
    ("テスト手順仕様書\n（Test Procedure Specification）",
     "# 测试程序规格书 - <场景>\n\n"
     "## 前置条件 / 测试数据准备 / 测试环境 / 操作步骤（含截图）\n\n"
     "## 预期与判定 / 异常处理 / 清理步骤"),
    ("欠陥報告\n（Defect Report）",
     "# 缺陷报告 - <缺陷 ID>\n\n"
     "## 基本信息\n"
     "- 标题 / 等级 (P0-P4) / 状态 / 报告人 / 报告日期\n"
     "- 关联需求 (F1-F11) / 关联测试项 / 关联风险\n\n"
     "## 复现\n"
     "- 环境 / 前置条件 / 操作步骤 / 预期 / 实际 / Trace ID\n\n"
     "## 分析\n"
     "- 根因 / 影响范围 / 修复方案\n\n"
     "## 修复\n"
     "- 修复人 / 修复日期 / 修复 commit / 验证人 / 验证日期"),
]
r = 4
for k, v in tpls:
    ws.cell(r, 1, k).alignment = topleft()
    ws.cell(r, 1).fill = fill(LIGHT)
    ws.cell(r, 1).font = f(10, True)
    ws.cell(r, 2, v).alignment = topleft()
    ws.cell(r, 2).font = f(9)
    ws.row_dimensions[r].height = 200
    r += 1
grid(ws, 4, r - 1, 1, 2)
ws.freeze_panes = "A4"

# ---------------- Sheet 16: 术语集 ----------------
ws = wb.create_sheet("术语集")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 18, "C": 18, "D": 40})
title_block(ws, "JIS X 0129 / IPA 术语集（中日英对照）", 4)
headers = ["分类", "中文", "日本語", "English"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
glos = [
    ("测试级别", "单元测试", "単体テスト / コンポーネントテスト", "Unit Test / Component Test"),
    ("测试级别", "集成测试", "結合テスト / 統合テスト", "Integration Test"),
    ("测试级别", "系统测试", "システムテスト", "System Test"),
    ("测试级别", "验收测试", "受入テスト / 受け入れテスト", "Acceptance Test"),
    ("测试级别", "回归测试", "回帰テスト", "Regression Test"),
    ("测试级别", "冒烟测试", "スモークテスト", "Smoke Test"),
    ("测试级别", "健全性测试", "サニティテスト", "Sanity Test"),
    ("测试类型", "功能测试", "機能テスト", "Functional Test"),
    ("测试类型", "性能测试", "性能テスト", "Performance Test"),
    ("测试类型", "负荷测试", "負荷テスト", "Load Test"),
    ("测试类型", "压力测试", "ストレステスト", "Stress Test"),
    ("测试类型", "容量测试", "容量テスト", "Volume Test"),
    ("测试类型", "尖峰测试", "スパイクテスト", "Spike Test"),
    ("测试类型", "浸泡测试", "ソークテスト", "Soak Test"),
    ("测试类型", "可靠性测试", "信頼性テスト", "Reliability Test"),
    ("测试类型", "安全测试", "セキュリティテスト", "Security Test"),
    ("测试类型", "兼容性测试", "互換性テスト", "Compatibility Test"),
    ("测试类型", "探索性测试", "探索的テスト", "Exploratory Test"),
    ("测试类型", "故障注入测试", "フォールト挿入テスト", "Fault Injection Test"),
    ("测试类型", "变异测试", "ミューテーションテスト", "Mutation Test"),
    ("测试类型", "契约测试", "契約テスト", "Contract Test"),
    ("测试类型", "合规测试", "コンプライアンステスト", "Compliance Test"),
    ("测试技法", "等价类划分", "同値分割", "Equivalence Partitioning"),
    ("测试技法", "边界值分析", "境界値分析", "Boundary Value Analysis"),
    ("测试技法", "判定表", "デシジョンテーブル", "Decision Table"),
    ("测试技法", "状态迁移测试", "状態遷移テスト", "State Transition Testing"),
    ("测试技法", "用例测试", "ユースケーステスト", "Use Case Testing"),
    ("测试技法", "全对组合", "ペアワイズ", "Pairwise"),
    ("测试技法", "语句覆盖", "ステートメントカバレッジ", "Statement Coverage"),
    ("测试技法", "分支覆盖", "ブランチカバレッジ", "Branch Coverage"),
    ("测试技法", "条件覆盖", "コンディションカバレッジ", "Condition Coverage"),
    ("测试技法", "MC/DC", "MC/DC", "Modified Condition/Decision Coverage"),
    ("测试技法", "错误推测", "エラー推測", "Error Guessing"),
    ("测试技法", "检查表", "チェックリスト", "Checklist"),
    ("测试文档", "测试方针", "テスト方針", "Test Policy"),
    ("测试文档", "测试计划", "テスト計画書", "Test Plan"),
    ("测试文档", "测试设计书", "テスト設計書", "Test Design Specification"),
    ("测试文档", "测试用例规格书", "テストケース仕様書", "Test Case Specification"),
    ("测试文档", "测试程序规格书", "テスト手順仕様書", "Test Procedure Specification"),
    ("测试文档", "测试日志", "テストログ", "Test Log"),
    ("测试文档", "测试结果报告", "テスト結果報告", "Test Result Report"),
    ("测试文档", "缺陷报告", "不具合報告 / 欠陥報告", "Defect Report"),
    ("质量度量", "缺陷密度", "欠陥密度", "Defect Density"),
    ("质量度量", "缺陷去除率", "欠陥除去率", "Defect Removal Efficiency"),
    ("质量度量", "缺陷注入率", "欠陥挿入率", "Fault Injection Rate"),
    ("质量度量", "测试覆盖率", "テストカバレッジ", "Test Coverage"),
    ("质量度量", "变异得分", "ミューテーションスコア", "Mutation Score"),
    ("质量度量", "平均修复时长", "平均修復時間", "Mean Time to Repair (MTTR)"),
    ("质量度量", "缺陷逃逸率", "欠陥流出率", "Defect Escape Rate"),
    ("过程", "测试左移", "シフトレフト", "Shift-Left"),
    ("过程", "风险驱动测试", "リスクベースドテスト", "Risk-Based Testing"),
    ("过程", "V 字模型", "V 字モデル", "V-Model"),
]
r = 4
for row in glos:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)
ws.freeze_panes = "A4"

# ---------------- Sheet 17: 引用标准 ----------------
ws = wb.create_sheet("引用标准")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 12, "B": 30, "C": 60, "D": 30})
title_block(ws, "引用标准与参考文献清单", 4)
headers = ["类别", "编号", "名称", "用途"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
refs = [
    ("日本国家标准", "JIS X 0129-1:2013", "ソフトウェアテストの概念と定義（ISO/IEC/IEEE 29119-1）", "测试概念与定义"),
    ("日本国家标准", "JIS X 0129-2:2013", "テストプロセス（ISO/IEC/IEEE 29119-2）", "测试过程（组织、过程模型）"),
    ("日本国家标准", "JIS X 0129-3:2013", "テスト文書（ISO/IEC/IEEE 29119-3）", "测试文档结构（本文档遵循此标准）"),
    ("日本国家标准", "JIS X 0129-4:2016", "テスト技法（ISO/IEC/IEEE 29119-4）", "测试技法（§4 引用）"),
    ("日本国家标准", "JIS X 0129-5:2016", "キーワード駆動テスト（ISO/IEC/IEEE 29119-5）", "关键字驱动测试"),
    ("日本国家标准", "JIS Q 25000:2017", "システム及びソフトウェアの品質モデル", "质量模型"),
    ("日本国家标准", "JIS Q 25001:2017", "システム及びソフトウェアの品質モデル（改訂版）", "质量模型修订版"),
    ("IPA 实践", "IPA/SEC", "ソフトウェア開発データ白書", "测试数据/文档最佳实践"),
    ("IPA 实践", "IPA", "非機能要求グレード", "非功能需求等级分类（用于性能/可靠性等级）"),
    ("IPA 实践", "IPA", "セキュリティ要件チェックリスト", "安全需求检查表（用于 §4.7）"),
    ("ISO 标准", "ISO/IEC 25010:2011", "システム及びソフトウェア製品の品質モデル", "8 大质量属性"),
    ("ISO 标准", "ISO/IEC/IEEE 14764", "ソフトウェア保守", "维护性测试参考"),
    ("安全标准", "OWASP ASVS 4.0", "Application Security Verification Standard", "Web 应用安全验证基线"),
    ("安全标准", "CWE/SANS Top 25", "Most Dangerous Software Errors", "错误类型分类"),
    ("安全标准", "NIST SP 800-53", "Security and Privacy Controls", "安全控制目录"),
    ("V 字模型", "JIS X 0129-2 §5", "テストプロセス V 字モデル", "V 字模型与各级测试定位（§2.4 引用）"),
    ("V 字模型", "ISO/IEC/IEEE 12207", "ソフトウェアライフサイクルプロセス", "软件生命周期过程"),
    ("V 字模型", "ISO/IEC/IEEE 15288", "システムライフサイクルプロセス", "系统生命周期过程"),
]
r = 4
for row in refs:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)
ws.freeze_panes = "A4"

# ---------------- Save ----------------
_here = os.path.dirname(os.path.abspath(__file__))
out_dir = os.path.dirname(_here)
out = os.path.join(out_dir, "CATs_测试设计书_v1.0.xlsx")
wb.save(out)
print(f"Saved {out} with sheets: {wb.sheetnames}")
