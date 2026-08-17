# -*- coding: utf-8 -*-
"""OFCAT 详细设计（模块/接口/数据库）-> 一线大厂规格中文 Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

FONT="Microsoft YaHei"
NAVY="1F3864"; HEADER="2E5496"; SUB="8EAADB"; LIGHT="D9E1F2"; ZEBRA="F2F5FB"; WHITE="FFFFFF"
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def f(sz=10,b=False,color="000000"): return Font(name=FONT,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
def center(wrap=True): return Alignment(horizontal="center",vertical="center",wrap_text=wrap)
def left(wrap=True): return Alignment(horizontal="left",vertical="center",wrap_text=wrap)
def topleft(): return Alignment(horizontal="left",vertical="top",wrap_text=True)
def style_header(ws,row,n,fc=HEADER):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c); cell.fill=fill(fc); cell.font=f(10,True,WHITE)
        cell.alignment=center(); cell.border=BORDER
def grid(ws,r1,r2,c1,c2,zebra=True):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            cell=ws.cell(row=r,column=c); cell.border=BORDER
            if cell.alignment is None or cell.alignment.vertical is None: cell.alignment=topleft()
            if zebra and (r-r1)%2==1 and cell.fill.fgColor.rgb in (None,"00000000"): cell.fill=fill(ZEBRA)
def widths(ws,mp):
    for col,w in mp.items(): ws.column_dimensions[col].width=w
def title_block(ws,title,span):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    c=ws.cell(1,1,title); c.fill=fill(NAVY); c.font=f(13,True,WHITE); c.alignment=center(False)
    ws.row_dimensions[1].height=28
def table(ws,headers,rows,wmap,title,center_cols=(),rowh=None,freeze_col="A",autofilter=True):
    ws.sheet_view.showGridLines=False
    title_block(ws,title,len(headers)); widths(ws,wmap)
    for i,h in enumerate(headers): ws.cell(3,i+1,h)
    style_header(ws,3,len(headers))
    r=4
    for row in rows:
        for i,x in enumerate(row): ws.cell(r,i+1,x)
        if rowh: ws.row_dimensions[r].height=rowh
        r+=1
    grid(ws,4,r-1,1,len(headers))
    for rr in range(4,r):
        for cc in center_cols: ws.cell(rr,cc).alignment=center()
    ws.freeze_panes=f"{freeze_col}4"
    if autofilter: ws.auto_filter.ref=f"A3:{get_column_letter(len(headers))}{r-1}"
    return r-1
def cover(ws,title,sub,meta,stats):
    ws.sheet_view.showGridLines=False
    widths(ws,{"A":3,"B":26,"C":56,"D":3})
    ws.merge_cells("B2:C2"); ws["B2"]=title; ws["B2"].font=f(22,True,NAVY); ws["B2"].alignment=center(False); ws.row_dimensions[2].height=44
    ws.merge_cells("B3:C3"); ws["B3"]=sub; ws["B3"].font=f(12,True,"404040"); ws["B3"].alignment=center(False); ws.row_dimensions[3].height=24
    r=5
    for k,v in meta:
        ws.cell(r,2,k).fill=fill(SUB); ws.cell(r,2).font=f(10,True,WHITE); ws.cell(r,2).alignment=center()
        ws.cell(r,3,v).alignment=left(); ws.cell(r,3).font=f(10)
        ws.cell(r,2).border=BORDER; ws.cell(r,3).border=BORDER; ws.row_dimensions[r].height=20; r+=1
    r+=1; ws.cell(r,2,"摘要（自动统计）").font=f(10,True,NAVY); r+=1
    for k,fml in stats:
        ws.cell(r,2,k).fill=fill(LIGHT); ws.cell(r,2).font=f(10,True); ws.cell(r,2).alignment=center()
        ws.cell(r,3,fml).font=f(10,True,NAVY); ws.cell(r,3).alignment=left(False)
        ws.cell(r,2).border=BORDER; ws.cell(r,3).border=BORDER; ws.row_dimensions[r].height=20; r+=1
def tabcolors(wb,mp):
    for n,c in mp.items(): wb[n].sheet_properties.tabColor=c
    try: wb.calculation.fullCalcOnLoad=True
    except Exception: pass

META=lambda no,name,extra: [("文档编号",no),("文档名",name),("版本","第 1.0 版（草稿）"),
   ("创建日","2026-06-25"),("作者","架构师"),("状态","评审前草稿"),("密级","仅社内"),("上游文档",extra)]

# ============================================================== 模块设计书
wb=Workbook(); cover(wb.active,"模块设计书","OFCAT — 详细设计 / 模块内部逻辑·算法",
  META("OFCAT-DD-M-001","模块设计书","基础设计书 v1.0"),
  [("模块数","=COUNTA(模块一览!A4:A100)"),("全局参数","=COUNTA(全局参数!A4:A100)"),
   ("编排节点","=COUNTA(编排节点!A4:A100)"),("QA 规则","=COUNTA(QA规则!A4:A100)")])
wb.active.title="封面"
table(wb.create_sheet("目录"),["No","工作表名","内容"],
 [("1","封面","文档管理信息·摘要"),("2","目录","本工作表"),("3","修订履历","版本履历"),
  ("4","模块一览","M-01~06"),("5","全局参数","阈值/超时等默认值"),("6","编排节点","LangGraph 节点流转"),
  ("7","TM匹配算法","L0/L1 算法与复杂度"),("8","术语校验","匹配/注入/强制校验"),
  ("9","保护识别模式","标签/占位符正则"),("10","QA规则","QA-01~05"),
  ("11","网关与合规","路由/提示/回退"),("12","复杂度性能","各模块复杂度")],
 {"A":6,"B":16,"C":56},"目录（Index）",center_cols=(1,),autofilter=False)
table(wb.create_sheet("修订履历"),["版本","日期","修订者","修订内容"],
 [("1.0","2026-06-25","架构师","初版。M-01~M-06 模块算法、参数、异常与复杂度")],
 {"A":8,"B":14,"C":18,"D":80},"修订履历",center_cols=(1,2))
table(wb.create_sheet("模块一览"),["模块ID","模块名","承接组件","关联功能"],
 [("M-01","Orchestrator（编排）","CMP-05","F2–F6"),
  ("M-02","TM Service（记忆库匹配）","CMP-06","F2,F9"),
  ("M-03","Term Service（术语匹配/校验）","CMP-07","F3,F4"),
  ("M-04","Tag/Placeholder Protector（保护）","CMP-08","F5"),
  ("M-05","QA Service（质量校验）","CMP-09","F4,F5"),
  ("M-06","AI Gateway（模型网关/合规）","CMP-10","F6,F10")],
 {"A":10,"B":34,"C":14,"D":14},"模块一览（Modules）",center_cols=(1,3,4))
table(wb.create_sheet("全局参数"),["参数","默认","说明"],
 [("FUZZY_THRESHOLD","85","模糊匹配采纳阈值（%）"),("FUZZY_TOPK","5","模糊候选返回上限"),
  ("LEN_WINDOW","0.30","候选长度窗口（±30%）"),("VEC_TOPK","20","语义召回候选数"),
  ("VEC_MIN_SCORE","0.80","语义召回最低余弦相似度"),("MODEL_TIMEOUT_S","20","单次模型调用超时（秒）"),
  ("MODEL_RETRY","2","失败重试次数"),("EMBED_DIM","1024","bge-m3 向量维度")],
 {"A":22,"B":12,"C":62},"全局参数（默认值）",center_cols=(2,))
table(wb.create_sheet("编排节点"),["节点","类型","输入","输出","说明"],
 [("capture","入口","选区","标准化 seg","获取上下文/源语言"),
  ("tm_match","确定性","seg","TM 命中/miss","调 M-02；L0/L1≥阈值直接 respond"),
  ("term_match","确定性","seg","术语约束","调 M-03"),
  ("protect","确定性","seg","哨兵文本+token_map","调 M-04"),
  ("translate","模型","哨兵文本+约束","流式译文","调 M-06；high_quality→多模型子图"),
  ("restore","确定性","译文+map","回填译文","调 M-04，校验数量一致"),
  ("term_enforce","确定性","译文","违反列表","调 M-03"),
  ("qa","确定性","译文","QA 结果","调 M-05"),
  ("respond","出口","状态","SSE 事件","下发扩展"),
  ("persist","出口","确认译文","TM 写入","调 M-02（F9，确认后）")],
 {"A":14,"B":10,"C":20,"D":20,"E":40},"编排节点（LangGraph）",center_cols=(2,),rowh=28)
table(wb.create_sheet("TM匹配算法"),["阶段","处理","复杂度"],
 [("规范化","NFKC+trim+折叠空白；hash=sha1(source_norm)","O(len)"),
  ("L0 精确","按 source_hash 查最优（quality,usage 排序）","O(1)"),
  ("L1 候选预过滤","同语言对+domain+长度窗口 ±30%","O(1) 索引"),
  ("L1 词面相似","RapidFuzz token_sort_ratio，主排序依据","O(N)"),
  ("L1 语义召回","vec KNN top20，cos≥0.80，并入候选","O(logN)"),
  ("采纳","综合分≥85 取前 5，最高分为 best+diff","—")],
 {"A":16,"B":58,"C":14},"TM 匹配算法（M-02）",center_cols=(3,))
table(wb.create_sheet("术语校验"),["项","说明"],
 [("匹配模式","word（CJK 用 Aho-Corasick 词边界）/ substring / regex"),
  ("领域优先","同一 source_term 命中多条时 domain 精确 > 全局"),
  ("missing","约定译法缺失 → 标记 + 建议 target_term"),
  ("forbidden","命中禁止译法 → 标记 + 建议（唯一且安全时可自动替换）"),
  ("注入","约束清单注入 system 提示（建议）；强制性由译后校验保证")],
 {"A":14,"B":80},"术语匹配/强制校验（M-03）")
table(wb.create_sheet("保护识别模式"),["类型","正则（示意）"],
 [("HTML/XML 标签","</?[a-zA-Z][^>]*>"),("双花括号变量","{{ [\\w.]+ }}"),
  ("单花括号占位","{[\\w.]+}"),("printf 占位","%(\\d+$)?[sdfx]"),
  ("命名占位","%\\w+%  /  :[a-zA-Z_]\\w*")],
 {"A":18,"B":60},"标签/占位符识别（M-04）")
r_qa=table(wb.create_sheet("QA规则"),["规则ID","名称","判定","严重度"],
 [("QA-01","标签/占位符数量一致","原文与译文计数相等","阻断"),
  ("QA-02","术语合规","M-03 无 missing/forbidden","阻断"),
  ("QA-03","数字一致","数字多重集一致（本地化格式可配置）","警告"),
  ("QA-04","非空/长度","译文非空且长度比合理","警告"),
  ("QA-05","禁用字符","无控制字符、无残留哨兵","阻断")],
 {"A":10,"B":24,"C":40,"D":12},"QA 规则集（M-05）",center_cols=(1,4))
qa=wb["QA规则"]
qa.conditional_formatting.add(f"D4:D{r_qa}",CellIsRule(operator="equal",formula=['"阻断"'],fill=fill("FFC7CE"),font=f(10,True,"9C0006")))
qa.conditional_formatting.add(f"D4:D{r_qa}",CellIsRule(operator="equal",formula=['"警告"'],fill=fill("FFEB9C"),font=f(10,True,"9C6500")))
table(wb.create_sheet("网关与合规"),["项","设计"],
 [("判定优先级","项目标记 > 域名黑名单 > 域名白名单 > 默认策略"),
  ("敏感路由","强制本地模型；不可达 → ComplianceBlocked（不回退云端）"),
  ("非敏感路由","默认云端（可配置）"),
  ("提示拼装","领域设定 + 占位保留指令 + 术语约束清单 + 目标语言/风格"),
  ("调用回退","超时重试 MODEL_RETRY；非敏感可切备用；敏感禁止回退云端"),
  ("流式","首 token 经 SSE delta 透传（NFR-05）")],
 {"A":16,"B":78},"AI 网关 / 合规（M-06）")
table(wb.create_sheet("复杂度性能"),["模块","关键路径复杂度","备注"],
 [("M-02 L0","O(1)","哈希索引"),("M-02 L1","O(N) 词面 + O(log) 向量","N 受预过滤约束（百量级）"),
  ("M-03","O(len)","Aho-Corasick 缓存自动机"),("M-04","O(len)","单遍正则扫描"),
  ("M-05","O(len)","计数/集合比较"),("M-06","受模型 RTT 支配","流式首字 ~1s（NFR-05）")],
 {"A":14,"B":24,"C":40},"复杂度与性能要点",center_cols=())
tabcolors(wb,{"封面":NAVY,"目录":"7F7F7F","TM匹配算法":"C00000","QA规则":"BF8F00","网关与合规":"7030A0"})
o1=r"D:\OFCAT\doc\03-详细设计\模块设计\OFCAT_模块设计书_v1.0.xlsx"; wb.save(o1); print("SAVED",o1)

# ============================================================== 接口设计书
wb=Workbook(); cover(wb.active,"接口设计书","OFCAT — 详细设计 / API 契约",
  META("OFCAT-DD-I-001","接口设计书","基础设计书 v1.0"),
  [("API 数","=COUNTA(API一览!A4:A100)"),("错误码数","=COUNTA(错误码!A4:A100)"),
   ("SSE 事件","=COUNTA(SSE事件!A4:A100)"),("字段定义","=COUNTA(请求字段!A4:A200)")])
wb.active.title="封面"
table(wb.create_sheet("目录"),["No","工作表名","内容"],
 [("1","封面","文档管理信息·摘要"),("2","目录","本工作表"),("3","修订履历","版本履历"),
  ("4","通用约定","基址/认证/安全"),("5","错误码","错误码表"),("6","API一览","API-01~10"),
  ("7","SSE事件","/translate 事件流"),("8","请求字段","主要端点请求字段")],
 {"A":6,"B":16,"C":56},"目录（Index）",center_cols=(1,),autofilter=False)
table(wb.create_sheet("修订履历"),["版本","日期","修订者","修订内容"],
 [("1.0","2026-06-25","架构师","初版。扩展↔引擎 API-01~10 完整契约、SSE 事件、错误码")],
 {"A":8,"B":14,"C":18,"D":80},"修订履历",center_cols=(1,2))
table(wb.create_sheet("通用约定"),["项","约定"],
 [("基址","http://127.0.0.1:<port>（握手确定/可配置）"),
  ("编码","application/json; charset=utf-8；时间 ISO-8601 UTC"),
  ("协议版本","请求头 X-OFCAT-API: 1"),
  ("鉴权","Authorization: Bearer <token>（引擎启动随机生成，经扩展安装态获取）"),
  ("Origin 校验","校验 Origin/扩展 ID 白名单，否则 403"),
  ("Host 校验","仅接受 127.0.0.1/localhost，防 DNS rebinding"),
  ("速率","单令牌限流（默认 20 req/s）"),
  ("错误信封","{ error: { code, message, details } }")],
 {"A":16,"B":78},"通用约定（Conventions）")
r_ec=table(wb.create_sheet("错误码"),["code","HTTP","含义","关联"],
 [("VALIDATION_ERROR","400","请求体校验失败","全部"),
  ("UNAUTHORIZED","401","令牌缺失/失效","E-02"),
  ("FORBIDDEN_ORIGIN","403","Origin/扩展 ID 不在白名单","安全"),
  ("NOT_FOUND","404","资源不存在（如 job_id）","API-09"),
  ("COMPLIANCE_BLOCKED","409","敏感且本地模型不可达，fail-closed","E-04"),
  ("QA_BLOCKED","422","QA 阻断项存在，禁止写回","E-05"),
  ("RATE_LIMITED","429","超出限流","—"),
  ("MODEL_ERROR","502","模型调用失败（重试/回退后仍失败）","E-03"),
  ("ENGINE_ERROR","500","引擎内部错误","—"),
  ("MODEL_TIMEOUT","504","模型超时","E-03")],
 {"A":22,"B":8,"C":42,"D":12},"错误码表（Error Codes）",center_cols=(2,4))
table(wb.create_sheet("API一览"),["接口ID","方法/路径","用途","鉴权","关联"],
 [("API-01","GET /health","健康检查/版本/能力","否","—"),
  ("API-02","POST /session/handshake","令牌校验/能力协商","是","安全"),
  ("API-03","POST /tm/match","仅 TM 匹配（不翻译）","是","F2"),
  ("API-04","POST /translate（SSE）","完整翻译管道，流式","是","F3–F6"),
  ("API-05","POST /tm/commit","确认译文回存 TM","是","F9"),
  ("API-06","POST /qa/check","独立 QA 校验","是","F4,F5"),
  ("API-07","GET/PUT /settings","设置读写","是","F10"),
  ("API-08","POST /import","存量导入任务（异步）","是","F11"),
  ("API-09","GET /import/{job_id}","导入进度/报告","是","F11"),
  ("API-10","POST /sync/pull · /push","中心同步代理","是","C4")],
 {"A":10,"B":30,"C":34,"D":8,"E":12},"API 一览（Extension↔Engine）",center_cols=(1,4,5))
table(wb.create_sheet("SSE事件"),["event","data","说明"],
 [("tm_hit","{level,score,target,diff}","命中 L0/L1 时下发（随后直接 done）"),
  ("terms","{matched:[{source_term,target_term}]}","命中术语"),
  ("delta","{text}","流式译文增量"),
  ("qa","{pass,blocks,warns}","QA 结果"),
  ("done","{target,level,terms,qa,elapsed_ms}","终态最终译文与元数据"),
  ("error","{code,message}","异常终止（对应错误码表）")],
 {"A":12,"B":44,"C":42},"/translate SSE 事件流")
table(wb.create_sheet("请求字段"),["端点","字段","类型","必填","约束/说明"],
 [("/tm/match","segment","string","是","1..5000"),
  ("/tm/match","source_lang/target_lang","string","是","BCP-47，如 ja/zh-CN"),
  ("/tm/match","domain","string","否","默认 ''"),
  ("/tm/match","project_id","int","否","—"),
  ("/translate","segment","string","是","1..5000，含标签/占位符原样"),
  ("/translate","source_lang/target_lang","string","是","BCP-47"),
  ("/translate","domain","string","否","—"),
  ("/translate","project_id","int","否","—"),
  ("/translate","url","string","否","用于合规判定"),
  ("/translate","mode","string","否","default(L2)/high_quality(L3)"),
  ("/tm/commit","segment/target","string","是","—"),
  ("/tm/commit","source_lang/target_lang","string","是","—"),
  ("/tm/commit","domain/context","string","否","—"),
  ("/tm/commit","origin","string","否","默认 human")],
 {"A":14,"B":24,"C":10,"D":8,"E":40},"主要端点请求字段",center_cols=(3,4))
tabcolors(wb,{"封面":NAVY,"目录":"7F7F7F","API一览":"C00000","错误码":"BF8F00","SSE事件":"2E75B6"})
o2=r"D:\OFCAT\doc\03-详细设计\接口设计\OFCAT_接口设计书_v1.0.xlsx"; wb.save(o2); print("SAVED",o2)

# ============================================================== 数据库设计书
wb=Workbook(); cover(wb.active,"数据库设计书","OFCAT — 详细设计 / DDL·索引·迁移",
  META("OFCAT-DD-D-001","数据库设计书","基础设计书 v1.0"),
  [("表数","=COUNTA(表一览!A4:A100)"),("索引数","=COUNTA(索引一览!A4:A100)"),
   ("TM 列数","=COUNTA('列-TM'!A4:A100)"),("术语列数","=COUNTA('列-terms'!A4:A100)")])
wb.active.title="封面"
table(wb.create_sheet("目录"),["No","工作表名","内容"],
 [("1","封面","文档管理信息·摘要"),("2","目录","本工作表"),("3","修订履历","版本履历"),
  ("4","表一览","全部表"),("5","列-projects","projects 列定义"),("6","列-terms","terms 列定义"),
  ("7","列-TM","translation_memory 列定义"),("8","列-辅助表","history/import_jobs/…"),
  ("9","索引一览","索引设计"),("10","同步合并","C4 合并策略")],
 {"A":6,"B":16,"C":56},"目录（Index）",center_cols=(1,),autofilter=False)
table(wb.create_sheet("修订履历"),["版本","日期","修订者","修订内容"],
 [("1.0","2026-06-25","架构师","初版。物理模型、DDL、索引、向量表、迁移与合并策略")],
 {"A":8,"B":14,"C":18,"D":80},"修订履历",center_cols=(1,2))
table(wb.create_sheet("表一览"),["表名","说明","主键"],
 [("projects","项目/语言对/领域/敏感策略","id"),("terms","术语库条目","id"),
  ("translation_memory","TM 句段对","id"),("tm_vectors","TM 语义向量（vec0 虚拟表）","tm_id"),
  ("history","变更/操作审计","id"),("import_jobs","导入任务与报告","id(uuid)"),
  ("sync_state","同步游标","entity"),("settings","键值设置","key"),
  ("schema_migrations","迁移版本","version")],
 {"A":22,"B":40,"C":12},"表一览（Tables）",center_cols=(3,))
COL=["列名","类型","NULL","默认","键","说明"]
table(wb.create_sheet("列-projects"),COL,
 [("id","INTEGER","否","-","PK","主键"),("name","TEXT","否","-","-","项目名"),
  ("source_lang","TEXT","是","-","-","源语言"),("target_lang","TEXT","是","-","-","目标语言"),
  ("domain","TEXT","否","''","-","领域"),("sensitivity","TEXT","否","'normal'","-","normal|sensitive"),
  ("created_at","TEXT","否","now","-","创建时间"),("updated_at","TEXT","是","-","-","更新时间")],
 {"A":16,"B":12,"C":7,"D":12,"E":16,"F":36},"列定义 — projects",center_cols=(3,5))
table(wb.create_sheet("列-terms"),COL,
 [("id","INTEGER","否","-","PK","主键"),("project_id","INTEGER","是","-","FK→projects","NULL=全局"),
  ("source_lang","TEXT","否","-","索引","源语言"),("target_lang","TEXT","否","-","索引","目标语言"),
  ("domain","TEXT","否","''","索引","领域"),("source_term","TEXT","否","-","-","原术语"),
  ("source_term_norm","TEXT","否","-","唯一键","规范化匹配键"),("target_term","TEXT","否","-","-","约定译法"),
  ("forbidden","TEXT","否","'[]'","-","JSON 禁止译法"),("case_sensitive","INTEGER","否","0","-","是否区分大小写"),
  ("match_mode","TEXT","否","'word'","-","word|substring|regex"),("note","TEXT","是","-","-","备注"),
  ("status","TEXT","否","'active'","索引","active|disabled"),("created_at","TEXT","否","now","-","创建时间"),
  ("updated_at","TEXT","是","-","-","更新时间")],
 {"A":18,"B":10,"C":7,"D":10,"E":14,"F":34},"列定义 — terms",center_cols=(3,5))
table(wb.create_sheet("列-TM"),COL,
 [("id","INTEGER","否","-","PK","主键"),("project_id","INTEGER","是","-","FK→projects","-"),
  ("source_lang","TEXT","否","-","索引","源语言"),("target_lang","TEXT","否","-","索引","目标语言"),
  ("domain","TEXT","否","''","索引","领域"),("source_text","TEXT","否","-","-","原文"),
  ("source_norm","TEXT","否","-","-","规范化文本（模糊基准）"),("source_hash","TEXT","否","-","索引/唯一","sha1(source_norm)（L0）"),
  ("source_len","INTEGER","否","-","索引","长度预过滤"),("target_text","TEXT","否","-","-","译文"),
  ("context","TEXT","是","-","-","上下文/URL"),("origin","TEXT","否","'human'","-","human|import|mt"),
  ("quality","INTEGER","否","0","-","质量等级"),("usage_count","INTEGER","否","0","-","使用次数"),
  ("created_at","TEXT","否","now","-","创建时间"),("updated_at","TEXT","是","-","-","更新时间")],
 {"A":16,"B":10,"C":7,"D":10,"E":14,"F":36},"列定义 — translation_memory",center_cols=(3,5))
table(wb.create_sheet("列-辅助表"),["表","列","类型","说明"],
 [("tm_vectors","tm_id","INTEGER PK","向量主键（=tm.id）"),("tm_vectors","embedding","FLOAT[1024]","bge-m3 向量"),
  ("history","id","INTEGER PK","主键"),("history","entity_type/entity_id","TEXT/INT","tm|term + 实体ID"),
  ("history","action","TEXT","create|update|delete|import|sync"),("history","before/after","TEXT(JSON)","变更前后"),
  ("history","actor/ts","TEXT","操作者/时间"),
  ("import_jobs","id","TEXT PK","uuid"),("import_jobs","status","TEXT","pending|running|done|failed"),
  ("import_jobs","total/succeeded/duplicated/failed","INTEGER","统计"),("import_jobs","mapping/report","TEXT(JSON)","映射/错误明细"),
  ("sync_state","entity","TEXT PK","terms|tm"),("sync_state","last_pull_at/last_push_at/cursor","TEXT","同步游标"),
  ("settings","key/value","TEXT","键 / JSON 值"),
  ("schema_migrations","version/applied_at","INTEGER/TEXT","迁移版本")],
 {"A":18,"B":34,"C":16,"D":40},"列定义 — 辅助表",center_cols=())
table(wb.create_sheet("索引一览"),["索引名","表","列","唯一","服务于"],
 [("idx_tm_hash","translation_memory","source_lang,target_lang,source_hash","否","M-02 L0"),
  ("idx_tm_filter","translation_memory","source_lang,target_lang,domain,source_len","否","M-02 L1 候选"),
  ("uq_tm","translation_memory","source_lang,target_lang,domain,source_hash,IFNULL(project_id,0)","是","回存去重"),
  ("idx_term_lookup","terms","source_lang,target_lang,domain,status","否","术语载入"),
  ("uq_term","terms","IFNULL(project_id,0),source_lang,target_lang,domain,source_term_norm","是","术语唯一"),
  ("idx_hist","history","entity_type,entity_id,ts","否","审计")],
 {"A":16,"B":22,"C":50,"D":8,"E":14},"索引一览（Indexes）",center_cols=(4,))
table(wb.create_sheet("同步合并"),["实体","拉取冲突","推送冲突"],
 [("terms","以中心为准（管理者维护权威），本地差异记 history","仅管理者推送，普通客户端只读拉取"),
  ("translation_memory","后写优先（updated_at 新者胜），覆盖者留 history；同 hash 不同译文双方都留并标记 conflict 待人工","同左")],
 {"A":20,"B":52,"C":36},"同步合并策略（C4）",rowh=44)
tabcolors(wb,{"封面":NAVY,"目录":"7F7F7F","列-TM":"C00000","索引一览":"2E75B6","同步合并":"7030A0"})
o3=r"D:\OFCAT\doc\03-详细设计\数据库设计\OFCAT_数据库设计书_v1.0.xlsx"; wb.save(o3); print("SAVED",o3)
