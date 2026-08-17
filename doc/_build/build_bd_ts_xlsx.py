# -*- coding: utf-8 -*-
"""OFCAT 基础设计书 + 技术选型书 -> 一线大厂规格中文 Excel 工作簿"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

FONT="Microsoft YaHei"
NAVY="1F3864"; HEADER="2E5496"; SUB="8EAADB"; LIGHT="D9E1F2"; ZEBRA="F2F5FB"; WHITE="FFFFFF"
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def f(sz=10,b=False,color="000000"): return Font(name=FONT,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
def center(wrap=True): return Alignment(horizontal="center",vertical="center",wrap_text=wrap)
def left(wrap=True): return Alignment(horizontal="left",vertical="center",wrap_text=wrap)
def topleft(): return Alignment(horizontal="left",vertical="top",wrap_text=True)

def style_header(ws,row,ncols,fillc=HEADER):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c); cell.fill=fill(fillc); cell.font=f(10,True,WHITE)
        cell.alignment=center(); cell.border=BORDER
def grid(ws,r1,r2,c1,c2,zebra=True):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            cell=ws.cell(row=r,column=c); cell.border=BORDER
            if cell.alignment is None or cell.alignment.vertical is None: cell.alignment=topleft()
            if zebra and (r-r1)%2==1 and cell.fill.fgColor.rgb in (None,"00000000"): cell.fill=fill(ZEBRA)
def widths(ws,mp):
    for col,w in mp.items(): ws.column_dimensions[col].width=w
def title_block(ws,title,span):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    c=ws.cell(1,1,title); c.fill=fill(NAVY); c.font=f(13,True,WHITE); c.alignment=center(False)
    ws.row_dimensions[1].height=28

def table(ws,headers,rows,wmap,title,center_cols=(),rowh=None,freeze_col="A",autofilter=True):
    ws.sheet_view.showGridLines=False
    title_block(ws,title,len(headers)); widths(ws,wmap)
    for i,h in enumerate(headers): ws.cell(3,i+1,h)
    style_header(ws,3,len(headers))
    r=4
    for row in rows:
        for i,x in enumerate(row): ws.cell(r,i+1,x)
        if rowh: ws.row_dimensions[r].height=rowh
        r+=1
    grid(ws,4,r-1,1,len(headers))
    for rr in range(4,r):
        for cc in center_cols: ws.cell(rr,cc).alignment=center()
    ws.freeze_panes=f"{freeze_col}4"
    if autofilter:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref=f"A3:{get_column_letter(len(headers))}{r-1}"
    return r-1

def cover(ws,title,sub,meta,stats):
    ws.sheet_view.showGridLines=False
    widths(ws,{"A":3,"B":26,"C":56,"D":3})
    ws.merge_cells("B2:C2"); ws["B2"]=title; ws["B2"].font=f(22,True,NAVY); ws["B2"].alignment=center(False); ws.row_dimensions[2].height=44
    ws.merge_cells("B3:C3"); ws["B3"]=sub; ws["B3"].font=f(12,True,"404040"); ws["B3"].alignment=center(False); ws.row_dimensions[3].height=24
    r=5
    for k,v in meta:
        ws.cell(r,2,k).fill=fill(SUB); ws.cell(r,2).font=f(10,True,WHITE); ws.cell(r,2).alignment=center()
        ws.cell(r,3,v).alignment=left(); ws.cell(r,3).font=f(10)
        ws.cell(r,2).border=BORDER; ws.cell(r,3).border=BORDER; ws.row_dimensions[r].height=20; r+=1
    r+=1; ws.cell(r,2,"摘要（自动统计）").font=f(10,True,NAVY); r+=1
    for k,fml in stats:
        ws.cell(r,2,k).fill=fill(LIGHT); ws.cell(r,2).font=f(10,True); ws.cell(r,2).alignment=center()
        ws.cell(r,3,fml).font=f(10,True,NAVY); ws.cell(r,3).alignment=left(False)
        ws.cell(r,2).border=BORDER; ws.cell(r,3).border=BORDER; ws.row_dimensions[r].height=20; r+=1

# ====================================================================== 基础设计书
wb=Workbook()
cv=wb.active; cv.title="封面"
cover(cv,"基础设计书","OFCAT — AI 增强型 CAT 浏览器工作台",
  [("文档编号","OFCAT-BD-001"),("文档名","基础设计书"),("版本","第 1.0 版（草稿）"),
   ("创建日","2026-06-25"),("作者","架构师"),("状态","评审前草稿"),
   ("密级","仅社内"),("上游文档","需求定义书 v1.1")],
  [("组件数","=COUNTA(组件一览!A4:A100)"),("接口数","=COUNTA(接口一览!A4:A100)"),
   ("数据实体数","=COUNTA(数据实体!A4:A100)"),("画面数","=COUNTA(画面一览!A4:A100)")])

ws=wb.create_sheet("目录")
table(ws,["No","工作表名","内容"],
 [("1","封面","文档管理信息·摘要"),("2","目录","本工作表"),("3","修订履历","版本履历"),
  ("4","设计原则","设计原则与前提假设"),("5","组件一览","CMP-01~14 组件职责"),
  ("6","部署构成","节点·部署物"),("7","接口一览","API-01~10 扩展↔引擎"),
  ("8","数据实体","实体·物理表·维度"),("9","字段-TM","translation_memory 字段"),
  ("10","字段-术语","terms 字段"),("11","画面一览","UI1~4"),
  ("12","错误处理","E-01~08 处理方针"),("13","非功能实现","NFR 实现方针"),
  ("14","安全设计","安全项·设计")],
 {"A":6,"B":14,"C":58},"目录（Index）",center_cols=(1,),autofilter=False)
ws.column_dimensions["A"].width=6

table(wb.create_sheet("修订履历"),["版本","日期","修订者","修订内容"],
 [("1.0","2026-06-25","架构师","初版。确立薄扩展+本地引擎方式、组件分解、数据模型、接口、非功能与安全设计")],
 {"A":8,"B":14,"C":18,"D":80},"修订履历（Revision History）",center_cols=(1,2))

table(wb.create_sheet("设计原则"),["No/课题","内容"],
 [("原则1","确定性优先：TM/术语/标签正确性由确定性逻辑保证，不依赖模型自觉"),
  ("原则2","分层延迟：L0/L1 本地即时解决，绝不调用云端；AI 仅用于必要环节"),
  ("原则3","合规 fail-closed：判定失败或本地模型不可达时宁可中止，不外泄敏感内容"),
  ("原则4","薄扩展：浏览器侧只做捕获/渲染/写回，重逻辑下沉本地引擎"),
  ("原则5","接口契约稳定：扩展↔引擎、引擎↔模型以稳定契约解耦"),
  ("前提 O1","TM≤50万/术语≤10万，SQLite 直接索引"),
  ("前提 O2","Edge + Chrome 双支持（同源码）"),
  ("前提 O6","域名清单+项目标记混合策略，默认 fail-closed")],
 {"A":12,"B":86},"设计原则与前提",center_cols=(1,))

table(wb.create_sheet("组件一览"),["组件ID","组件名","所属","职责","关联功能"],
 [("CMP-01","Content Script","扩展","选区捕获、overlay 注入与编辑、写回页面","F1,F7,F8"),
  ("CMP-02","Service Worker","扩展","消息路由、与引擎会话、SSE 中转、快捷键/右键","F1,F6"),
  ("CMP-03","Options/Popup UI","扩展","模型/合规/语言对/术语·TM 路径设置","F10,UI2"),
  ("CMP-04","API 层","引擎","鉴权、Origin 校验、REST/SSE 端点","全部"),
  ("CMP-05","Orchestrator","引擎","LangGraph 编排翻译管道","F2–F6"),
  ("CMP-06","TM Service","引擎","精确/模糊匹配、回存","F2,F9"),
  ("CMP-07","Term Service","引擎","术语匹配、注入、强制校验","F3,F4"),
  ("CMP-08","Tag/Placeholder Protector","引擎","标签/占位符哨兵化与回填校验","F5"),
  ("CMP-09","QA Service","引擎","数量/标签/术语一致性 QA","F4,F5"),
  ("CMP-10","AI Gateway","引擎","模型路由、合规开关、流式","F6,F10"),
  ("CMP-11","Storage","引擎","SQLite + 向量、事务、迁移","F2,F9,F11"),
  ("CMP-12","Import Pipeline","引擎","存量清洗/映射/去重/导入","F11"),
  ("CMP-13","Sync Client","引擎","与中心同步服务的异步同步","C4"),
  ("CMP-14","中心同步服务","服务端","术语/核心 TM 的服务端同步 API","C4")],
 {"A":10,"B":24,"C":10,"D":50,"E":14},"组件一览（Components）",center_cols=(1,3,5),rowh=30)

table(wb.create_sheet("部署构成"),["节点","部署物","说明"],
 [("用户 PC","浏览器扩展（store/sideload）","薄扩展"),
  ("用户 PC","本地引擎（PyInstaller，托盘/自启）","监听 127.0.0.1:<port>"),
  ("用户 PC 或内网","本地 LLM（Ollama 单机 / vLLM 内网）","合规路径"),
  ("内网/云","中心同步服务（可选）","术语/核心 TM 异步同步"),
  ("外网","云端模型 API","非敏感内容")],
 {"A":18,"B":40,"C":40},"部署构成（Deployment）")

rmax=table(wb.create_sheet("接口一览"),["接口ID","方法/路径","用途","关联功能"],
 [("API-01","GET /health","引擎健康检查、版本、能力探测","—"),
  ("API-02","POST /session/handshake","令牌校验、能力协商","安全"),
  ("API-03","POST /tm/match","TM 精确/模糊匹配","F2"),
  ("API-04","POST /translate（SSE）","完整翻译管道（术语/保护/QA），流式","F3–F6"),
  ("API-05","POST /tm/commit","确认译文回存 TM","F9"),
  ("API-06","POST /qa/check","独立 QA 校验（标签/术语/数量）","F4,F5"),
  ("API-07","GET/PUT /settings","设置读写（模型/合规/语言对）","F10"),
  ("API-08","POST /import（job）","存量导入任务（异步，返回 job_id）","F11"),
  ("API-09","GET /import/{job_id}","导入进度/结果报告","F11"),
  ("API-10","POST /sync/pull · /sync/push","与中心同步服务对接（引擎代理）","C4")],
 {"A":10,"B":30,"C":52,"D":14},"接口一览（Extension↔Engine API）",center_cols=(1,4))

table(wb.create_sheet("数据实体"),["实体","物理表","说明","关键维度"],
 [("项目","projects","项目/语言对/领域/敏感策略","—"),
  ("术语","terms","术语库条目","source_lang,target_lang,domain"),
  ("翻译记忆","translation_memory","TM 句段对","source_lang,target_lang,domain"),
  ("TM 向量","tm_vectors","句段语义向量（sqlite-vec）","tm_id"),
  ("历史","history","变更/操作审计","—"),
  ("同步状态","sync_state","各实体同步游标","entity"),
  ("设置","settings","键值设置","key")],
 {"A":14,"B":22,"C":34,"D":34},"数据实体一览（Entities）")

table(wb.create_sheet("字段-TM"),["字段","类型","说明"],
 [("id","INTEGER PK","主键"),("source_lang / target_lang","TEXT","语言对（C5），如 ja/zh-CN"),
  ("domain","TEXT","领域（游戏/反馈/技术…）"),("source_text","TEXT","原文"),
  ("source_hash","TEXT","原文规范化哈希（精确匹配索引）"),("target_text","TEXT","译文"),
  ("context","TEXT","上下文/来源 URL"),("project_id","INTEGER FK","所属项目"),
  ("origin","TEXT","来源（human/import/mt）"),("quality","INTEGER","质量等级"),
  ("created_at / updated_at","TEXT","时间戳")],
 {"A":26,"B":16,"C":56},"字段定义 — translation_memory")

table(wb.create_sheet("字段-术语"),["字段","类型","说明"],
 [("id","INTEGER PK","主键"),("source_lang / target_lang","TEXT","语言对"),
  ("domain","TEXT","领域"),("source_term","TEXT","原术语"),("target_term","TEXT","约定译法"),
  ("forbidden","TEXT(JSON)","禁止译法列表（供 F4 校验）"),("note","TEXT","备注"),
  ("status","TEXT","状态（启用/停用）"),("updated_at","TEXT","时间戳")],
 {"A":26,"B":16,"C":56},"字段定义 — terms")

table(wb.create_sheet("画面一览"),["画面ID","画面名","触发","主要元素","关联"],
 [("UI1","翻译 overlay","选区翻译触发","原文、流式译文、术语提示条、QA 警告、接受/编辑/取消、相似度标识(L0/L1/L2)","F2,F4,F5,F7"),
  ("UI2","设置页（Options）","扩展设置入口","模型与密钥、合规策略(域名清单/项目标记)、语言对、术语/TM 路径、引擎连接状态","F10"),
  ("UI3","导入向导","设置页内","文件选择、字段映射、预览、去重策略、导入报告","F11"),
  ("UI4","Popup（快捷面板）","工具栏图标","引擎状态、当前项目/语言对快切、近期记录","—")],
 {"A":10,"B":18,"C":16,"D":58,"E":14},"画面一览（Screens）",center_cols=(1,5),rowh=44)

table(wb.create_sheet("错误处理"),["区分","场景","处理方针","用户可见"],
 [("E-01","引擎未启动/不可达","提示并给启动指引；不降级到云端旁路","是"),
  ("E-02","令牌失效","重新握手；失败则提示重置","是"),
  ("E-03","模型超时/失败","重试 N 次→回退备用模型（非敏感）；仍失败报错","是"),
  ("E-04","合规阻断","敏感且本地不可达→409 中止并说明","是"),
  ("E-05","标签/占位符回填失败","QA 标红，禁止自动写回，要求人工确认","是"),
  ("E-06","TM 写入失败","本地队列重试；不阻塞当前翻译","否（后台）"),
  ("E-07","导入非法行","跳过并记入导入报告","是（报告）"),
  ("E-08","同步冲突","按合并策略处理，冲突项标记待人工","部分")],
 {"A":10,"B":24,"C":50,"D":14},"错误处理方针（Error Handling）",center_cols=(1,4))

table(wb.create_sheet("非功能实现"),["NFR","需求","实现方针"],
 [("NFR-03/04","L0/L1 即时、不调 API","source_hash 索引精确命中；模糊候选预过滤+RapidFuzz，全程本地"),
  ("NFR-05","L2 ~1s 首字","SSE 流式；术语/保护为轻量本地预处理，不阻塞首字"),
  ("NFR-06","L3 高质量","独立路径并行多模型+评审，手动触发，不进默认链"),
  ("NFR-01/02","离线/同步故障可用","本地 SQLite + 本地模型；同步异步旁路，断网不影响主流程"),
  ("NFR-08/09","可维护/接口稳定","扩展与引擎分进程、契约化 API；引擎可独立升级"),
  ("NFR-10/11","迁移/备份","导入管道；SQLite 文件可备份/迁移"),
  ("NFR-12/13","合规/加密","合规 fail-closed；同步通道 TLS；本地数据本地保管")],
 {"A":12,"B":24,"C":62},"非功能实现方针（NFR Realization）",center_cols=(1,))

table(wb.create_sheet("安全设计"),["项","设计"],
 [("本地引擎暴露面","仅绑定 127.0.0.1，不监听外网接口"),
  ("越权访问防护","随机 Bearer Token（扩展+引擎本地）每请求校验；Origin/扩展 ID 白名单，拒绝网页直连"),
  ("CSRF/DNS rebinding","校验 Host/Origin 头；令牌不可被网页脚本读取（仅扩展持有）"),
  ("密钥管理","云端 API Key 存引擎侧（非扩展），不下发到页面上下文"),
  ("合规","敏感判定在网关强制执行，fail-closed；敏感内容永不出本地"),
  ("数据","本地 SQLite；同步链路 TLS+鉴权；可选静态加密（详细设计评估）"),
  ("审计","history 表记录关键变更，支持追溯")],
 {"A":22,"B":80},"安全设计（Security）")

for name,clr in {"封面":NAVY,"目录":"7F7F7F","组件一览":"C00000","接口一览":"2E75B6",
                 "数据实体":"548235","错误处理":"BF8F00","安全设计":"7030A0"}.items():
    wb[name].sheet_properties.tabColor=clr
try: wb.calculation.fullCalcOnLoad=True
except Exception: pass
out1=r"D:\OFCAT\doc\02-基础设计\架构设计\OFCAT_基础设计书_v1.0.xlsx"
wb.save(out1); print("SAVED",out1)

# ====================================================================== 技术选型书
wb2=Workbook()
cv2=wb2.active; cv2.title="封面"
cover(cv2,"技术选型书","OFCAT — AI 增强型 CAT 浏览器工作台",
  [("文档编号","OFCAT-TS-001"),("文档名","技术选型书（含 ADR）"),("版本","第 1.0 版（草稿）"),
   ("创建日","2026-06-25"),("作者","架构师"),("状态","评审前草稿"),
   ("密级","仅社内"),("上游文档","需求定义书 v1.1 / 基础设计书 v1.0")],
  [("选型层数","=COUNTA(选型总览!A4:A100)"),("ADR 条数","=COUNTA(ADR一览!A4:A100)"),
   ("被否决方案","=COUNTA(被否决方案!A4:A100)"),("风险条数","=COUNTA(风险一览!A4:A100)")])

table(wb2.create_sheet("目录"),["No","工作表名","内容"],
 [("1","封面","文档管理信息·摘要"),("2","目录","本工作表"),("3","修订履历","版本履历"),
  ("4","评估维度","维度·权重"),("5","选型总览","各层决策与综合分"),
  ("6","ADR一览","ADR-01~13 决策记录"),("7","被否决方案","否决理由"),("8","风险一览","选型风险与对策")],
 {"A":6,"B":14,"C":58},"目录（Index）",center_cols=(1,),autofilter=False)

table(wb2.create_sheet("修订履历"),["版本","日期","修订者","修订内容"],
 [("1.0","2026-06-25","架构师","初版。各层技术选型对比、评分与决策记录")],
 {"A":8,"B":14,"C":18,"D":80},"修订履历（Revision History）",center_cols=(1,2))

table(wb2.create_sheet("评估维度"),["维度","权重","说明"],
 [("适配性","30%","与需求/约束（C1–C6）、生态的契合度"),
  ("体验/性能","25%","对分层延迟与极简体验的支撑"),
  ("可维护性","20%","长期维护、社区、稳定性"),
  ("落地成本","15%","开发/打包/运维成本"),
  ("合规/安全","10%","对数据合规与安全的支撑")],
 {"A":16,"B":10,"C":62},"评估维度与权重",center_cols=(2,))

r_ov=table(wb2.create_sheet("选型总览"),["层","决策","主要替代","综合分"],
 [("部署形态","薄扩展 + 本地引擎","纯扩展 / 纯桌面端",4.6),
  ("扩展 UI","Svelte + TS（Vite/CRXJS）","React / 原生",4.3),
  ("引擎语言","Python + FastAPI","Node + Fastify",4.4),
  ("编排","LangGraph","自研状态机",4.2),
  ("存储","SQLite + sqlite-vec","Qdrant / DuckDB",4.5),
  ("TM 匹配","RapidFuzz + bge-m3","纯向量 / 纯编辑距离",4.5),
  ("AI 网关","LiteLLM + 自研合规路由","全自研 / OpenRouter",4.3),
  ("本地模型","Qwen 系（Ollama/vLLM）","Llama / DeepSeek 本地",4.4),
  ("嵌入模型","bge-m3","multilingual-e5",4.2),
  ("通信","localhost HTTP+SSE+Token","Native Messaging / WebSocket",4.3),
  ("OCR（后续）","PaddleOCR","tesseract.js / 云 OCR",4.1),
  ("PDF（后续）","pdf.js","PDFium",4.2),
  ("引擎打包","PyInstaller + 托盘","Tauri sidecar / Docker",3.9)],
 {"A":14,"B":30,"C":30,"D":10},"选型总览（Overview）",center_cols=(4,))
ws_ov=wb2["选型总览"]
for rr in range(4,r_ov+1): ws_ov.cell(rr,4).number_format="0.0"
ws_ov.conditional_formatting.add(f"D4:D{r_ov}",
  ColorScaleRule(start_type="num",start_value=3.8,start_color="F8696B",
                 mid_type="num",mid_value=4.3,mid_color="FFEB84",
                 end_type="num",end_value=4.6,end_color="63BE7B"))

table(wb2.create_sheet("ADR一览"),["ADR-ID","主题","候选","决策","理由","取舍/备注"],
 [("ADR-01","部署形态","纯扩展/薄扩展+引擎/纯桌面","薄扩展+本地引擎","一次性解决SQLite/OCR/本地LLM/编排落地；扩展轻稳可独立升级","需装伴随程序（已确认可接受）"),
  ("ADR-02","扩展 UI","React/Svelte/原生","Svelte+TS（Vite/CRXJS）","overlay 需极小体积高性能，Svelte 无虚拟DOM产物小","团队熟 React 可换，架构不依赖"),
  ("ADR-03","引擎语言","Python/Node","Python+FastAPI","PaddleOCR/LangGraph/嵌入/sqlite-vec/RapidFuzz 均 Python 一等；FastAPI async+SSE","打包不如 Node 轻→PyInstaller+托盘"),
  ("ADR-04","编排","LangGraph/自研状态机","LangGraph","管道本是多节点有状态；多 Agent 可平滑扩展避免重写","MVP 略重，仅用图与状态，节点轻量"),
  ("ADR-05","存储","SQLite+sqlite-vec/Qdrant/DuckDB","SQLite+sqlite-vec","本地权威副本零运维可迁移；同文件内向量 KNN","超大规模不及 Qdrant，当前量级足够"),
  ("ADR-06","TM 匹配","编辑距离/向量/混合","RapidFuzz 主+bge-m3 辅","fuzzy% 可解释符合 CAT 预期；纯向量误判改数字句","维护两套索引，向量为可选增强"),
  ("ADR-07","AI 网关","全自研/LiteLLM/OpenRouter","LiteLLM+自研合规路由","LiteLLM 省多家适配；合规须自研 fail-closed","OpenRouter 过第三方违规排除"),
  ("ADR-08","本地模型","Qwen/Llama/DeepSeek","Qwen 系（Ollama/vLLM）","中日英最强、指令遵循好；Ollama 省事 vLLM 并发优","显存受限选小档；OpenAI 兼容切换无成本"),
  ("ADR-09","嵌入模型","bge-m3/multilingual-e5","bge-m3","多语种长文本检索强，CJK 稳","e5 亦可"),
  ("ADR-10","通信","Native Messaging/HTTP/WebSocket","localhost HTTP+SSE+Token","简单可调试、SSE 流式；令牌+Origin 防越权","Native Messaging 流式受限调试难为备选"),
  ("ADR-11","OCR（后续）","PaddleOCR/tesseract.js/云","PaddleOCR","中日英、表格版面强；必须在引擎","tesseract 精度不足/云有合规风险"),
  ("ADR-12","PDF（后续）","pdf.js/PDFium","pdf.js","渲染+文本层；版面对齐为主要工作量","PDFium 集成更重"),
  ("ADR-13","引擎打包","PyInstaller/Tauri/Docker","PyInstaller+托盘","单可执行+托盘常驻+自启","首启慢/体积大可接受；后续评估 Tauri")],
 {"A":10,"B":14,"C":28,"D":24,"E":44,"F":30},"选型决策记录（ADR）",center_cols=(1,),rowh=46)

table(wb2.create_sheet("被否决方案"),["方案","否决理由"],
 [("纯浏览器扩展（WASM SQLite/OCR）","MV3 沙箱无法稳定承载 OCR/本地 LLM/长任务，体验与能力受限"),
  ("Qdrant 向量库","独立服务需运维，违背本地零运维（C4）；当前量级 sqlite-vec 足够"),
  ("OpenRouter 等云中转网关","内容经第三方，违反数据合规（C3）"),
  ("多模型投票进默认链","与极简/低延迟体验冲突，改为手动 L3"),
  ("纯向量 TM 匹配","fuzzy% 不可解释，改数字类高相似句被判低，违背 CAT 预期")],
 {"A":34,"B":66},"被否决方案（Rejected）")

r_rk=table(wb2.create_sheet("风险一览"),["风险ID","风险","影响","对策"],
 [("R-01","富文本编辑器写回破坏文档","高","早期对 ProseMirror/CodeMirror 做写回 PoC（F8）"),
  ("R-02","PyInstaller 包体大/杀软误报","中","代码签名；评估 Tauri 壳"),
  ("R-03","本地模型显存/性能不足","中","提供模型档位选择；vLLM 共享部署"),
  ("R-04","sqlite-vec 规模上限","中","监控规模，超阈值切分或迁移向量库"),
  ("R-05","合规误判（漏判敏感）","高","fail-closed + 域名清单可维护 + 审计（O6）")],
 {"A":10,"B":40,"C":10,"D":46},"选型相关风险（Risks）",center_cols=(1,3))
ws_rk=wb2["风险一览"]
ws_rk.conditional_formatting.add(f"C4:C{r_rk}",CellIsRule(operator="equal",formula=['"高"'],fill=fill("FFC7CE"),font=f(10,True,"9C0006")))
ws_rk.conditional_formatting.add(f"C4:C{r_rk}",CellIsRule(operator="equal",formula=['"中"'],fill=fill("FFEB9C"),font=f(10,True,"9C6500")))

for name,clr in {"封面":NAVY,"目录":"7F7F7F","选型总览":"C00000","ADR一览":"2E75B6","风险一览":"BF8F00"}.items():
    wb2[name].sheet_properties.tabColor=clr
try: wb2.calculation.fullCalcOnLoad=True
except Exception: pass
out2=r"D:\OFCAT\doc\02-基础设计\技术选型\OFCAT_技术选型书_v1.0.xlsx"
wb2.save(out2); print("SAVED",out2)
