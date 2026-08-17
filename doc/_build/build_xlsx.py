# -*- coding: utf-8 -*-
"""OFCAT 需求定义书 -> 一线大厂规格的中文 Excel 工作簿生成（保留日系 SIer 结构与颗粒度）"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

FONT = "Microsoft YaHei"
NAVY   = "1F3864"
HEADER = "2E5496"
SUB    = "8EAADB"
LIGHT  = "D9E1F2"
ZEBRA  = "F2F5FB"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=10, b=False, color="000000"): return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True):   return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def topleft():         return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

wb = Workbook()

def style_header(ws, row, ncols, fillc=HEADER):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(fillc); cell.font = f(10, True, WHITE)
        cell.alignment = center(); cell.border = BORDER

def grid(ws, r1, r2, c1, c2, zebra=True):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.vertical is None:
                cell.alignment = topleft()
            if zebra and (r-r1) % 2 == 1 and cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = fill(ZEBRA)

def widths(ws, mp):
    for col, w in mp.items(): ws.column_dimensions[col].width = w

def title_block(ws, title, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1,1, title); c.fill=fill(NAVY); c.font=f(13, True, WHITE)
    c.alignment=center(False); ws.row_dimensions[1].height=28

# ============================================================ 1. 封面
ws = wb.active; ws.title = "封面"
ws.sheet_view.showGridLines = False
widths(ws, {"A":3,"B":26,"C":52,"D":3})
ws.merge_cells("B2:C2"); ws["B2"]="需求定义书"; ws["B2"].font=f(22,True,NAVY); ws["B2"].alignment=center(False)
ws.row_dimensions[2].height=44
ws.merge_cells("B3:C3"); ws["B3"]="OFCAT — AI 增强型 CAT 浏览器工作台"; ws["B3"].font=f(12,True,"404040"); ws["B3"].alignment=center(False)
ws.row_dimensions[3].height=24
meta=[("文档编号","OFCAT-REQ-001"),("文档名","需求定义书"),
      ("版本","第 1.1 版（草稿）"),("创建日","2026-06-25"),
      ("作者","架构师"),("状态","评审前草稿"),
      ("密级","仅社内（Confidential / Internal Only）"),
      ("上游文档","初版构想 / 需求规格说明 v0.1")]
r=5
for k,v in meta:
    ws.cell(r,2,k).fill=fill(SUB); ws.cell(r,2).font=f(10,True,WHITE); ws.cell(r,2).alignment=center()
    ws.cell(r,3,v).alignment=left(); ws.cell(r,3).font=f(10)
    ws.cell(r,2).border=BORDER; ws.cell(r,3).border=BORDER; ws.row_dimensions[r].height=20; r+=1
r+=1
ws.cell(r,2,"摘要（自动统计）").font=f(10,True,NAVY)
r+=1
stats=[("功能条数","=COUNTA(功能一览!A4:A100)"),
       ("MVP 必须功能数",'=COUNTIF(功能一览!D4:D100,"必须")'),
       ("未确定课题数",'=COUNTIF(课题管理!F4:F100,"未确定")'),
       ("术语登记数","=COUNTA(术语表!A4:A200)")]
for k,fml in stats:
    ws.cell(r,2,k).fill=fill(LIGHT); ws.cell(r,2).font=f(10,True); ws.cell(r,2).alignment=center()
    ws.cell(r,3,fml).font=f(10,True,NAVY); ws.cell(r,3).alignment=left(False)
    ws.cell(r,2).border=BORDER; ws.cell(r,3).border=BORDER; ws.row_dimensions[r].height=20; r+=1

# ============================================================ 2. 目录
ws = wb.create_sheet("目录"); ws.sheet_view.showGridLines=False
title_block(ws,"目录（Index）",4)
widths(ws,{"A":3,"B":10,"C":34,"D":58})
for i,h in enumerate(["No","工作表名","内容"]): ws.cell(3,i+2,h)
style_header(ws,3,4); ws.cell(3,1).fill=fill(HEADER); ws.cell(3,1).border=BORDER
toc=[("1","封面","文档管理信息·摘要"),
     ("2","目录","本工作表"),
     ("3","修订履历","版本·修订内容履历"),
     ("4","术语表","术语·定义"),
     ("5","约束条件","已确定架构约束 C1–C6"),
     ("6","功能一览","功能ID·优先级·阶段（F1–F11）"),
     ("7","功能规格","功能明细卡（输入/处理/输出/业务规则/例外）"),
     ("8","非功能需求","按 IPA 等级分类的非功能需求"),
     ("9","课题管理","Open Issues O1–O6（状态·期限）"),
     ("10","可追溯性","课题→效果→功能 对应矩阵"),
     ("11","里程碑","M1–M5 发布计划")]
r=4
for no,sh,desc in toc:
    ws.cell(r,2,no); ws.cell(r,3,sh); ws.cell(r,4,desc); r+=1
grid(ws,4,r-1,1,4)
for rr in range(4,r): ws.cell(rr,2).alignment=center()
ws.freeze_panes="A4"

# ============================================================ 3. 修订履历
ws=wb.create_sheet("修订履历"); ws.sheet_view.showGridLines=False
title_block(ws,"修订履历（Revision History）",5)
widths(ws,{"A":8,"B":14,"C":18,"D":66,"E":12})
for i,h in enumerate(["版本","日期","修订者","修订内容","审批"]): ws.cell(3,i+1,h)
style_header(ws,3,5)
rev=[("0.1","2026-06-25","架构师","需求沟通后初稿",""),
     ("1.0","2026-06-25","架构师","按日系 SIer 需求定义书规范与颗粒度重整；补充功能规格卡/非功能需求/术语定义/课题管理",""),
     ("1.1","2026-06-25","架构师","全文中文化（原日语正文译为中文，保留 SIer 结构与颗粒度）","")]
r=4
for v in rev:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    r+=1
grid(ws,4,r-1,1,5)
for rr in range(4,r):
    ws.cell(rr,1).alignment=center(); ws.cell(rr,2).alignment=center(); ws.cell(rr,5).alignment=center()
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:E{r-1}"

# ============================================================ 4. 术语表
ws=wb.create_sheet("术语表"); ws.sheet_view.showGridLines=False
title_block(ws,"术语表（Glossary）",3)
widths(ws,{"A":22,"B":34,"C":74})
for i,h in enumerate(["术语","英文/全称","定义"]): ws.cell(3,i+1,h)
style_header(ws,3,3)
terms=[("CAT","Computer-Assisted Translation","计算机辅助翻译。译员主导、工具辅助的翻译作业方式。"),
("TM","Translation Memory（翻译记忆库）","翻译记忆库。已确认的原文-译文对的复用库。"),
("术语库","Termbase","规定特定术语在特定领域的统一译法。"),
("句段","Segment","翻译作业的最小单位（通常为一句）。"),
("模糊匹配","Fuzzy Match","TM 中与当前原文相似但非完全一致的命中（以相似度 % 表示）。"),
("精确匹配","100% Match","TM 中与当前原文完全一致的命中。"),
("占位符","Placeholder","{count}、%s、{{name}} 等需原样保留的可变标记。"),
("标签保护","Tag Protection","翻译过程中保护 HTML/标记标签不被破坏的机制。"),
("薄扩展","Thin Extension","仅承担 UI/捕获、不含重逻辑的浏览器扩展。"),
("本地引擎","Local Engine","本机运行的 CAT 核心服务（TM/术语/QA/编排/OCR/本地 LLM）。"),
("AI 网关","AI Gateway","统一封装云端/本地各 LLM、负责模型路由与合规开关的组件。"),
("合规开关","Compliance Switch","按域名/项目把敏感内容强制路由到本地模型的机制。")]
r=4
for v in terms:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    r+=1
grid(ws,4,r-1,1,3)
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:C{r-1}"

# ============================================================ 5. 约束条件
ws=wb.create_sheet("约束条件"); ws.sheet_view.showGridLines=False
title_block(ws,"约束条件一览（Constraints）",3)
widths(ws,{"A":8,"B":24,"C":92})
for i,h in enumerate(["ID","约束名","内容"]): ws.cell(3,i+1,h)
style_header(ws,3,3)
cons=[("C1","部署形态","薄扩展 + 本地引擎。扩展仅做选区捕获/UI/快捷键；CAT 引擎放本机服务，经 127.0.0.1 或 Native Messaging 通信。"),
("C2","产品定位","社内工具。重稳定与可维护性；账号体系·计费·对外多租户为对象外。"),
("C3","数据合规","AI 网关必须支持模型路由 + 合规开关。敏感内容按域名/项目单位强制路由到本地模型。云端/本地以 OpenAI 兼容协议统一。"),
("C4","数据归属","TM/术语库本地为主 + 可选同步。本机 SQLite 为权威副本（离线/快速）；中心同步服务同步术语库与核心 TM（非实时协作）。"),
("C5","多语种","TM/术语条目主键必须含 源语言→目标语言 维度。不写死中/日/英。"),
("C6","迁移数据","存量 TM/术语格式杂乱。需要清洗+字段映射+去重的导入管道。不以标准 TMX/TBX 为前提。")]
r=4
for v in cons:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    r+=1
grid(ws,4,r-1,1,3)
for rr in range(4,r): ws.cell(rr,1).alignment=center()
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:C{r-1}"

# ============================================================ 6. 功能一览
ws=wb.create_sheet("功能一览"); ws.sheet_view.showGridLines=False
title_block(ws,"功能一览（Function List）",6)
widths(ws,{"A":10,"B":24,"C":58,"D":10,"E":10,"F":16})
for i,h in enumerate(["功能ID","功能名","概要","优先级","阶段","关联功能"]): ws.cell(3,i+1,h)
style_header(ws,3,6)
funcs=[("F1","选区捕获","网页文本选区捕获与触发","必须","MVP","F2,F8,F10"),
("F2","TM 匹配","精确匹配/模糊匹配","必须","MVP","F6,F9"),
("F3","术语匹配与注入","术语提示与约束注入","必须","MVP","F4,F6"),
("F4","术语强制校验","译后术语校验/替换","必须","MVP","F3,F7"),
("F5","标签/占位符保护","标签占位符保护与校验","必须","MVP","F6,F7"),
("F6","单模型流式翻译","默认流式翻译（L2）","必须","MVP","F3,F5,F10"),
("F7","行内 overlay 编辑","译文行内展示与编辑","必须","MVP","F4,F5,F8"),
("F8","写回页面","安全写回选区","必须","MVP","F1,F7"),
("F9","保存进 TM","确认译文回存 TM","必须","MVP","F2,F11"),
("F10","合规路由","云端/本地模型路由","必须","MVP","F6"),
("F11","存量数据导入","存量术语/TM 导入","高","M2","F2,F9")]
r=4
for v in funcs:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    r+=1
grid(ws,4,r-1,1,6)
for rr in range(4,r):
    for cc in (1,4,5,6): ws.cell(rr,cc).alignment=center()
dv=DataValidation(type="list",formula1='"必须,高,中,低"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"D4:D{r-1}")
dvp=DataValidation(type="list",formula1='"MVP,M2,M3,M4,M5"',allow_blank=True); ws.add_data_validation(dvp); dvp.add(f"E4:E{r-1}")
ws.conditional_formatting.add(f"D4:D{r-1}", CellIsRule(operator="equal",formula=['"必须"'],fill=fill("FFC7CE"),font=f(10,True,"9C0006")))
ws.conditional_formatting.add(f"D4:D{r-1}", CellIsRule(operator="equal",formula=['"高"'],fill=fill("FFEB9C"),font=f(10,True,"9C6500")))
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:F{r-1}"

# ============================================================ 7. 功能规格
ws=wb.create_sheet("功能规格"); ws.sheet_view.showGridLines=False
title_block(ws,"功能规格（Function Specification — 明细）",10)
cols=["功能ID","功能名","触发","输入","处理","输出","业务规则","例外/错误","优先级","关联功能"]
widths(ws,{"A":9,"B":22,"C":30,"D":30,"E":42,"F":26,"G":40,"H":36,"I":9,"J":14})
for i,h in enumerate(cols): ws.cell(3,i+1,h)
style_header(ws,3,len(cols))
spec=[
("F1","选区捕获","①快捷键 ②右键菜单「AI Translate」③选区旁浮动按钮","选中文本/所在 DOM 节点/页面 URL·域名/(可能的)语言信息","获取选区文本与上下文；识别源语言；判定域名是否敏感(供 F10)","标准化句段(原文+元信息)","空选区/纯符号不触发；过长选区按句段切分","不可编辑区域·跨 iframe·富文本节点→标记「需特殊写回」(F8)","必须","F2,F8,F10"),
("F2","TM 匹配","句段确定时","原文/源语言/目标语言/(可选)领域·项目","①精确匹配检索 ②模糊匹配(以编辑距离/相似度为主、语义向量为辅)；计算相似度%","命中候选(译文/相似度%/来源)；100%→L0，模糊→L1","精确匹配不调用 API、即时填充；模糊(默认 85%以上)高亮差异+等待确认","TM 未初始化/为空→视为未命中转 L2","必须","F6,F9"),
("F3","术语匹配与注入","句段确定时","原文/源·目标语言/领域","术语抽取→术语库匹配→overlay 提示+作为约束注入翻译提示词","命中术语列表(原词/约定译法/领域)","按多语种维度(C5)匹配；领域有不同译法时领域设定优先","术语库未设定→不注入继续翻译","必须","F4,F6"),
("F4","术语强制校验","翻译结果生成后","翻译结果/F3 命中术语列表","扫描译文，检出未使用约定译法的术语；标红+一键替换","违反列表+修正候选","不依赖模型自觉，确保确定性","上下文不可替换→仅警告(不自动替换)","必须","F3,F7"),
("F5","标签/占位符保护","翻译前后","原文(含标签/占位符)","译前替换为哨兵 token→翻译→译后回填→校验数量与完整性","标签/占位符保全的译文","数量不一致视为 QA 错误(F7 警告)；位置可移动，欠缺/增加不可","无法回填(数量不一致)→向用户警告并要求确认","必须","F6,F7"),
("F6","单模型流式翻译","TM 未命中(L2)","受保护句段/术语约束/源·目标语言/领域","经 AI 网关向模型请求；流式输出；~1s 内显示首字","译文(流式)","默认单模型；多模型投票(L3)为手动触发的独立功能(后续)","响应失败→重试/回退到备用模型(可配置)","必须","F3,F5,F10"),
("F7","行内 overlay 编辑","收到译文时","译文/术语提示/QA 警告(F4·F5)","overlay 展示；高亮术语命中/QA 违反；可编辑；回车确认","确认译文","可仅用键盘完成(操作数最小化)","用户取消→不写回/不保存","必须","F4,F5,F8"),
("F8","写回页面","确认时","确认译文/目标 DOM 节点类型","普通文本直接替换；富文本编辑器经其 API/模拟输入安全写入","更新后的页面","禁止直接改写 innerText(防止文档破坏)","不可写入区域→仅提供译文供复制","必须","F1,F7"),
("F9","保存进 TM","确认后","原文/确认译文/源·目标语言/领域/项目","按 源语言→目标语言 维度(C5)保存进本机 SQLite TM；重复则更新","更新后的 TM","同步对象的核心 TM 推送至中心服务(C4，非实时)","保存失败→本地队列重试","必须","F2,F11"),
("F10","合规路由","翻译请求时","域名/项目/敏感判定/模型设定","敏感为真→强制路由本地模型；否则→默认(可云端)；OpenAI 兼容仅切换 base_url","选定的模型端点","敏感内容上云率 0%；默认策略可配置(O6)","本地模型不可达→中止翻译并警告(不把敏感内容流向云端)","必须","F6"),
("F11","存量数据导入","导入操作时","Excel/CSV/TMX/TBX 等(格式不统一)","清洗→字段映射(源/目标/领域)→去重→导入本机 SQLite","导入结果报告(成功/重复/错误条数)","不以标准 TMX/TBX 为前提(C6)；映射定义须经用户确认","非法行跳过并记入报告","高","F2,F9"),
]
r=4
for v in spec:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    ws.row_dimensions[r].height=70
    r+=1
grid(ws,4,r-1,1,len(cols))
for rr in range(4,r):
    ws.cell(rr,1).alignment=center(); ws.cell(rr,9).alignment=center(); ws.cell(rr,10).alignment=center()
ws.freeze_panes="C4"; ws.auto_filter.ref=f"A3:J{r-1}"

# ============================================================ 8. 非功能需求
ws=wb.create_sheet("非功能需求"); ws.sheet_view.showGridLines=False
title_block(ws,"非功能需求一览（Non-Functional Requirements / 按 IPA 等级分类）",5)
widths(ws,{"A":10,"B":18,"C":28,"D":72,"E":14})
for i,h in enumerate(["NFR-ID","类别","项目","需求内容","目标/区分"]): ws.cell(3,i+1,h)
style_header(ws,3,5)
nfr=[("NFR-01","可用性","离线运行","TM/术语匹配·本地模型路径在离线时仍可用","必达"),
("NFR-02","可用性","同步故障时继续","中心同步服务停止时本地作业仍可继续","必达"),
("NFR-03","性能·扩展性","L0 精确匹配","TM 精确匹配即时响应·不调用云端 API","即时"),
("NFR-04","性能·扩展性","L1 模糊匹配","TM 模糊匹配即时填充+等待确认","即时"),
("NFR-05","性能·扩展性","L2 默认翻译","TM 未命中以单模型流式 ~1s 内显示首字","~1s"),
("NFR-06","性能·扩展性","L3 高质量","多模型投票+评审为手动触发","5〜10s"),
("NFR-07","性能·扩展性","TM 索引规模","按 TM 条数预想规模(O1 确定后)确定索引方式","O1 依赖"),
("NFR-08","运维·可维护性","职责分离","薄扩展与本地引擎分离，可独立升级","必达"),
("NFR-09","运维·可维护性","接口契约稳定","稳定扩展↔引擎/AI 网关的接口契约","必达"),
("NFR-10","迁移性","存量导入","提供存量术语/TM 的导入管道(F11)","必达"),
("NFR-11","迁移性","备份","本机 SQLite 可备份/迁移","必达"),
("NFR-12","安全","上云限制","敏感域名/项目内容不发往云端","上云率 0%"),
("NFR-13","安全","加密","同步通道加密；本地数据本地保管","必达"),
("NFR-14","系统环境","客户端","Chromium 系浏览器 + 本地引擎","前提"),
("NFR-15","系统环境","模型","云端 + 本地 LLM(推荐 Qwen 系·Ollama/vLLM)","前提")]
r=4
for v in nfr:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    r+=1
grid(ws,4,r-1,1,5)
for rr in range(4,r):
    ws.cell(rr,1).alignment=center(); ws.cell(rr,5).alignment=center()
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:E{r-1}"

# ============================================================ 9. 课题管理
ws=wb.create_sheet("课题管理"); ws.sheet_view.showGridLines=False
title_block(ws,"课题管理表（Open Issues）",7)
widths(ws,{"A":8,"B":48,"C":34,"D":12,"E":16,"F":12,"G":22})
for i,h in enumerate(["ID","课题","影响","负责人","期限","状态","备注"]): ws.cell(3,i+1,h)
style_header(ws,3,7)
issues=[("O1","团队规模/日均翻译量/TM 预想条数规模","TM 索引·同步服务规格","","基础设计前","未确定",""),
("O2","目标浏览器：仅 Edge / Edge + Chrome","打包·兼容性","","基础设计前","未确定",""),
("O3","云端模型是否有社内账号·预算约束","默认模型·成本设计","","基础设计前","未确定",""),
("O4","中心同步服务的运维主体·部署(内网/云)","C4 落地·安全设计","","M2 前","未确定",""),
("O5","存量数据的真实样本/格式","导入管道设计(F11)","","M2 前","未确定",""),
("O6","「敏感」的定义(域名清单 / 项目标记)","合规开关默认策略(F10)","","MVP 前","未确定","最优先")]
r=4
for v in issues:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    r+=1
grid(ws,4,r-1,1,7)
for rr in range(4,r):
    for cc in (1,4,5,6): ws.cell(rr,cc).alignment=center()
dvs=DataValidation(type="list",formula1='"未确定,处理中,确定,保留"',allow_blank=True); ws.add_data_validation(dvs); dvs.add(f"F4:F{r-1}")
ws.conditional_formatting.add(f"F4:F{r-1}", CellIsRule(operator="equal",formula=['"未确定"'],fill=fill("FFC7CE"),font=f(10,True,"9C0006")))
ws.conditional_formatting.add(f"F4:F{r-1}", CellIsRule(operator="equal",formula=['"确定"'],fill=fill("C6EFCE"),font=f(10,True,"006100")))
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:G{r-1}"

# ============================================================ 10. 可追溯性
ws=wb.create_sheet("可追溯性"); ws.sheet_view.showGridLines=False
title_block(ws,"可追溯性矩阵（课题→效果→功能）",4)
widths(ws,{"A":18,"B":46,"C":18,"D":40})
for i,h in enumerate(["课题","期望效果","关联约束","对应功能"]): ws.cell(3,i+1,h)
style_header(ws,3,4)
trace=[("课题1 操作步骤多","效果1 操作数≤2 / 效果5 效率3〜10倍","—","F1,F6,F7,F8"),
("课题2 通用插件缺 CAT 能力","效果2 精确匹配即时 / 效果3 术语符合100%","C5","F2,F3,F4,F5,F9"),
("课题3 AI 术语不统一/语义污染","效果3 术语符合率 100%","C5","F3,F4"),
("课题4 敏感内容上云风险","效果4 上云率 0%","C3","F10"),
("课题5 存量数据杂乱","—（基础整备）","C6","F11")]
r=4
for v in trace:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    r+=1
grid(ws,4,r-1,1,4)
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:D{r-1}"

# ============================================================ 11. 里程碑
ws=wb.create_sheet("里程碑"); ws.sheet_view.showGridLines=False
title_block(ws,"里程碑（Milestones）",4)
widths(ws,{"A":10,"B":34,"C":40,"D":30})
for i,h in enumerate(["阶段","内容","主要功能","备注"]): ws.cell(3,i+1,h)
style_header(ws,3,4)
ms=[("M1","MVP 闭环（网页选区翻译）","F1–F10","最优先。O6 须在 MVP 前确定"),
("M2","数据底座（存量导入+数据层）","F11 + 同步","O4,O5 须在 M2 前确定"),
("M3","文档场景","PDF 双语对照",""),
("M4","工作流场景","Jira 反馈模式",""),
("M5","增强能力","OCR / 多模型投票 / 本地 LLM / 文档级实体记忆","")]
r=4
for v in ms:
    for i,x in enumerate(v): ws.cell(r,i+1,x)
    r+=1
grid(ws,4,r-1,1,4)
for rr in range(4,r): ws.cell(rr,1).alignment=center()
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:D{r-1}"

# tab colors
for name,clr in {"封面":NAVY,"目录":"7F7F7F","功能一览":"C00000","功能规格":"C00000",
                 "非功能需求":"548235","课题管理":"BF8F00","可追溯性":"7030A0"}.items():
    wb[name].sheet_properties.tabColor=clr

try:
    wb.calculation.fullCalcOnLoad=True
except Exception:
    pass

out=r"D:\OFCAT\doc\01-需求\需求规格说明\OFCAT_需求定义书_v1.1.xlsx"
wb.save(out)
print("SAVED", out)
