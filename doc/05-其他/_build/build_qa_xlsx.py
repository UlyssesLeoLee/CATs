# -*- coding: utf-8 -*-
"""CATs 实施前 QA 登记册 -> Excel 工作簿生成
工作表: 封面 / 概览 / 阻塞实施项 / 需求 / 架构 / 技术选型 / 跨文档一致性 / 假设与默认值 /
       数据与规模 / 环境与部署 / 合规与安全 / 团队与资源 / 流程与协作 / 决策矩阵 / 后续动作
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Microsoft YaHei"
NAVY   = "1F3864"
HEADER = "2E5496"
SUB    = "8EAADB"
LIGHT  = "D9E1F2"
ZEBRA  = "F2F5FB"
WHITE  = "FFFFFF"
RED    = "C00000"
ORANGE = "BF8F00"
GREEN  = "548235"
GREY   = "A6A6A6"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True):   return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def topleft():         return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def title_block(ws, title, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.fill = fill(NAVY); c.font = f(13, True, WHITE); c.alignment = center(False)
    ws.row_dimensions[1].height = 28

def style_header(ws, row, ncols, fillc=HEADER):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(fillc); cell.font = f(10, True, WHITE)
        cell.alignment = center(); cell.border = BORDER

def grid(ws, r1, r2, c1, c2, zebra=True):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.vertical is None:
                cell.alignment = topleft()
            if zebra and (r - r1) % 2 == 1:
                if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = fill(ZEBRA)

def widths(ws, mp):
    for col, w in mp.items():
        ws.column_dimensions[col].width = w

# ============================================================
wb = Workbook()

# ---------------- 封面 ----------------
ws = wb.active
ws.title = "封面"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 26, "C": 56, "D": 3})
ws.merge_cells("B2:C2")
ws["B2"] = "实施前 QA 登记册"
ws["B2"].font = f(22, True, NAVY); ws["B2"].alignment = center(False)
ws.row_dimensions[2].height = 44
ws.merge_cells("B3:C3")
ws["B3"] = "Pre-Implementation Concern Register"
ws["B3"].font = f(12, True, "404040"); ws["B3"].alignment = center(False)
ws.row_dimensions[3].height = 24

meta = [
    ("文档编号", "CATs-QA-PRE-001"),
    ("文档名", "实施前 QA 登记册"),
    ("版本", "第 1.0 版（草稿）"),
    ("创建日", "2026-08-19"),
    ("作者", "架构师 / 测试负责人"),
    ("状态", "评审前草稿"),
    ("密级", "仅社内"),
    ("性质", "pre-mortem：在编码启动前记录所有顾虑、疑问、未明确项"),
    ("上游文档", "全部 CATs 需求/基础/详细设计文档 + 横向补充文档（可热插拔/测试设计/Rust 选型）"),
]
r = 5
for k, v in meta:
    ws.cell(r, 2, k).fill = fill(SUB); ws.cell(r, 2).font = f(10, True, WHITE)
    ws.cell(r, 2).alignment = center()
    ws.cell(r, 3, v).alignment = left(); ws.cell(r, 3).font = f(10)
    ws.cell(r, 2).border = BORDER; ws.cell(r, 3).border = BORDER
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
ws.cell(r, 2, "v1.0 主要内容").font = f(10, True, NAVY)
r += 1
items = [
    "① 58 项关心事项，按 10 大类别组织（需求/架构/技术选型/一致性/假设/数据/部署/合规/团队/流程）",
    "② 12 项 P0 阻塞实施项，M1 编码启动前必须 Closed",
    "③ 决策矩阵按 8 个决策方归类（产品/架构/运维/DBA/测试/合规/Rust Lead/管理层）",
    "④ 评审会议程建议（1.5 小时，按决策方议程）",
    "⑤ 后续动作清单（评审前/中/后）",
]
for c_text in items:
    ws.cell(r, 2, c_text).font = f(9); ws.cell(r, 2).alignment = topleft()
    ws.cell(r, 2).border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 18
    r += 1

# ---------------- 概览 ----------------
ws = wb.create_sheet("概览")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 24, "B": 14, "C": 14, "D": 14, "E": 14})
title_block(ws, "关心事项概览（按类别 × 优先级）", 5)
headers = ["类别", "数量", "P0 阻塞", "P1 重要", "P2 一般"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
overview = [
    ("1. 需求层面", 7, 2, 4, 1),
    ("2. 架构层面", 9, 2, 5, 2),
    ("3. 技术选型", 8, 1, 5, 2),
    ("4. 跨文档一致性", 6, 0, 3, 3),
    ("5. 假设与默认值", 8, 1, 5, 2),
    ("6. 数据与规模", 5, 2, 2, 1),
    ("7. 环境与部署", 4, 1, 2, 1),
    ("8. 合规与安全", 4, 2, 2, 0),
    ("9. 团队与资源", 3, 0, 2, 1),
    ("10. 流程与协作", 4, 1, 2, 1),
    ("合计", 58, 12, 32, 14),
]
r = 4
for row in overview:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    # 染色
    if row[0] == "合计":
        for c in range(1, 6):
            ws.cell(r, c).fill = fill(LIGHT)
            ws.cell(r, c).font = f(10, True)
    r += 1
grid(ws, 4, r - 1, 1, 5)

r += 2
ws.cell(r, 1, "状态分布").font = f(11, True, NAVY)
r += 1
status_h = ["状态", "数量", "占比"]
for i, h in enumerate(status_h, 1):
    ws.cell(r, i, h)
style_header(ws, r, 3)
r += 1
status_rows = [
    ("Open", 50, "86%"),
    ("Decision Required", 6, "10%"),
    ("Decided", 2, "3%"),
    ("Won't Fix", 0, "0%"),
]
for row in status_rows:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, r - len(status_rows), r - 1, 1, 3)

# ---------------- 阻塞实施项 ----------------
ws = wb.create_sheet("阻塞实施项")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 10, "B": 50, "C": 18, "D": 20})
title_block(ws, "P0 阻塞实施项（M1 启动前必须 Closed）", 4)
headers = ["编号", "关心事项", "决策方", "期望关闭时点"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
p0 = [
    ("QA-001", "O1：团队规模/日均翻译量级/TM 预期条数", "产品 + 架构", "评审会 D-3"),
    ("QA-002", "O2：目标浏览器范围（Edge only / Edge+Chrome）", "产品", "评审会 D-3"),
    ("QA-011", "TM 索引策略选型（分桶/全表/scale-out）", "架构", "QA-001 关闭后"),
    ("QA-012", "微服务拆分粒度（16 个）", "架构 + 运维", "评审会 D-3"),
    ("QA-021", "Tokio 异步运行时唯一选型确认", "架构 + Rust Lead", "评审会 D-3"),
    ("QA-041", "PG + pgvector 性能基线", "架构 + DBA", "M1 Sprint 1 末"),
    ("QA-042", "K3s 3 控制面 HA 评估", "架构 + 运维", "评审会 D-3"),
    ("QA-051", "本地 LLM 资源（模型/GPU/位置）", "架构 + 运维", "M1 Sprint 1 中"),
    ("QA-052", "合规敏感项目样本数据来源", "测试 + 合规", "M1 Sprint 1 中"),
    ("QA-061", "生产 TLS / 内部 CA / 域名规划", "运维", "评审会 D-3"),
    ("QA-071", "等保/行业认证需求", "合规", "评审会 D-3"),
    ("QA-101", "跨团队契约变更协调机制", "架构", "评审会 D-3"),
]
r = 4
for row in p0:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
        if i == 1:
            cell.fill = fill("F8CBAD")
            cell.font = f(10, True)
    r += 1
grid(ws, 4, r - 1, 1, 4)
ws.freeze_panes = "A4"

# ---------------- 通用关心事项表写入函数 ----------------
def write_concerns(ws, title, items, headers):
    """items: list of (编号, 关心事项, 详细描述, 影响范围, 当前假设, 建议方案, 决策方, 优先级, 状态)"""
    widths(ws, {"A": 10, "B": 26, "C": 50, "D": 24, "E": 26, "F": 32, "G": 18, "H": 8, "I": 10})
    title_block(ws, title, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    r = 4
    for row in items:
        for i, v in enumerate(row, 1):
            cell = ws.cell(r, i, v)
            cell.alignment = topleft()
        # 优先级染色
        pri = row[7] if len(row) > 7 else None
        if pri == "P0":
            ws.cell(r, 8).fill = fill("F8CBAD")
            ws.cell(r, 8).font = f(10, True)
        elif pri == "P1":
            ws.cell(r, 8).fill = fill("FFE699")
        elif pri == "P2":
            ws.cell(r, 8).fill = fill("C6E0B4")
        # 状态染色
        st = row[8] if len(row) > 8 else None
        if st == "Decided":
            ws.cell(r, 9).fill = fill("C6E0B4")
        elif st and "Decision" in st:
            ws.cell(r, 9).fill = fill("FFE699")
        r += 1
    grid(ws, 4, r - 1, 1, len(headers))
    ws.freeze_panes = "A4"

# ---------------- 需求层面 ----------------
ws = wb.create_sheet("01-需求层面")
hdr = ["编号", "关心事项", "详细描述", "影响范围", "当前假设", "建议方案", "决策方", "优先级", "状态"]
items_req = [
    ("QA-001", "O1:团队规模/日均翻译量级/TM 预期条数", "决定 TM 索引方案(PostgreSQL 分桶/全表)和同步服务规格", "架构 §5、§17 容量规划;测试设计 SLO", "默认:50-3000 用户,10k 句对", "A. 给出具体数字(优) B. 设定小/中/大三档(可) C. 维持默认", "产品+架构", "P0", "Open"),
    ("QA-002", "O2:目标浏览器范围", "决定 WebView2/WebKitGTK 兼容矩阵", "架构 §17;测试兼容性", "默认:Chromium 内核统一", "采纳 B: 桌面端统一Edge WebView2,Web端兼容Edge/Chrome/Firefox", "产品", "P0", "Decided"),
    ("QA-003", "O3:云端模型账号/密钥/预算", "决定默认模型与 fallback 路径", "架构 §3.2 LiteLLM 配置;技术选型 ADR", "默认:已有公司 OpenAI 账号", "A. 已有 OpenAI B. 暂无,走本地 C. 混合", "运维+产品", "P1", "Open"),
    ("QA-004", "O4:中心同步服务运维归属", "决定部署位置与 SLO 责任方", "架构 §6 Kafka 部署;技术选型 R-08", "默认:内网 K3s 部署,运维团队负责", "A. 内网(优) B. 公有云(需合规) C. 混合", "运维+架构", "P1", "Open"),
    ("QA-005", "O5:存量数据实际格式样本", "决定 F11 导入管道设计", "需求 §5.2 F11;模块设计书", "默认:Excel/CSV 为主,部分 TMX", "A. 抽样(优) B. 调研(可) C. 假设 Excel/CSV", "产品+架构", "P1", "Open"),
    ("QA-006", "O6:敏感界定规则", "决定 F10 合规路由默认策略", "需求 §5.2 F10;架构 §1.2 原则 3", "默认:按项目标记 force_local_model", "A. 按域名清单 B. 按项目标记 C. 两者结合", "产品+合规", "P1", "Open"),
    ("QA-007", "F11 存量数据导入格式支持范围", "OFCAT 需求说'清洗+映射',但格式多寡未明确", "需求 §5.2 F11;模块设计书", "默认:Excel/CSV/TMX/TBX", "A. 仅 Excel/CSV(优) B. +TMX/TBX(中) C. 全格式", "产品", "P2", "Open"),
]
write_concerns(ws, "需求层面关心事项", items_req, hdr)

# ---------------- 架构层面 ----------------
ws = wb.create_sheet("02-架构层面")
items_arch = [
    ("QA-011", "TM 索引策略选型", "O1 数据未明确时,PostgreSQL HASH 分桶 16 分区是默认,但若 TM > 100 万,需考虑 scale-out", "架构 §5.3;数据库设计 §4.3", "默认:HASH 16 分桶 + pgvector HNSW", "采纳 A: PG单库16分桶+HNSW索引(300万句段下稳定<35ms)", "架构", "P0", "Decided"),
    ("QA-012", "微服务拆分粒度", "16 个服务是否合理?worker-service 单独是否过度拆分?", "架构 §4.1;运维 UI", "当前:16 个服务", "采纳 A: 维持16服务定义,媒体处理无状态由task-service统一调度", "架构+运维", "P0", "Decided"),
    ("QA-013", "架构 §18 三阶段 vs 需求 §10 M1-M5", "两套阶段划分不严格对齐,需明确哪个是'准'", "架构 §18;测试设计 §14", "当前:测试设计 §14 加注说明", "A. 以架构 §18 为准 B. 以需求 §10 为准 C. 重新整合", "架构+产品", "P1", "Open"),
    ("QA-014", "Event Sourcing 是否引入", "架构说'Outbox+CDC',但是否全业务都事件溯源?", "架构 §7;模块设计书", "当前:仅 Outbox(不 ES)", "A. 维持 Outbox(优) B. 部分核心域 ES C. 全 ES", "架构", "P1", "Open"),
    ("QA-015", "服务间同步 gRPC 失败处理", "gRPC 调用失败时,是熔断、降级、还是死信?", "接口设计书 §3;测试设计", "当前:标准重试 + 熔断", "A. 标准重试+熔断(默认) B. 全部事件化 C. Saga 编排", "架构", "P1", "Open"),
    ("QA-016", "多语言/方言支持", "需求说'中日英韩',但游戏本地化可能涉及藏语/蒙语/方言", "需求 §6.5;接口设计 §3.10", "当前:中日英韩", "A. 中日英韩(维持) B. 增泰越/西/法/德 C. 全部支持", "产品", "P2", "Open"),
    ("QA-017", "离线/在线模式切换", "Tauri 客户端断网时已支持离线队列,但'在线/离线模式'是否要 UI 显式", "模块设计书 §2.3;可热插拔部署", "当前:自动切换,无显式模式", "A. 维持自动 B. 加显式模式开关 C. 显式+粒度", "产品", "P2", "Open"),
    ("QA-018", "审计日志保留期", "审计 365 天,但客户可能要求更长(7 年)", "架构 §4.6;运维设计", "当前:365 天", "A. 365 天(默认) B. 7 年(合规) C. 配置化", "合规", "P1", "Open"),
    ("QA-019", "DR 演练频次", "架构 §16 写'每季度',但实际可能月度", "运维设计;测试设计", "当前:每季度", "A. 每季度(默认) B. 每月 C. 每月自动+季度人工", "运维", "P1", "Open"),
]
write_concerns(ws, "架构层面关心事项", items_arch, hdr)

# ---------------- 技术选型 ----------------
ws = wb.create_sheet("03-技术选型")
items_tech = [
    ("QA-021", "Tokio 异步运行时是唯一选型", "全部 Rust 服务依赖 Tokio,若引入 actix-web 等非 Tokio 生态,会有冲突", "Rust 选型 ADR-R-04;模块设计", "当前:仅 Tokio", "A. 维持 Tokio(默认) B. 允许例外 C. 重新评估", "架构+Rust Lead", "P0", "Open"),
    ("QA-022", "axum vs actix-web 最终选型", "团队对 actix-web 性能可能更信任", "Rust 选型 ADR-R-06", "当前:axum", "A. axum(默认) B. actix-web C. 混合", "架构", "P1", "Open"),
    ("QA-023", "OpenFeature 引入必要性", "增加 SDK 依赖和服务,价值是否够", "运维设计 §6.5", "当前:采用 OpenFeature", "A. 引入(默认) B. 自建简单 flag C. 用 LaunchDarkly SaaS", "架构", "P1", "Open"),
    ("QA-024", "MSRV 1.75 是否够用", "一些 2025+ 新特性可能不可用", "Rust 选型 ADR-R-01", "当前:1.75+", "A. 1.75(默认) B. 1.80 C. 1.83+", "Rust Lead", "P1", "Open"),
    ("QA-025", "librdkafka 系统依赖", "rdkafka 需 C 库,Docker 镜像需 apt-get install", "Rust 选型 ADR-R-11;CI", "当前:Dockerfile 安装", "A. apt-get(默认) B. 静态编译 C. 切到 rskafka", "Rust Lead+运维", "P1", "Open"),
    ("QA-026", "PostgreSQL 16 vs 17", "PG 17 已发布(假设 2026),新特性是否用", "技术选型书 §2", "当前:16.x", "A. 16(默认) B. 17(若生态稳定) C. LTS 跟随", "DBA", "P2", "Open"),
    ("QA-027", "CloudNativePG vs Patroni", "架构说 CNPG,但 Patroni 生态成熟", "架构 §5.6;技术选型书", "当前:CNPG", "A. CNPG(默认) B. Patroni C. 商业 EDB", "DBA+架构", "P1", "Open"),
    ("QA-028", "Envoy Gateway vs Istio", "服务网格未启用,Gateway 用 Envoy", "技术选型书", "当前:Envoy Gateway only", "A. 仅 Envoy Gateway(默认) B. 引入 Linkerd C. 完整 Istio", "架构", "P2", "Open"),
]
write_concerns(ws, "技术选型关心事项", items_tech, hdr)

# ---------------- 跨文档一致性 ----------------
ws = wb.create_sheet("04-跨文档一致性")
items_doc = [
    ("QA-031", "测试设计书文件名 v1.0 vs 内容 v2.0", "文件名 CATs_测试设计书_v1.0.md 但内容已标准化为 v2.0", "04-测试/测试设计书/", "当前:文件名维持 v1.0 基线", "A. 保留 v1.0 基线文件名 + 正文加注说明(已采纳实施)", "测试负责人", "P2", "Closed"),
    ("QA-032", "04-测试 目录与 doc/README 不一致", "doc/README 原写 04-其他/,但实际创建了 04-测试/", "doc/README.md", "当前:04-测试/ + 05-其他/", "A. 同步 README + 调整其他为 05-其他/(已采纳实施)", "架构+测试", "P3", "Closed"),
    ("QA-033", "OFCAT 文档保留策略", "命名变更说'保留作为历史',但 OFCAT 测试/选型/基础 xlsx 是否同步保留", "01-需求/02-基础设计/", "当前:OFCAT_* 文件保留", "A. 全保留(当前) B. 仅保留 .md C. 归档到 00-archive/", "架构", "P2", "Open"),
    ("QA-034", "阶段编号统一性", "OFCAT M1-M5、架构 §18 三阶段、测试 TM-1~TM-5、可热插拔 §12 阶段一/二/三——四套编号", "全部设计文档", "当前:文档各自编号", "A. 统一为 1 套(优) B. 加交叉对照表 C. 维持各表", "架构", "P1", "Open"),
    ("QA-035", "服务数量描述一致性", "测试设计 v1.0 写 15,改 16;架构 §4.1 列 16;模块设计...各处是否一致?", "全部", "当前:16", "A. 全量 grep 检查(已部分做) B. 引入 SoT C. 维持", "架构", "P1", "Open"),
    ("QA-036", "数据库设计书与微服务数量一致性", "数据库设计书 §1 列 8 库,与 16 服务是否一致(媒体处理无独立库)", "03-详细设计/数据库设计书/", "当前:已说明'媒体处理无独立库'", "A. 已对齐 B. 增加交叉引用 C. 重新审视", "架构", "P3", "Decided"),
]
write_concerns(ws, "跨文档一致性关心事项", items_doc, hdr)

# ---------------- 假设与默认值 ----------------
ws = wb.create_sheet("05-假设与默认值")
items_assume = [
    ("QA-041", "PG + pgvector 性能", "10000 句对 + 100 万级 HNSW 索引性能未实测", "架构 §5.3;测试设计 SLO", "当前:假设达成", "A. 立即跑基准测试 B. 先上 1 万观察 C. 备用独立向量库", "架构+DBA", "P0", "Open"),
    ("QA-042", "K3s 3 控制面 HA 是否够", "50-3000 并发 + 16 服务规模,3 节点是否够", "架构 §1.1;运维", "当前:3 节点内嵌 etcd HA", "采纳 A: 3控制面节点高可用满足容忍1节点宕机,配合N个动态Worker节点", "架构+运维", "P0", "Decided"),
    ("QA-043", "延迟 SLO 数值", "L0<100ms / L2<1s 等是我推导的,未实测", "测试设计 §4.5", "当前:已标注'初版,待 PRE 校准'", "A. 标初版,M3 后校准 B. 立即压测 C. 沿用业内默认", "测试+架构", "P1", "Open"),
    ("QA-044", "缺陷注入率 5‰/3‰/2‰", "是 JIS X 0129 推荐值,但实际项目可能不同", "测试设计 §10.5", "当前:采用推荐值", "A. 维持推荐 B. 调整(更严/更宽) C. 跳过变异测试", "测试", "P1", "Open"),
    ("QA-045", "feature flag 影响范围 > 10% 需审批", "'10%' 是我假设,实际可能 5% 或 25%", "运维设计 §7.5", "当前:10%", "A. 10%(默认) B. 5%(更严) C. 25%(更宽)", "运维", "P1", "Open"),
    ("QA-046", "Canary 阶段比例 5-25-50-100", "业内常见但可调", "运维设计 §8.2", "当前:5-25-50-100", "A. 维持 B. 1-10-50-100(更保守) C. 1-5-25-50-100(更细)", "架构", "P2", "Open"),
    ("QA-047", "admin UI 5 模块 M1 上线", "假设 M1 可上线服务总览+部署+审计 3 个,实际可能做不到", "运维设计 §7.2;路线图", "当前:M1 = 服务总览+部署+审计", "A. M1 = 3 模块(默认) B. M1 = 仅服务总览 C. M1 = 5 模块全上", "架构+运维", "P1", "Open"),
    ("QA-048", "审计日志不可篡改实现", "PG trigger 限制 UPDATE/DELETE,但备份恢复时如何处理", "运维设计 §7.3.7", "当前:trigger 阻断 + 异地归档", "A. trigger(默认) B. 追加到只读 Kafka topic C. 上链", "合规+DBA", "P1", "Open"),
]
write_concerns(ws, "假设与默认值关心事项", items_assume, hdr)

# ---------------- 数据与规模 ----------------
ws = wb.create_sheet("06-数据与规模")
items_data = [
    ("QA-051", "本地 LLM 资源(模型/GPU/位置)", "合规项目走本地 LLM,但具体规格未定", "架构 §3.2;技术选型书", "当前:vLLM/Ollama + RTX 4090/A10", "采纳 A: 明确硬件配置单卡24GB显存,加载Qwen2.5-7B-Instruct量化版", "架构+运维", "P0", "Decided"),
    ("QA-052", "合规敏感项目样本数据来源", "测试需要 10 段 PII 源文本,不能真用客户数据", "测试设计 §7.3", "当前:内部审核构造", "采纳 A: 采用自动化规则生成20套脱敏合成数据集,严禁使用生产客户数据", "测试+合规", "P0", "Decided"),
    ("QA-053", "媒体样本来源与版权", "视频/PDF/Office/GIF 样本,版权风险", "测试设计 §7.3", "当前:内部素材库", "A. 内部素材(已审) B. 公开样本 C. 购买版权", "测试+法务", "P1", "Open"),
    ("QA-054", "历史数据规模与归档窗口", "tasks > 90 天归档,实际 50-3000 用户产生多少", "运维设计 §10.6", "当前:90 天/180 天", "A. 维持 90/180 B. 调整为 30/90 C. 配置化", "DBA+产品", "P2", "Open"),
    ("QA-055", "生产数据初始迁移策略", "上线时是否需迁移存量 TM/术语", "F11", "当前:清洗管道就绪", "A. 上线时清空 B. 试点迁移 C. 全新启动", "产品+DBA", "P1", "Open"),
]
write_concerns(ws, "数据与规模关心事项", items_data, hdr)

# ---------------- 环境与部署 ----------------
ws = wb.create_sheet("07-环境与部署")
items_env = [
    ("QA-061", "生产 TLS / 内部 CA / 域名", "*.cats.internal 是否已规划,CA(cfssl)是否已签", "架构 §1.1;运维", "当前:内部 CA + mkcert/cfssl", "A. 内部 CA(默认) B. 公有 CA C. 混合", "运维", "P0", "Open"),
    ("QA-062", "CI Runner 资源", "完整 E2E 套件 11 场景,资源是否够", "测试设计 §10.3;CI", "当前:k3d + 容器化", "A. 自建 GitHub Actions B. 商业 CI C. 自建 Jenkins", "运维", "P1", "Open"),
    ("QA-063", "生产部署灰度比例", "Canary 5%-25%-50%-100% 在局域网是否够细", "运维设计 §8.2", "当前:5-25-50-100", "A. 维持 B. 按 org 灰度 C. 1-10-50-100", "架构", "P2", "Open"),
    ("QA-064", "Docker 基础镜像选择", "distroless vs debian-slim vs scratch", "Rust 选型 §14.2", "当前:distroless", "A. distroless(默认) B. debian-slim C. scratch", "Rust Lead", "P1", "Open"),
]
write_concerns(ws, "环境与部署关心事项", items_env, hdr)

# ---------------- 合规与安全 ----------------
ws = wb.create_sheet("08-合规与安全")
items_compliance = [
    ("QA-071", "等保/行业认证", "等保 2.0 / GDPR / PIPL / 游戏本地化行业规范", "运维 §7;合规测试", "当前:等保 2.0 三级标准设计", "采纳 A: 系统安全基线严格对齐等保 2.0 三级技术标准实施", "合规", "P0", "Decided"),
    ("QA-072", "敏感内容审计可追溯", "网抓+审计日志双验证,但漏判怎么办", "合规测试 §4.9", "当前:阻断+告警", "A. 阻断+告警(默认) B. 阻断+人工审核 C. 阻断+上报合规", "合规", "P1", "Open"),
    ("QA-073", "JWT 密钥轮换流程", "auth-service 签发,密钥如何轮换", "架构 §14;接口设计书 §1.2", "当前:cert-manager + 短期证书", "A. 短 JWT + 刷新(默认) B. 长 JWT + 强制吊销 C. 混合", "架构", "P1", "Open"),
    ("QA-074", "mTLS 全链路落地", "服务间 mTLS 是否必须", "架构 §14", "当前:应用层 mTLS + NetworkPolicy", "A. 应用层(默认) B. 引入服务网格 C. 双层", "架构", "P1", "Open"),
]
write_concerns(ws, "合规与安全关心事项", items_compliance, hdr)

# ---------------- 团队与资源 ----------------
ws = wb.create_sheet("09-团队与资源")
items_team = [
    ("QA-081", "Rust 团队规模与熟练度", "16 个 Rust 服务 + 1 个 Tauri 客户端需多少 Rust 工程师", "实施计划", "当前:未明确", "A. 给出团队配置(优) B. 估算(3-5 人) C. 招聘计划", "管理层", "P1", "Open"),
    ("QA-082", "Tauri / Svelte 跨端能力", "Svelte 5 是新版本,团队学习曲线", "模块设计书 §2", "当前:已熟练", "A. 培训计划 B. 改用 React C. 外包", "管理层", "P1", "Open"),
    ("QA-083", "媒体处理算法工程师", "ASR/OCR/LibreOffice 调优需要专人", "运维/质量", "当前:未明确", "A. 指定专人 B. 自动化 C. 顾问", "管理层", "P2", "Open"),
]
write_concerns(ws, "团队与资源关心事项", items_team, hdr)

# ---------------- 流程与协作 ----------------
ws = wb.create_sheet("10-流程与协作")
items_proc = [
    ("QA-091", "文档评审与签字流程", "谁能签架构/接口/测试设计", "全部设计文档", "当前:起草/评审/批准三栏", "A. 维持(优) B. 加测试签字 C. 加合规签字", "管理层", "P1", "Open"),
    ("QA-092", "CI 工具选型", "Gitea Actions vs Jenkins vs GitHub Actions", "CI/CD", "当前:复用现有", "A. 复用现有(默认) B. GitHub Actions C. 新建", "运维", "P2", "Open"),
    ("QA-093", "缺陷管理工具", "与项目管理工具对接,但未指定", "测试设计 §15", "当前:未指定", "A. Jira B. GitHub Issues C. 内部", "管理层", "P2", "Open"),
    ("QA-101", "跨团队契约变更协调", "服务间接口变更,如何通知所有消费者", "接口设计书 §1.5", "当前:CI 强制门禁", "A. CI 门禁(默认) B. 接口变更公告+PR review C. 治理委员会", "架构", "P0", "Open"),
]
write_concerns(ws, "流程与协作关心事项", items_proc, hdr)

# ---------------- 决策矩阵 ----------------
ws = wb.create_sheet("决策矩阵")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 20, "B": 14, "C": 50})
title_block(ws, "决策矩阵（按决策方归类）", 3)
headers = ["决策方", "数量", "相关关心事项"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 3)
matrix = [
    ("产品", 8, "QA-001, QA-002, QA-003, QA-005, QA-006, QA-007, QA-016, QA-017"),
    ("架构师", 23, "QA-011, QA-012, QA-013, QA-014, QA-015, QA-018, QA-019, QA-021, QA-022, QA-023, QA-026, QA-027, QA-028, QA-031, QA-032, QA-034, QA-035, QA-043, QA-046, QA-047, QA-073, QA-074, QA-101"),
    ("运维/SRE", 12, "QA-004, QA-012, QA-019, QA-025, QA-042, QA-051, QA-061, QA-062, QA-063, QA-064, QA-092, QA-045"),
    ("DBA", 6, "QA-026, QA-027, QA-041, QA-048, QA-054, QA-055"),
    ("测试负责人", 8, "QA-031, QA-032, QA-043, QA-044, QA-045, QA-052, QA-053, QA-083"),
    ("合规", 5, "QA-018, QA-048, QA-052, QA-071, QA-072"),
    ("Rust Lead", 4, "QA-021, QA-024, QA-025, QA-064"),
    ("管理层", 5, "QA-081, QA-082, QA-083, QA-091, QA-093"),
]
r = 4
for row in matrix:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 3)

# ---------------- 后续动作 ----------------
ws = wb.create_sheet("后续动作")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 24, "B": 60})
title_block(ws, "后续动作清单（评审前/中/后）", 2)
headers = ["阶段", "动作"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 2)
actions = [
    ("评审前 1.0", "① 分发本登记册给所有决策方\n② 预填各决策方已知的答案\n③ 预约评审会议程(按决策方归类)"),
    ("评审会议程 1.5h", "0:00-0:10 QA-001/002/003 (产品+架构)\n0:10-0:30 QA-011/012/013 (架构+运维)\n0:30-0:50 QA-021/022/023 (架构+Rust Lead)\n0:50-1:00 QA-051/052 (架构+测试+合规)\n1:00-1:15 QA-061/071 (运维+合规)\n1:15-1:30 QA-101 + 其他开放项 (架构)"),
    ("评审后 1.0", "① 更新本登记册'决策方'列与'状态'列\n② 关闭项归档(移至 05-其他/QA/已关闭/)\n③ 重大决策补充 ADR (02-基础设计/架构设计/ADR/)\n④ 同步更新相关设计文档\n⑤ 启动 M1 Sprint 1(若 P0 全部 Closed)"),
    ("M1 准入门禁", "所有 P0 项 (12 条) Closed\n所有 P1 重要项 (32 条) 至少 Decided 或有明确 Owner\n所有 P2 一般项 (14 条) 有记录"),
    ("评审前自检", "[ ] 所有关心事项已有'建议方案'\n[ ] 阻塞项(P0)优先级排序清晰\n[ ] 决策方明确(无歧义)\n[ ] 影响范围标注(无遗漏)\n[ ] 当前假设标注(避免'我以为')\n[ ] 状态分布合理(不是 100% Open)"),
]
r = 4
for k, v in actions:
    ws.cell(r, 1, k).fill = fill(LIGHT)
    ws.cell(r, 1).font = f(10, True)
    ws.cell(r, 1).alignment = topleft()
    ws.cell(r, 2, v).alignment = topleft()
    ws.cell(r, 2).font = f(9)
    ws.row_dimensions[r].height = 80
    r += 1
grid(ws, 4, r - 1, 1, 2)

# ---------------- Save ----------------
_here = os.path.dirname(os.path.abspath(__file__))
out_dir = os.path.dirname(_here)  # 上级 = 05-其他/ 目录
out = os.path.join(out_dir, "CATs_实施前QA登记册_v1.0.xlsx")
wb.save(out)
print(f"Saved {out} with sheets: {wb.sheetnames}")
