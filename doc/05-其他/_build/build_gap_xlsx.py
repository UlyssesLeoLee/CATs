# -*- coding: utf-8 -*-
"""CATs 应补文档清单 -> Excel 工作簿生成
工作表: 封面 / 总览 / 28 份清单 / 优先级分布 / 行动计划 / 与 150 任务映射
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Microsoft YaHei"
NAVY   = "1F3864"
HEADER = "2E5496"
SUB    = "8EAADB"
LIGHT  = "D9E1F2"
ZEBRA  = "F2F5FB"
WHITE  = "FFFFFF"

DONE   = "C6E0B4"  # 绿 = 已完成 / 本次已建
DOING  = "FFE699"  # 黄 = 进行中 / 后续整合
TODO   = "F2F2F2"  # 灰 = 未启动 / 待阶段触发
BLOCK  = "F8CBAD"  # 红 = 阻塞

P0_FILL = "F8CBAD"
P1_FILL = "FFE699"
P2_FILL = "D9E1F2"
M3_FILL = "E4DFEC"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True):   return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def topleft():         return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def title_block(ws, title, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.fill = fill(NAVY); c.font = f(13, True, WHITE); c.alignment = center(False)
    ws.row_dimensions[1].height = 28

def style_header(ws, row, ncols, fillc=HEADER):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(fillc); cell.font = f(10, True, WHITE)
        cell.alignment = center(); cell.border = BORDER

def grid(ws, r1, r2, c1, c2, zebra=True):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.vertical is None:
                cell.alignment = topleft()
            if zebra and (r - r1) % 2 == 1:
                if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = fill(ZEBRA)

def widths(ws, mp):
    for col, w in mp.items():
        ws.column_dimensions[col].width = w

def pri_color(pri):
    return {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL, "M3": M3_FILL}.get(pri, TODO)

def status_color(status):
    if "✅" in status: return fill(DONE)
    if "🟡" in status: return fill(DOING)
    if "⚪" in status: return fill(TODO)
    return None

# ============================================================
wb = Workbook()

# ============================================================
# Sheet 1: 封面
# ============================================================
ws = wb.active
ws.title = "封面"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 26, "C": 56, "D": 3})
ws.merge_cells("B2:C2")
ws["B2"] = "应补文档清单"
ws["B2"].font = f(22, True, NAVY); ws["B2"].alignment = center(False)
ws.row_dimensions[2].height = 44
ws.merge_cells("B3:C3")
ws["B3"] = "基于 150 任务工作流（CATs 项目）— 缺失文档识别与落地计划"
ws["B3"].font = f(12, True, "404040"); ws["B3"].alignment = center(False)
ws.row_dimensions[3].height = 24

meta = [
    ("文档编号", "CATs-PMO-002"),
    ("文档名", "应补文档清单"),
    ("版本", "v1.0"),
    ("创建日", "2026-08-20"),
    ("作者", "PMO / 架构师"),
    ("状态", "草稿（待评审会确认）"),
    ("密级", "仅社内"),
    ("性质", "150 任务中『该产出但缺失』的文档清单 + 落地计划"),
    ("上游文档", "CATs_工作流文档_v1.0.md + CATs_实施前QA登记册_v1.0.md"),
    ("应补文档总数", "28 份（P0=7 / P1=8 / P2=8 / M3=5）"),
    ("本次立即建", "12 份（P0 7 + P1 5）"),
]
r = 5
for k, v in meta:
    ws.cell(r, 2, k).fill = fill(SUB); ws.cell(r, 2).font = f(10, True, WHITE)
    ws.cell(r, 2).alignment = center()
    ws.cell(r, 3, v).alignment = left(); ws.cell(r, 3).font = f(10)
    ws.cell(r, 2).border = BORDER; ws.cell(r, 3).border = BORDER
    ws.row_dimensions[r].height = 22
    r += 1

# ============================================================
# Sheet 2: 总览
# ============================================================
ws = wb.create_sheet("总览")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 8, "B": 30, "C": 14, "D": 12, "E": 14, "F": 24})
title_block(ws, "优先级总览", 6)
headers = ["#", "优先级", "数量", "占比", "期望关闭时点", "本次动作"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 6)
pri_rows = [
    (1, "P0（评审会前必须）",     7, "25%", "2026-08-25 评审会前", "本次建 7 份"),
    (2, "P1（M1 启动前应有）",    8, "29%", "2026-09-10 M1 Sprint 0", "本次建 5 份，3 份后续整合"),
    (3, "P2（M1 Sprint 1~3 内补）", 8, "29%", "2026-10 末",            "清单标注，待触发"),
    (4, "M3（上线/收尾时点）",     5, "17%", "2027-Q2 上线期",         "清单标注，待触发"),
]
r = 4
for row in pri_rows:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    pri = row[1]
    if "P0" in pri: ws.cell(r, 2).fill = fill(P0_FILL)
    elif "P1" in pri: ws.cell(r, 2).fill = fill(P1_FILL)
    elif "P2" in pri: ws.cell(r, 2).fill = fill(P2_FILL)
    elif "M3" in pri: ws.cell(r, 2).fill = fill(M3_FILL)
    r += 1
# 合计
ws.cell(r, 1, "合计").font = f(10, True); ws.cell(r, 1).fill = fill(LIGHT); ws.cell(r, 1).alignment = center()
for c, v in enumerate(["", "4 优先级", 28, "100%", "—", "本次 12 份"], 2):
    cell = ws.cell(r, c, v); cell.font = f(10, True); cell.fill = fill(LIGHT); cell.alignment = topleft()
grid(ws, 4, r, 1, 6)
ws.freeze_panes = "A4"

# ============================================================
# Sheet 3: 28 份清单
# ============================================================
ws = wb.create_sheet("28份清单")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 6, "B": 20, "C": 10, "D": 36, "E": 36, "F": 8, "G": 8, "H": 8, "I": 18, "J": 24})
title_block(ws, "28 份应补文档完整清单", 10)
headers = ["#", "フェーズ", "任务ID", "文档名", "拟保存路径", "优先级", "状态", "本次", "关闭时点", "备注"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 10)

gaps = [
    # (#, フェーズ, ID, 文档名, 路径, P, 状态, 本次, 关闭时点, 备注)
    (1,  "超上流",       3,  "CATs_系统化计划书_v1.0.md",        "05-其他\\立项材料\\",                          "P0", "⚪ 缺失",                "✅ 本次", "2026-08-25", "投资/期间/资源"),
    (2,  "超上流",       5,  "CATs_项目章程_v1.0.md",            "05-其他\\立项材料\\",                          "P0", "⚪ 缺失",                "✅ 本次", "2026-08-25", "PJ Charter"),
    (3,  "要件定義",     21, "CATs_要件承認決議書_v1.0.md",      "05-其他\\评审记录\\",                          "P0", "⚪ 缺失",                "✅ 本次", "2026-08-25", "要件 Baseline"),
    (4,  "基本設計",     41, "CATs_BD评审纪要_v1.0.md",          "05-其他\\评审记录\\",                          "P0", "⚪ 缺失",                "✅ 本次", "2026-08-25", "基本设计评审"),
    (5,  "管理",         131,"CATs_项目管理计划书_v1.0.md",      "05-其他\\管理\\",                              "P0", "⚪ 缺失",                "✅ 本次", "2026-08-25", "PJ Plan"),
    (6,  "管理",         132,"CATs_WBS_v1.0.xlsx",               "05-其他\\管理\\",                              "P0", "⚪ 缺失",                "✅ 本次", "2026-08-25", "工作分解结构"),
    (7,  "管理",         144,"CATs_Baseline一览_v1.0.md",        "05-其他\\管理\\",                              "P0", "⚪ 缺失",                "✅ 本次", "2026-08-25", "基线登记册"),
    (8,  "要件定義",     17, "CATs_安全要件定义书_v1.0.md",      "05-其他\\安全\\",                              "P1", "⚪ 缺失(仅顺带)",         "✅ 本次", "2026-09-10", "等保/合规"),
    (9,  "要件定義",     19, "CATs_迁移要件定义书_v1.0.md",      "05-其他\\迁移\\",                              "P1", "⚪ 缺失",                "✅ 本次", "2026-09-10", "数据/系统切换"),
    (10, "基本設計",     27, "CATs_报表设计书_v1.0.md",          "02-基础设计\\报表\\",                          "P1", "⚪ 缺失",                "✅ 本次", "2026-09-10", "导出/打印格式"),
    (11, "詳細設計",     48, "CATs_SQL设计一览_v1.0.md",         "03-详细设计\\SQL\\",                           "P1", "⚪ 缺失",                "✅ 本次", "2026-09-10", "DDL+关键 SQL"),
    (12, "受入試験",     90, "CATs_UAT计划书_v1.0.md",           "04-测试\\UAT\\",                               "P1", "⚪ 缺失",                "✅ 本次", "2026-09-10", "UAT 范围/环境/通过准则"),
    (13, "基本設計",     33, "CATs_权限矩阵_v1.0.md",            "05-其他\\安全\\",                              "P1", "🟡 原则有(可热插拔§7.4)","⏳ 后续整合","2026-09-10", "RBAC 角色×资源"),
    (14, "基本設計",     40, "CATs_迁移设计书_v1.0.md",          "05-其他\\迁移\\",                              "P1", "🟡 F11 设计有",          "⏳ 后续整合","2026-09-10", "生产迁移"),
    (15, "詳細設計",     44, "CATs_类图_v1.0.md",                "03-详细设计\\类图\\",                          "P1", "🟡 原则有",              "⏳ 后续整合","2026-09-10", "UML 类图"),
    (16, "超上流",       6,  "CATs_AsIs业务流程图_v1.0.md",      "01-需求\\原始需求\\",                          "P2", "🟡 OFCAT §2 部分覆盖",   "⚪ 待触发", "2026-10 末", "业务流程现状"),
    (17, "超上流",       7,  "CATs_AsIs系统构成图_v1.0.md",      "01-需求\\原始需求\\",                          "P2", "🟡 OFCAT §3 部分覆盖",   "⚪ 待触发", "2026-10 末", "系统现状"),
    (18, "基本設計",     26, "CATs_UI_UX设计书_v1.0.md",         "02-基础设计\\UI\\",                            "P2", "🟡 Next.js 有",          "⚪ 待触发", "2026-10 末", "UI 完整设计"),
    (19, "詳細設計",     49, "CATs_批处理详细设计_v1.0.md",      "03-详细设计\\批处理\\",                        "P2", "🟡 接口有",              "⚪ 待触发", "2026-10 末", "worker 详细"),
    (20, "要件定義",     20, "CATs_要件评审纪要_v1.0.md",        "05-其他\\评审记录\\",                          "P2", "⚪ 缺失",                "⚪ 待触发", "M1-S1",    "RD Review 纪要"),
    (21, "詳細設計",     52, "CATs_DD评审纪要_v1.0.md",          "05-其他\\评审记录\\",                          "P2", "⚪ 缺失",                "⚪ 待触发", "M1-S2",    "DD Review 纪要"),
    (22, "管理",         133,"CATs_进度报告_初版_v1.0.md",       "05-其他\\管理\\",                              "P2", "🟡 工作流文档作快照",     "⚪ 待触发", "M1-S1",    "周/月进度"),
    (23, "管理",         137,"CATs_构成管理台账_v1.0.md",        "05-其他\\管理\\",                              "P2", "🟡 命名/版本有",          "⚪ 待触发", "M1-S1",    "制品/文档版本"),
    (24, "システム試験",  "—", "CATs_系统测试报告_v1.0.md",      "04-测试\\ST\\",                                "M3", "⚪ 缺失",                "⚪ 待触发", "ST 阶段",  "PT/Load/Sec 等"),
    (25, "受入試験",     "—", "CATs_UAT报告_v1.0.md",            "04-测试\\UAT\\",                               "M3", "⚪ 缺失",                "⚪ 待触发", "UAT 阶段", "92-95 输出"),
    (26, "運用",         109,"CATs_运维手册_v1.0.md",            "05-其他\\运维\\",                              "M3", "⚪ 缺失",                "⚪ 待触发", "上线前",    "日常运维"),
    (27, "管理",         140,"CATs_会议报告模板_v1.0.md",        "05-其他\\管理\\",                              "M3", "⚪ 缺失",                "⚪ 待触发", "M1 起",    "周会/里程碑"),
    (28, "終結",         147,"CATs_项目结项报告模板_v1.0.md",    "05-其他\\管理\\",                              "M3", "⚪ 缺失",                "⚪ 待触发", "收尾",      "结项模板"),
]
r = 4
for row in gaps:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    # 优先级染色
    pc = pri_color(row[5])
    if pc:
        ws.cell(r, 6).fill = fill(pc)
        ws.cell(r, 6).alignment = center()
        ws.cell(r, 6).font = f(10, True)
    # 状态列染色
    sc = status_color(row[6])
    if sc:
        ws.cell(r, 7).fill = sc
    # 本次列
    if "✅ 本次" in str(row[7]):
        ws.cell(r, 8).fill = fill(DONE)
        ws.cell(r, 8).font = f(10, True, "375623")
    elif "⏳ 后续" in str(row[7]):
        ws.cell(r, 8).fill = fill(DOING)
    else:
        ws.cell(r, 8).fill = fill(TODO)
    r += 1
grid(ws, 4, r - 1, 1, 10)
ws.freeze_panes = "A4"

# ============================================================
# Sheet 4: 优先级分布
# ============================================================
ws = wb.create_sheet("优先级分布")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 14, "B": 14, "C": 50, "D": 14})
title_block(ws, "各优先级详解", 4)
headers = ["优先级", "数量", "主要类别", "期望关闭"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
pri_detail = [
    ("P0", 7, "立项材料(2) + 评审决议(2) + 管理基础(3)",  "2026-08-25 评审会前",  P0_FILL),
    ("P1", 8, "安全/迁移/报表/SQL/UAT 核心要件",          "2026-09-10 M1 Sprint 0", P1_FILL),
    ("P2", 8, "AsIs/UI/批处理/进度/构成 + 评审纪要",      "2026-10 末",            P2_FILL),
    ("M3", 5, "ST/UAT 报告 + 运维手册 + 模板类",          "2027-Q2 上线期",        M3_FILL),
]
r = 4
for p, n, desc, when, c in pri_detail:
    ws.cell(r, 1, p).fill = fill(c); ws.cell(r, 1).font = f(11, True); ws.cell(r, 1).alignment = center()
    ws.cell(r, 2, n).alignment = center()
    ws.cell(r, 3, desc).alignment = topleft()
    ws.cell(r, 4, when).alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)

# ============================================================
# Sheet 5: 行动计划
# ============================================================
ws = wb.create_sheet("行动计划")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 8, "B": 30, "C": 14, "D": 18, "E": 50})
title_block(ws, "本次立即执行计划（12 份）", 5)
headers = ["#", "文档名", "优先级", "保存路径", "基于的现有文档"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
plan = [
    (1,  "CATs_系统化计划书_v1.0.md",     "P0", "05-其他\\立项材料\\",          "初版构想 + 工作流文档"),
    (2,  "CATs_项目章程_v1.0.md",         "P0", "05-其他\\立项材料\\",          "初版构想"),
    (3,  "CATs_要件承認決議書_v1.0.md",   "P0", "05-其他\\评审记录\\",          "OFCAT v1.1 + 评审会"),
    (4,  "CATs_BD评审纪要_v1.0.md",       "P0", "05-其他\\评审记录\\",          "架构/技术选型/接口/数据库/模块 + 可热插拔"),
    (5,  "CATs_项目管理计划书_v1.0.md",   "P0", "05-其他\\管理\\",              "工作流文档 + 实施前 QA 登记册"),
    (6,  "CATs_WBS_v1.0.xlsx",            "P0", "05-其他\\管理\\",              "工作流文档 150 任务拆分"),
    (7,  "CATs_Baseline一览_v1.0.md",     "P0", "05-其他\\管理\\",              "命名变更 + 全部 CATs 文档"),
    (8,  "CATs_安全要件定义书_v1.0.md",   "P1", "05-其他\\安全\\",              "架构§14 + 可热插拔§12"),
    (9,  "CATs_迁移要件定义书_v1.0.md",   "P1", "05-其他\\迁移\\",              "F11 设计 + 可热插拔§11"),
    (10, "CATs_报表设计书_v1.0.md",       "P1", "02-基础设计\\报表\\",          "需求定义 F1-F11 + 接口设计"),
    (11, "CATs_SQL设计一览_v1.0.md",      "P1", "03-详细设计\\SQL\\",           "数据库设计书 §4 + DDL"),
    (12, "CATs_UAT计划书_v1.0.md",        "P1", "04-测试\\UAT\\",               "测试设计 §3.4 + 客户期望"),
]
r = 4
for row in plan:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    pc = pri_color(row[2])
    if pc: ws.cell(r, 3).fill = fill(pc); ws.cell(r, 3).alignment = center(); ws.cell(r, 3).font = f(10, True)
    r += 1
grid(ws, 4, r - 1, 1, 5)
ws.freeze_panes = "A4"

# ============================================================
# Sheet 6: 与 150 任务映射
# ============================================================
ws = wb.create_sheet("与150任务映射")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 8, "B": 16, "C": 12, "D": 32})
title_block(ws, "应补文档 → 150 任务 ID 反向索引", 4)
headers = ["#", "任务 ID", "优先级", "文档名"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 4)
# 任务 ID 排序
m150 = [
    (3,  "P0", "CATs_系统化计划书"),
    (5,  "P0", "CATs_项目章程"),
    (6,  "P2", "CATs_AsIs业务流程图"),
    (7,  "P2", "CATs_AsIs系统构成图"),
    (17, "P1", "CATs_安全要件定义书"),
    (19, "P1", "CATs_迁移要件定义书"),
    (20, "P2", "CATs_要件评审纪要"),
    (21, "P0", "CATs_要件承認決議書"),
    (26, "P2", "CATs_UI_UX设计书"),
    (27, "P1", "CATs_报表设计书"),
    (33, "P1", "CATs_权限矩阵"),
    (40, "P1", "CATs_迁移设计书"),
    (41, "P0", "CATs_BD评审纪要"),
    (44, "P1", "CATs_类图"),
    (48, "P1", "CATs_SQL设计一览"),
    (49, "P2", "CATs_批处理详细设计"),
    (52, "P2", "CATs_DD评审纪要"),
    (90, "P1", "CATs_UAT计划书"),
    (109,"M3", "CATs_运维手册"),
    (131,"P0", "CATs_项目管理计划书"),
    (132,"P0", "CATs_WBS"),
    (133,"P2", "CATs_进度报告_初版"),
    (137,"P2", "CATs_构成管理台账"),
    (140,"M3", "CATs_会议报告模板"),
    (144,"P0", "CATs_Baseline一览"),
    (147,"M3", "CATs_项目结项报告模板"),
    ("—","M3", "CATs_系统测试报告(80-87)"),
    ("—","M3", "CATs_UAT报告(92-95)"),
]
r = 4
for i, (tid, p, name) in enumerate(m150, 1):
    ws.cell(r, 1, i).alignment = center()
    ws.cell(r, 2, str(tid)).alignment = center()
    ws.cell(r, 2).font = f(10, True)
    ws.cell(r, 3, p).alignment = center()
    ws.cell(r, 3).fill = fill(pri_color(p))
    ws.cell(r, 3).font = f(10, True)
    ws.cell(r, 4, name).alignment = topleft()
    r += 1
grid(ws, 4, r - 1, 1, 4)
ws.freeze_panes = "A4"

# ============================================================
_here = os.path.dirname(os.path.abspath(__file__))
out_dir = os.path.dirname(_here)
out = os.path.join(out_dir, "CATs_应补文档清单_v1.0.xlsx")
wb.save(out)
print(f"Saved {out} with sheets: {wb.sheetnames}")
