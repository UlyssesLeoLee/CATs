# -*- coding: utf-8 -*-
"""CATs 应补文档清单 v1.1 -> Excel
54 份（v1.0 28 + v1.1 新增 26）
工作表: 封面 / 总览 / 54份清单 / 优先级分布 / 行动计划 / 与150任务映射 / v1.1新增
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Microsoft YaHei"
NAVY   = "1F3864"
HEADER = "2E5496"
SUB    = "8EAADB"
LIGHT  = "D9E1F2"
ZEBRA  = "F2F5FB"
WHITE  = "FFFFFF"

DONE   = "C6E0B4"
DOING  = "FFE699"
TODO   = "F2F2F2"
BLOCK  = "F8CBAD"

P0_FILL = "F8CBAD"
P1_FILL = "FFE699"
P2_FILL = "D9E1F2"
M3_FILL = "E4DFEC"
V11_FILL = "B4C7E7"  # v1.1 新增用

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True):   return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def topleft():         return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def title_block(ws, title, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.fill = fill(NAVY); c.font = f(13, True, WHITE); c.alignment = center(False)
    ws.row_dimensions[1].height = 28

def style_header(ws, row, ncols, fillc=HEADER):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(fillc); cell.font = f(10, True, WHITE)
        cell.alignment = center(); cell.border = BORDER

def grid(ws, r1, r2, c1, c2, zebra=True):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.vertical is None:
                cell.alignment = topleft()
            if zebra and (r - r1) % 2 == 1:
                if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = fill(ZEBRA)

def widths(ws, mp):
    for col, w in mp.items():
        ws.column_dimensions[col].width = w

def pri_color(pri):
    return {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL, "M3": M3_FILL}.get(pri, TODO)

# ============================================================
wb = Workbook()

# Sheet 1: 封面
# ============================================================
ws = wb.active
ws.title = "封面"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 26, "C": 56, "D": 3})
ws.merge_cells("B2:C2")
ws["B2"] = "应补文档清单 v1.1"
ws["B2"].font = f(22, True, NAVY); ws["B2"].alignment = center(False)
ws.row_dimensions[2].height = 44
ws.merge_cells("B3:C3")
ws["B3"] = "基于 150 任务工作流（CATs 项目）— 54 份应补文档（v1.0 28 + v1.1 新增 26）"
ws["B3"].font = f(12, True, "404040"); ws["B3"].alignment = center(False)
ws.row_dimensions[3].height = 24

meta = [
    ("文档编号", "CATs-PMO-002"),
    ("版本", "v1.1"),
    ("创建日", "2026-08-20"),
    ("v1.0", "28 份"),
    ("v1.1 新增", "26 份（实施/治理 3 + 决议模板 5 + 收尾模板 3 + 管理工具 6 + 运维 4 + ADR 1 + ST 报告 3 + 保守 1）"),
    ("合计", "54 份"),
    ("P0", "10 份（评审会前必建）"),
    ("P1", "9 份（M1 启动前）"),
    ("P2", "11 份（M1 Sprint 1~3）"),
    ("M3", "24 份（上线/收尾）"),
    ("本次立即建", "16 份（P0 10 + P1 6）"),
    ("后续整合/待触发", "38 份"),
]
r = 5
for k, v in meta:
    ws.cell(r, 2, k).fill = fill(SUB); ws.cell(r, 2).font = f(10, True, WHITE); ws.cell(r, 2).alignment = center()
    ws.cell(r, 3, v).alignment = left(); ws.cell(r, 3).font = f(10)
    ws.cell(r, 2).border = BORDER; ws.cell(r, 3).border = BORDER
    ws.row_dimensions[r].height = 22
    r += 1

# Sheet 2: 总览
ws = wb.create_sheet("总览")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 8, "B": 30, "C": 12, "D": 10, "E": 12, "F": 22, "G": 20})
title_block(ws, "优先级总览（v1.1）", 7)
headers = ["#", "优先级", "v1.0", "v1.1 新增", "合计", "期望关闭时点", "本次动作"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 7)
pri_rows = [
    (1, "P0 评审会前必须",       7, 3,  10, "2026-08-25 评审会前",     "本次建 10 份"),
    (2, "P1 M1 启动前应有",      8, 1,  9,  "2026-09-10 M1 Sprint 0", "本次建 6 份，3 后续整合"),
    (3, "P2 M1 Sprint 1~3",      8, 3,  11, "2026-10 末",            "清单标注，待触发"),
    (4, "M3 上线/收尾",          5, 19, 24, "2027-Q2 上线期",         "清单标注，待触发"),
]
r = 4
for row in pri_rows:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    pri = row[1]
    if "P0" in pri: ws.cell(r, 2).fill = fill(P0_FILL)
    elif "P1" in pri: ws.cell(r, 2).fill = fill(P1_FILL)
    elif "P2" in pri: ws.cell(r, 2).fill = fill(P2_FILL)
    elif "M3" in pri: ws.cell(r, 2).fill = fill(M3_FILL)
    r += 1
ws.cell(r, 1, "合计").font = f(10, True); ws.cell(r, 1).fill = fill(LIGHT); ws.cell(r, 1).alignment = center()
for c, v in enumerate(["", "4 优先级", 28, 26, 54, "—", "本次 16 份"], 2):
    cell = ws.cell(r, c, v); cell.font = f(10, True); cell.fill = fill(LIGHT); cell.alignment = topleft()
grid(ws, 4, r, 1, 7)
ws.freeze_panes = "A4"

# Sheet 3: 54 份清单
ws = wb.create_sheet("54份清单")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 6, "B": 20, "C": 10, "D": 32, "E": 36, "F": 8, "G": 8, "H": 8, "I": 18, "J": 12})
title_block(ws, "54 份应补文档完整清单", 10)
headers = ["#", "フェーズ", "任务ID", "文档名", "拟保存路径", "优先级", "v1.1", "状态", "关闭时点", "本次"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 10)

gaps = [
    # (序号, フェーズ, ID, 文档, 路径, 优先级, v1.1标记, 状态, 关闭时点, 本次)
    (1,  "超上流",       3,  "CATs_系统化计划书",          "05-其他\\立项材料\\",         "P0", "v1.0", "✅ 已建",  "2026-08-25", "本次"),
    (2,  "超上流",       5,  "CATs_项目章程",              "05-其他\\立项材料\\",         "P0", "v1.0", "✅ 已建",  "2026-08-25", "本次"),
    (3,  "要件定義",     21, "CATs_要件承認決議書",        "05-其他\\评审记录\\",         "P0", "v1.0", "✅ 已建",  "2026-08-25", "本次"),
    (4,  "基本設計",     41, "CATs_BD 评审纪要",           "05-其他\\评审记录\\",         "P0", "v1.0", "✅ 已建",  "2026-08-25", "本次"),
    (5,  "実装",         "53/57/58", "CATs_CI_CD 構築運用手順書", "05-其他\\治理\\",         "P0", "v1.1", "🆕 已建",  "2026-08-25", "本次"),
    (6,  "実装",         "54/56", "CATs_開発者ガイド",          "05-其他\\治理\\",            "P0", "v1.1", "🆕 已建",  "2026-08-25", "本次"),
    (7,  "実装",         "55/56/128", "CATs_品質ゲート運用手順書", "05-其他\\治理\\",         "P0", "v1.1", "🆕 已建",  "2026-08-25", "本次"),
    (8,  "管理",         131, "CATs_项目管理计划书",        "05-其他\\管理\\",             "P0", "v1.0", "✅ 已建",  "2026-08-25", "本次"),
    (9,  "管理",         132, "CATs_WBS",                  "05-其他\\管理\\",             "P0", "v1.0", "✅ 已建",  "2026-08-25", "本次"),
    (10, "管理",         144, "CATs_Baseline 一览",         "05-其他\\管理\\",             "P0", "v1.0", "✅ 已建",  "2026-08-25", "本次"),
    (11, "要件定義",     17, "CATs_安全要件定义书",        "05-其他\\安全\\",             "P1", "v1.0", "✅ 已建",  "2026-09-10", "本次"),
    (12, "要件定義",     19, "CATs_迁移要件定义书",        "05-其他\\迁移\\",             "P1", "v1.0", "✅ 已建",  "2026-09-10", "本次"),
    (13, "基本設計",     27, "CATs_报表设计书",            "02-基础设计\\报表\\",         "P1", "v1.0", "✅ 已建",  "2026-09-10", "本次"),
    (14, "基本設計",     33, "CATs_权限矩阵",              "05-其他\\安全\\",             "P1", "v1.0", "🟡 整合中", "2026-09-10", "后续"),
    (15, "基本設計",     40, "CATs_迁移设计书",            "05-其他\\迁移\\",             "P1", "v1.0", "🟡 整合中", "2026-09-10", "后续"),
    (16, "詳細設計",     44, "CATs_类图",                  "03-详细设计\\类图\\",         "P1", "v1.0", "🟡 整合中", "2026-09-10", "后续"),
    (17, "詳細設計",     48, "CATs_SQL 设计一览",          "03-详细设计\\SQL\\",          "P1", "v1.0", "✅ 已建",  "2026-09-10", "本次"),
    (18, "受入試験",     90, "CATs_UAT 计划书",            "04-测试\\UAT\\",              "P1", "v1.0", "✅ 已建",  "2026-09-10", "本次"),
    (19, "管理/架构",    "—", "CATs_ADR 库",               "02-基础设计\\决策\\",         "P1", "v1.1", "⚪ 待触发", "M1-S0",     "本次后续"),
    (20, "超上流",       6,  "CATs_AsIs 业务流程图",       "01-需求\\原始需求\\",         "P2", "v1.0", "⚪ 待触发", "M1-S1",     "后续"),
    (21, "超上流",       7,  "CATs_AsIs 系统构成图",       "01-需求\\原始需求\\",         "P2", "v1.0", "⚪ 待触发", "M1-S1",     "后续"),
    (22, "要件定義",     20, "CATs_要件评审纪要",          "05-其他\\评审记录\\",         "P2", "v1.0", "⚪ 待触发", "M1-S1",     "后续"),
    (23, "基本設計",     26, "CATs_UI/UX 设计书",          "02-基础设计\\UI\\",           "P2", "v1.0", "⚪ 待触发", "M1-S1",     "后续"),
    (24, "詳細設計",     45, "CATs_ロジック設計書",        "03-详细设计\\逻辑\\",         "P2", "v1.0", "⚪ 待触发", "M1-S2",     "后续"),
    (25, "詳細設計",     49, "CATs_批处理详细设计",        "03-详细设计\\批处理\\",       "P2", "v1.0", "⚪ 待触发", "M1-S2",     "后续"),
    (26, "詳細設計",     52, "CATs_DD 评审纪要",           "05-其他\\评审记录\\",         "P2", "v1.0", "⚪ 待触发", "M1-S2",     "后续"),
    (27, "管理",         "133/135", "CATs_风险登记册",   "05-其他\\管理\\",             "P2", "v1.1", "⚪ 待触发", "M1-S1",     "后续"),
    (28, "管理",         134, "CATs_课题管理表",            "05-其他\\管理\\",             "P2", "v1.0", "⚪ 待触发", "M1-S1",     "后续"),
    (29, "管理",         137, "CATs_构成管理台账",          "05-其他\\管理\\",             "P2", "v1.0", "⚪ 待触发", "M1-S1",     "后续"),
    (30, "管理",         140, "CATs_会议议事录模板",        "05-其他\\管理\\模板\\",       "P2", "v1.1", "⚪ 待触发", "M1-S0",     "后续"),
    (31, "受入試験",     94, "CATs_受入判定書テンプレート", "05-其他\\管理\\模板\\",       "M3", "v1.1", "⚪ 待触发", "UAT 末",    "后续"),
    (32, "受入試験",     95, "CATs_検収書テンプレート",    "05-其他\\管理\\模板\\",       "M3", "v1.1", "⚪ 待触发", "検収",      "后续"),
    (33, "リリース",     103, "CATs_GoNoGo 決議書テンプレート", "05-其他\\管理\\模板\\",   "M3", "v1.1", "⚪ 待触发", "发布前",    "后续"),
    (34, "リリース",     107, "CATs_GoLive 決議書テンプレート", "05-其他\\管理\\模板\\",   "M3", "v1.1", "⚪ 待触发", "上线",      "后续"),
    (35, "終結",         145, "CATs_完了判定書テンプレート", "05-其他\\管理\\模板\\",       "M3", "v1.1", "⚪ 待触发", "收尾",      "后续"),
    (36, "終結",         147, "CATs_完了報告テンプレート",   "05-其他\\管理\\模板\\",       "M3", "v1.1", "⚪ 待触发", "收尾",      "后续"),
    (37, "終結",         148, "CATs_振り返りテンプレート",  "05-其他\\管理\\模板\\",       "M3", "v1.1", "⚪ 待触发", "收尾",      "后续"),
    (38, "終結",         149, "CATs_KT 記録テンプレート",    "05-其他\\管理\\模板\\",       "M3", "v1.1", "⚪ 待触发", "收尾",      "后续"),
    (39, "管理",         133, "CATs_進捗報告書テンプレート", "05-其他\\管理\\模板\\",       "M3", "v1.1", "⚪ 待触发", "M1 起",     "后续"),
    (40, "管理",         138, "CATs_成果物管理台帳",        "05-其他\\管理\\",             "M3", "v1.1", "⚪ 待触发", "M2",        "后续"),
    (41, "管理",         141, "CATs_工数管理表",            "05-其他\\管理\\",             "M3", "v1.1", "⚪ 待触发", "M1",        "后续"),
    (42, "管理",         142, "CATs_コスト管理表",          "05-其他\\管理\\",             "M3", "v1.1", "⚪ 待触发", "M1",        "后续"),
    (43, "管理",         143, "CATs_スコープ変更管理表",    "05-其他\\管理\\",             "M3", "v1.1", "⚪ 待触发", "M1",        "后续"),
    (44, "管理",         144, "CATs_ベースライン変更管理表", "05-其他\\管理\\",             "M3", "v1.1", "⚪ 待触发", "M1",        "后续"),
    (45, "運用",         "109-117", "CATs_運用マニュアル", "05-其他\\运维\\",             "M3", "v1.1", "⚪ 待触发", "上线前",    "后续"),
    (46, "運用",         "114-116", "CATs_インシデント対応プレイブック", "05-其他\\运维\\",   "M3", "v1.1", "⚪ 待触发", "上线前",    "后续"),
    (47, "運用",         113, "CATs_キャパシティ管理計画",  "05-其他\\运维\\",             "M3", "v1.1", "⚪ 待触发", "上线后",    "后续"),
    (48, "運用",         117, "CATs_問い合わせ対応FAQ",     "05-其他\\运维\\",             "M3", "v1.1", "⚪ 待触发", "上线后",    "后续"),
    (49, "システム試験", "76-89", "CATs_系统测试报告",      "04-测试\\ST\\",               "M3", "v1.0", "⚪ 待触发", "ST 末",     "后续"),
    (50, "システム試験", 80, "CATs_性能試験レポート",       "04-测试\\ST\\",               "M3", "v1.1", "⚪ 待触发", "ST 末",     "后续"),
    (51, "システム試験", 83, "CATs_セキュリティ試験レポート", "04-测试\\ST\\",              "M3", "v1.1", "⚪ 待触发", "ST 末",     "后续"),
    (52, "システム試験", 84, "CATs_障害試験レポート",       "04-测试\\ST\\",               "M3", "v1.1", "⚪ 待触发", "ST 末",     "后续"),
    (53, "受入試験",     "92-95", "CATs_UAT 报告",          "04-测试\\UAT\\",              "M3", "v1.0", "⚪ 待触发", "UAT 末",    "后续"),
    (54, "保守",         "118-126", "CATs_保守マニュアル",  "05-其他\\运维\\",             "M3", "v1.1", "⚪ 待触发", "上线后",    "后续"),
]
r = 4
for row in gaps:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    pc = pri_color(row[5])
    if pc:
        ws.cell(r, 6).fill = fill(pc); ws.cell(r, 6).alignment = center(); ws.cell(r, 6).font = f(10, True)
    # v1.1 标记
    if "v1.1" in str(row[6]):
        ws.cell(r, 7).fill = fill(V11_FILL); ws.cell(r, 7).font = f(10, True, "1F3864")
    # 状态
    status = str(row[7])
    if "已建" in status:
        ws.cell(r, 8).fill = fill(DONE); ws.cell(r, 8).font = f(10, True, "375623")
    elif "整合" in status:
        ws.cell(r, 8).fill = fill(DOING)
    else:
        ws.cell(r, 8).fill = fill(TODO)
    # 本次
    if "本次" in str(row[9]):
        ws.cell(r, 10).fill = fill(DONE)
    else:
        ws.cell(r, 10).fill = fill(TODO)
    r += 1
grid(ws, 4, r - 1, 1, 10)
ws.freeze_panes = "A4"

# Sheet 4: 优先级分布
ws = wb.create_sheet("优先级分布")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 14, "B": 12, "C": 12, "D": 50, "E": 16, "F": 14})
title_block(ws, "各优先级详解", 6)
headers = ["优先级", "v1.0", "v1.1", "主要类别", "期望关闭", "本次"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 6)
pri_detail = [
    ("P0", 7, 3, "立项 + 评审 + 实施治理 + 管理基础",  "2026-08-25 评审会前",  P0_FILL, "10 本次"),
    ("P1", 8, 1, "核心要件 + ADR 库",                  "2026-09-10 M1 Sprint 0", P1_FILL, "6 本次+3 整合"),
    ("P2", 8, 3, "AsIs/UI/批处理/管理 + 风险/议事录",  "2026-10 末",            P2_FILL, "待触发"),
    ("M3", 5, 19, "决议模板 5 + 收尾 3 + 管理 6 + 运维 4 + 报告 3", "2027-Q2 上线期", M3_FILL, "待触发"),
]
r = 4
for p, n0, n1, desc, when, c, do in pri_detail:
    ws.cell(r, 1, p).fill = fill(c); ws.cell(r, 1).font = f(11, True); ws.cell(r, 1).alignment = center()
    ws.cell(r, 2, n0).alignment = center()
    ws.cell(r, 3, n1).alignment = center()
    ws.cell(r, 4, desc).alignment = topleft()
    ws.cell(r, 5, when).alignment = topleft()
    ws.cell(r, 6, do).alignment = center()
    r += 1
grid(ws, 4, r - 1, 1, 6)

# Sheet 5: 行动计划
ws = wb.create_sheet("行动计划")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 6, "B": 32, "C": 12, "D": 32, "E": 50})
title_block(ws, "本次立即执行计划（16 份）", 5)
headers = ["#", "文档名", "优先级", "保存路径", "基于的现有文档"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
plan = [
    (1,  "CATs_系统化计划书",            "P0", "05-其他\\立项材料\\",  "初版构想 + 工作流文档"),
    (2,  "CATs_项目章程",                "P0", "05-其他\\立项材料\\",  "初版构想"),
    (3,  "CATs_要件承認決議書",          "P0", "05-其他\\评审记录\\",  "OFCAT v1.1 + 评审会"),
    (4,  "CATs_BD 评审纪要",             "P0", "05-其他\\评审记录\\",  "架构/技术选型/接口/数据库/模块 + 可热插拔"),
    (5,  "CATs_CI_CD 構築運用手順書",    "P0", "05-其他\\治理\\",     "可热插拔 + Rust 选型 + 架构"),
    (6,  "CATs_開発者ガイド",            "P0", "05-其他\\治理\\",     "Rust 选型 + 架构 + 编码规范"),
    (7,  "CATs_品質ゲート運用手順書",    "P0", "05-其他\\治理\\",     "测试设计 §10 + 编码规范"),
    (8,  "CATs_项目管理计划书",          "P0", "05-其他\\管理\\",     "工作流文档 + QA 登记册"),
    (9,  "CATs_WBS",                     "P0", "05-其他\\管理\\",     "工作流文档 150 任务"),
    (10, "CATs_Baseline 一览",           "P0", "05-其他\\管理\\",     "命名变更 + 全部 CATs 文档"),
    (11, "CATs_安全要件定义书",          "P1", "05-其他\\安全\\",     "架构§14 + 可热插拔§12"),
    (12, "CATs_迁移要件定义书",          "P1", "05-其他\\迁移\\",     "F11 + 可热插拔§11"),
    (13, "CATs_报表设计书",              "P1", "02-基础设计\\报表\\", "F1-F11 + 接口设计"),
    (14, "CATs_SQL 设计一览",            "P1", "03-详细设计\\SQL\\",  "DB 设计书 §4 + DDL"),
    (15, "CATs_UAT 计划书",              "P1", "04-测试\\UAT\\",      "测试设计 §3.4"),
    (16, "CATs_ADR 库",                  "P1", "02-基础设计\\决策\\", "技术选型书 ADR + 架构 ADR"),
]
r = 4
for row in plan:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    pc = pri_color(row[2])
    if pc: ws.cell(r, 3).fill = fill(pc); ws.cell(r, 3).alignment = center(); ws.cell(r, 3).font = f(10, True)
    r += 1
grid(ws, 4, r - 1, 1, 5)
ws.freeze_panes = "A4"

# Sheet 6: 与 150 任务映射
ws = wb.create_sheet("与150任务映射")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 8, "B": 16, "C": 12, "D": 32, "E": 12})
title_block(ws, "应补文档 → 150 任务 ID 反向索引", 5)
headers = ["#", "任务 ID", "优先级", "文档名", "v1.1"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)

m150 = [
    (3,  "P0", "CATs_系统化计划书",       "v1.0"),
    (5,  "P0", "CATs_项目章程",           "v1.0"),
    (6,  "P2", "CATs_AsIs 业务流程图",    "v1.0"),
    (7,  "P2", "CATs_AsIs 系统构成图",    "v1.0"),
    (17, "P1", "CATs_安全要件定义书",     "v1.0"),
    (19, "P1", "CATs_迁移要件定义书",     "v1.0"),
    (20, "P2", "CATs_要件评审纪要",       "v1.0"),
    (21, "P0", "CATs_要件承認決議書",     "v1.0"),
    (26, "P2", "CATs_UI/UX 设计书",       "v1.0"),
    (27, "P1", "CATs_报表设计书",         "v1.0"),
    (33, "P1", "CATs_权限矩阵",           "v1.0"),
    (40, "P1", "CATs_迁移设计书",         "v1.0"),
    (41, "P0", "CATs_BD 评审纪要",        "v1.0"),
    (44, "P1", "CATs_类图",               "v1.0"),
    (45, "P2", "CATs_ロジック設計書",     "v1.0"),
    (48, "P1", "CATs_SQL 设计一览",       "v1.0"),
    (49, "P2", "CATs_批处理详细设计",     "v1.0"),
    (52, "P2", "CATs_DD 评审纪要",        "v1.0"),
    ("53/57/58", "P0", "CATs_CI_CD 構築運用手順書", "v1.1"),
    ("54/56",    "P0", "CATs_開発者ガイド",         "v1.1"),
    ("55/56/128","P0", "CATs_品質ゲート運用手順書",  "v1.1"),
    (80, "M3", "CATs_性能試験レポート",   "v1.1"),
    (83, "M3", "CATs_セキュリティ試験レポート", "v1.1"),
    (84, "M3", "CATs_障害試験レポート",   "v1.1"),
    (90, "P1", "CATs_UAT 计划书",         "v1.0"),
    (94, "M3", "CATs_受入判定書テンプレート", "v1.1"),
    (95, "M3", "CATs_検収書テンプレート", "v1.1"),
    (103,"M3", "CATs_GoNoGo 決議書テンプレート", "v1.1"),
    (107,"M3", "CATs_GoLive 決議書テンプレート", "v1.1"),
    (109,"M3", "CATs_運用マニュアル",     "v1.1"),
    ("114-116","M3", "CATs_インシデント対応プレイブック", "v1.1"),
    (113,"M3", "CATs_キャパシティ管理計画", "v1.1"),
    (117,"M3", "CATs_問い合わせ対応FAQ",  "v1.1"),
    (118,"M3", "CATs_保守マニュアル",     "v1.1"),
    (131,"P0", "CATs_项目管理计划书",     "v1.0"),
    (132,"P0", "CATs_WBS",                "v1.0"),
    (133,"M3", "CATs_進捗報告書テンプレート", "v1.1"),
    (134,"P2", "CATs_课题管理表",         "v1.0"),
    (135,"P2", "CATs_风险登记册",         "v1.1"),
    (137,"P2", "CATs_构成管理台账",       "v1.0"),
    (138,"M3", "CATs_成果物管理台帳",     "v1.1"),
    (140,"P2", "CATs_会议议事录模板",     "v1.1"),
    (141,"M3", "CATs_工数管理表",         "v1.1"),
    (142,"M3", "CATs_コスト管理表",       "v1.1"),
    (143,"M3", "CATs_スコープ変更管理表", "v1.1"),
    (144,"M3", "CATs_ベースライン変更管理表", "v1.1"),
    (145,"M3", "CATs_完了判定書テンプレート", "v1.1"),
    (147,"M3", "CATs_完了報告テンプレート", "v1.1"),
    (148,"M3", "CATs_振り返りテンプレート", "v1.1"),
    (149,"M3", "CATs_KT 記録テンプレート", "v1.1"),
    ("—", "P1", "CATs_ADR 库",            "v1.1"),
    ("76-89","M3", "CATs_系统测试报告",  "v1.0"),
    ("92-95","M3", "CATs_UAT 报告",       "v1.0"),
]
r = 4
for i, (tid, p, name, ver) in enumerate(m150, 1):
    ws.cell(r, 1, i).alignment = center()
    ws.cell(r, 2, str(tid)).alignment = center()
    ws.cell(r, 2).font = f(10, True)
    ws.cell(r, 3, p).alignment = center()
    ws.cell(r, 3).fill = fill(pri_color(p))
    ws.cell(r, 3).font = f(10, True)
    ws.cell(r, 4, name).alignment = topleft()
    if ver == "v1.1":
        ws.cell(r, 5, "v1.1").fill = fill(V11_FILL)
        ws.cell(r, 5).font = f(10, True, "1F3864")
    else:
        ws.cell(r, 5, "v1.0").fill = fill(LIGHT)
    ws.cell(r, 5).alignment = center()
    r += 1
grid(ws, 4, r - 1, 1, 5)
ws.freeze_panes = "A4"

# Sheet 7: v1.1 新增详情
ws = wb.create_sheet("v1.1新增")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 6, "B": 28, "C": 14, "D": 50, "E": 22})
title_block(ws, "v1.1 新增 26 份详解（5 类）", 5)
headers = ["#", "新增类别", "数量", "主要文档", "备注"]
for i, h in enumerate(headers, 1):
    ws.cell(3, i, h)
style_header(ws, 3, 5)
v11 = [
    (1, "A. 实施 / 治理",  5, "CI_CD 構築運用 / 開発者ガイド / 品質ゲート / ADR 库 / 后续整合", "3 份已建，1 待触发，1 整合"),
    (2, "B. 评审 / 决议模板", 5, "受入判定書 / 検収書 / GoNoGo / GoLive / 完了判定書", "M3 待触发"),
    (3, "C. 收尾 / KT / 复盘", 3, "完了報告 / 振り返り / KT 記録", "M3 待触发"),
    (4, "D. 持续管理工具",   8, "风险 / 课题 / 进度 / 工数 / 成本 / 范围 / 基线 / 议事录 / 成果物", "M1 起 + M3"),
    (5, "E. 运维 / Runbook", 4, "運用 / Runbook / 容量 / FAQ", "M3 上线前"),
    (6, "F. ST/UAT 专项",   3, "性能 / 安全 / 障害", "ST 末输出"),
    (7, "G. 保守",          1, "保守マニュアル", "M3 上线后"),
]
r = 4
for row in v11:
    for i, v in enumerate(row, 1):
        cell = ws.cell(r, i, v)
        cell.alignment = topleft()
    ws.cell(r, 2).fill = fill(V11_FILL)
    ws.cell(r, 2).font = f(10, True)
    r += 1
# 合计
ws.cell(r, 1, "合计").font = f(10, True); ws.cell(r, 1).fill = fill(LIGHT); ws.cell(r, 1).alignment = center()
for c, v in enumerate(["", "7 类别", 27, "—", "3 已建 + 24 待触发"], 2):
    cell = ws.cell(r, c, v); cell.font = f(10, True); cell.fill = fill(LIGHT); cell.alignment = topleft()
grid(ws, 4, r, 1, 5)

# Save
_here = os.path.dirname(os.path.abspath(__file__))
out_dir = os.path.dirname(_here)
out = os.path.join(out_dir, "CATs_应补文档清单_v1.1.xlsx")
wb.save(out)
print("Saved", out, "with sheets:", wb.sheetnames)
